import io
import re
import math
import json
import os
import unicodedata
from time import perf_counter
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from copy import deepcopy

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Font


SKIP_STREAMLIT_UI = os.environ.get("FRESH_FROZEN_SKIP_UI") == "1"

_OPENPYXL_MERGED_PATCHED = False


def _patch_openpyxl_merged_cell_writes():
    global _OPENPYXL_MERGED_PATCHED
    if _OPENPYXL_MERGED_PATCHED:
        return

    original_getitem = Worksheet.__getitem__
    original_cell = Worksheet.cell

    def anchor_cell(ws, row, column):
        cell = original_cell(ws, row=row, column=column)
        if not isinstance(cell, MergedCell):
            return cell
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= column <= merged_range.max_col:
                return original_cell(ws, row=merged_range.min_row, column=merged_range.min_col)
        return cell

    def patched_getitem(ws, key):
        cell = original_getitem(ws, key)
        if isinstance(cell, MergedCell):
            return anchor_cell(ws, cell.row, cell.column)
        return cell

    def patched_cell(ws, row, column, value=None):
        cell = anchor_cell(ws, row, column)
        if value is not None:
            cell.value = value
        return cell

    Worksheet.__getitem__ = patched_getitem
    Worksheet.cell = patched_cell
    _OPENPYXL_MERGED_PATCHED = True


_patch_openpyxl_merged_cell_writes()

if not SKIP_STREAMLIT_UI:
    st.set_page_config(page_title="鲜冻品占比通用模型", layout="wide")
    st.title("鲜冻品占比通用模型")
    st.caption("上传月系统成本与Q系统成本，自动提取并导出分表结果，表格与明细分开。")


APP_DIR = Path(__file__).resolve().parent
MONITOR_DIR = APP_DIR / "monitor_logs"
MONITOR_DIR.mkdir(parents=True, exist_ok=True)
RUN_LOG_FILE = MONITOR_DIR / "fresh_frozen_run_metrics.jsonl"
DEFAULT_TEMPLATE_CANDIDATES = [
    APP_DIR / "2601系统成本" / "鲜冻品占比-1.xlsx",
    APP_DIR / "鲜冻品占比-1.xlsx",
]
_WORKBOOK_CACHE = {}
_EXTRACT_CACHE = {}


def append_run_log(payload: dict):
    with RUN_LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def resolve_target_template_path(month_label: str) -> Path:
    digits = re.sub(r"\D", "", str(month_label or ""))
    if not digits:
        raise FileNotFoundError("未能识别月份，无法定位目标模板文件。")
    month_num = int(digits)
    candidates = [
        APP_DIR / f"26{month_num:02d}系统成本" / f"鲜冻品占比-{month_num}.xlsx",
        APP_DIR / f"鲜冻品占比-{month_num}.xlsx",
        *DEFAULT_TEMPLATE_CANDIDATES,
    ]
    for path in candidates:
        if path.exists():
            return path
    candidates_text = "、".join(str(path) for path in candidates)
    raise FileNotFoundError(f"未找到目标模板文件，请检查：{candidates_text}")


PLANT_CODE_MAP = {
    "蚌埠一厂": "BB1",
    "蚌埠二厂": "BB2",
    "辽阳": "LY",
    "天津": "TJ",
    "兖州": "YZ",
    "大连": "DL",
}


def normalize_plant_code(plant_name: str) -> str:
    return PLANT_CODE_MAP.get(str(plant_name).strip(), str(plant_name).strip())


MARKET_PRICE_TEMPLATE_SHEET = "行情价覆盖"
MARKET_PRICE_TEMPLATE_COLUMNS = ["工厂", "分类", "当前行情价", "基期行情价", "备注"]


def _market_price_template_df():
    plants = sorted(set(PLANT_CODE_MAP.values()))
    rows = []
    for plant in plants:
        for kind in ("腿肉", "胸肉", "其他"):
            rows.append({"工厂": plant, "分类": kind, "当前行情价": None, "基期行情价": None, "备注": ""})
    return pd.DataFrame(rows, columns=MARKET_PRICE_TEMPLATE_COLUMNS)


def _build_market_price_template_bytes():
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df = _market_price_template_df()
        df.to_excel(writer, index=False, sheet_name=MARKET_PRICE_TEMPLATE_SHEET)
        ws = writer.book[MARKET_PRICE_TEMPLATE_SHEET]
        widths = {"A": 10, "B": 10, "C": 18, "D": 18, "E": 24}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    out.seek(0)
    return out.getvalue()


def to_num(x):
    try:
        if x is None or pd.isna(x):
            return None
        s = str(x).strip()
        if s == "" or s == "-":
            return None
        s = s.replace(",", "").replace("，", "")
        # 支持 (123) 负数格式
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        v = float(s)
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except Exception:
        return None


def norm_code(x):
    if x is None or pd.isna(x):
        return ""
    s = str(x).strip()
    if s.lower() == "nan":
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s


def month_from_name(name: str) -> str:
    base = name.rsplit(".", 1)[0]
    m = re.search(r"(20)?(\d{2})(\d{2})(?:\s*\(\d+\))?$", base)
    if m:
        return str(int(m.group(3)))
    m2 = re.search(r"_(\d{4})$", base)
    if m2:
        return str(int(m2.group(1)[-2:]))
    return "11"


def quarter_from_name(name: str) -> str:
    base = name.rsplit(".", 1)[0].upper()
    if "Q2-Q4" in base or "25Q2-Q4" in base:
        return "Q2"

    m = re.search(r"Q(\d+)", base, flags=re.IGNORECASE)
    if not m:
        return "Q3"

    quarter_no = int(m.group(1))
    return f"Q{quarter_no}"


def quarter_from_files(files) -> str:
    counts = defaultdict(int)
    order = []
    for f in files or []:
        label = quarter_from_name(getattr(f, "name", str(f)))
        counts[label] += 1
        if label not in order:
            order.append(label)
    if not order:
        return "Q3"
    return max(order, key=lambda label: (counts[label], -order.index(label)))


def plant_from_name(name: str) -> str:
    base = name.rsplit(".", 1)[0]
    m = re.match(r"(.+?)系统成本", base)
    if m:
        return m.group(1).strip()
    if "_" in base:
        return base.split("_", 1)[0].strip()
    return base


def _norm_col(c):
    return str(c).strip()


def _score_tsc_columns(cols):
    colset = {_norm_col(c) for c in cols}
    must = {"产品族", "修行后原料", "使用半成品规格", "行类型", "影响口径"}
    strong = {"综合单价", "修形前原料综合耗用单价", "修形利用率", "损耗率", "半成品修形人工成本", "半成品总成本", "半成品入库量"}
    score = 0
    score += 6 * len(must & colset)
    score += 3 * len(strong & colset)
    if any(_norm_col(c).startswith("310") for c in cols):
        score += 2
    return score


def _score_part_columns(cols):
    colset = {_norm_col(c) for c in cols}
    keys = {"物料号", "原料号", "原料描述", "实际数量", "调整后实际量", "辅助"}
    score = 0
    score += 4 * len(keys & colset)
    if "分类" in colset:
        score += 1
    return score


def read_sheet_safe(xls: pd.ExcelFile, sheet_name: str, score_fn=None):
    best_df = None
    best_score = -1
    for h in (0, 1, None):
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=h)
        except Exception:
            continue
        score = score_fn(df.columns) if score_fn else 0
        if score > best_score:
            best_df = df
            best_score = score
    return best_df


def read_tsc_df(xls: pd.ExcelFile, kind: str):
    target = f"{kind}TSC"
    for sn in xls.sheet_names:
        if sn == target:
            return read_sheet_safe(xls, sn, score_fn=_score_tsc_columns)
    return None


def read_part_df(xls: pd.ExcelFile, kind: str):
    for sn in xls.sheet_names:
        if sn.endswith(kind) and not sn.endswith("TSC"):
            return read_sheet_safe(xls, sn, score_fn=_score_part_columns)
    return None


def find_col(cols, name, fallback=None):
    for c in cols:
        if str(c).strip() == name:
            return c
    if fallback is not None and 0 <= fallback < len(cols):
        return cols[fallback]
    return None


def parse_tsc(tsc_df: pd.DataFrame):
    def pick_col(preferred_names, fallback=None):
        for n in preferred_names:
            c = find_col(cols, n)
            if c is not None:
                return c
        if fallback is not None and 0 <= fallback < len(cols):
            return cols[fallback]
        return None

    def detect_col_by_tokens(candidates, tokens):
        best_col, best_score = None, 0
        sample_n = min(120, len(tsc_df))
        for c in candidates:
            score = 0
            for i in range(sample_n):
                v = tsc_df.iloc[i][c]
                if pd.isna(v):
                    continue
                s = str(v).strip()
                if any(t in s for t in tokens):
                    score += 1
            if score > best_score:
                best_col, best_score = c, score
        return best_col if best_score > 0 else None

    cols = list(tsc_df.columns)
    if len(cols) < 5:
        return [], {}

    c_prod = pick_col(["产品族"], 0)
    c_mat = pick_col(["修行后原料", "修形后原料"], 1)
    c_spec = pick_col(["使用半成品规格"], 2)
    c_type = pick_col(["行类型"], None)
    c_scope = pick_col(["影响口径"], None)

    # 部分Q表会把“行类型/影响口径”标题吞掉，改为按内容识别列
    if c_type is None:
        c_type = detect_col_by_tokens(cols[: min(10, len(cols))], ["实际单价", "实际价格", "规格占比", "差异", "对半成品成本的影响"])
    if c_scope is None:
        c_scope = detect_col_by_tokens(cols[: min(10, len(cols))], ["总成本", "单位成本"])

    if c_type is None:
        c_type = cols[3]
    if c_scope is None and c_type in cols:
        idx = cols.index(c_type)
        if idx + 1 < len(cols):
            c_scope = cols[idx + 1]
    if c_scope is None:
        c_scope = cols[4]
    c_mix = find_col(cols, "综合单价", 8)
    c_pre = find_col(cols, "修形前原料综合耗用单价", 9)
    c_util = find_col(cols, "修形利用率", 10)
    c_loss = find_col(cols, "损耗率", 11)
    c_lab = find_col(cols, "半成品修形人工成本", 13)
    c_total = find_col(cols, "半成品总成本", 14)
    c_in = find_col(cols, "半成品入库量", 15)

    raw_cols = cols[5:8] if len(cols) >= 8 else []
    if c_mix is not None and c_mix in cols:
        idx = cols.index(c_mix)
        raw_cols = cols[5:idx] if idx > 5 else []

    code_to_spec = {}
    for rc in raw_cols:
        v = None
        for i in range(min(10, len(tsc_df))):
            v = tsc_df.iloc[i][rc]
            if pd.notna(v) and str(v).strip() != "":
                break
        code_to_spec[norm_code(rc)] = str(v).strip() if pd.notna(v) else ""

    rows = []
    cur_prod, cur_mat, cur_spec = "", "", ""
    for _, r in tsc_df.iterrows():
        prod = r.get(c_prod)
        mat = norm_code(r.get(c_mat))
        spec = r.get(c_spec)
        if pd.notna(prod) and str(prod).strip() != "":
            cur_prod = str(prod).strip()
        if mat:
            cur_mat = mat
        if pd.notna(spec) and str(spec).strip() != "":
            cur_spec = str(spec).strip()

        rtype = "" if pd.isna(r.get(c_type)) else str(r.get(c_type)).strip()
        scope = "" if pd.isna(r.get(c_scope)) else str(r.get(c_scope)).strip()
        if not cur_mat or not rtype:
            continue

        raw_share = {}
        for rc in raw_cols:
            val = to_num(r.get(rc))
            if val is not None:
                raw_share[norm_code(rc)] = val

        rows.append(
            {
                "产品族": cur_prod,
                "修行后原料": cur_mat,
                "使用半成品规格": cur_spec,
                "行类型": rtype,
                "影响口径": scope,
                "综合单价": to_num(r.get(c_mix)) if c_mix is not None else None,
                "修形前原料综合耗用单价": to_num(r.get(c_pre)) if c_pre is not None else None,
                "修形利用率": to_num(r.get(c_util)) if c_util is not None else None,
                "损耗率": to_num(r.get(c_loss)) if c_loss is not None else None,
                "半成品修形人工成本": to_num(r.get(c_lab)) if c_lab is not None else None,
                "半成品总成本": to_num(r.get(c_total)) if c_total is not None else None,
                "半成品入库量": to_num(r.get(c_in)) if c_in is not None else None,
                "raw_share": raw_share,
            }
        )
    return rows, code_to_spec


def parse_part(part_df: pd.DataFrame):
    if part_df is None or part_df.empty:
        return []
    cols = list(part_df.columns)
    # 优先按列名识别，失败再回退位置
    def pick_col(*names, fallback=None):
        for n in names:
            for c in cols:
                if str(c).strip() == n:
                    return c
        if fallback is not None and 0 <= fallback < len(cols):
            return cols[fallback]
        return None

    c_mat = pick_col("物料号", "物料号　　", fallback=1)
    c_raw_code = pick_col("原料号", "原料号　　", fallback=4)
    c_raw_desc = pick_col("原料描述", "原料描述　", fallback=5)
    c_qty = pick_col("实际数量", "调整后实际量", fallback=6)
    c_aux = pick_col("辅助", fallback=9)

    out = []
    for _, row in part_df.iterrows():
        mat = norm_code(row.get(c_mat))
        raw_code = norm_code(row.get(c_raw_code))
        raw_desc = "" if c_raw_desc is None or pd.isna(row.get(c_raw_desc)) else str(row.get(c_raw_desc)).strip()
        qty = to_num(row.get(c_qty)) if c_qty is not None else None
        aux = to_num(row.get(c_aux)) if c_aux is not None else None
        if not mat or not raw_code or qty is None or qty <= 0:
            continue
        fresh_frozen = "冻品" if "冻品" in raw_desc else ("鲜品" if "鲜品" in raw_desc else "")
        out.append(
            {
                "修行后原料": mat,
                "原料号": raw_code,
                "原料描述": raw_desc,
                "数量": qty,
                "鲜冻": fresh_frozen,
                "辅助": aux,
            }
        )
    return out


def _row_type_text(row_or_text) -> str:
    if isinstance(row_or_text, dict):
        row_or_text = row_or_text.get("行类型", "")
    return "" if row_or_text is None else str(row_or_text).strip()


def _is_actual_price_row_type(row_or_text) -> bool:
    text = _row_type_text(row_or_text)
    return "实际单价" in text or "实际价格" in text


def _is_current_actual_price_row_type(row_or_text) -> bool:
    text = _row_type_text(row_or_text)
    return any(token in text for token in ("月实际单价", "月实际价格", "本月实际单价", "本月实际价格"))


def _actual_price_row_score(row):
    if not isinstance(row, dict):
        return 0
    score = 0
    for k in ("半成品总成本", "修形前原料综合耗用单价", "综合单价", "半成品入库量"):
        if to_num(row.get(k)) is not None:
            score += 1
    return score


def _baseline_actual_rank(row_or_text, quarter_label: str | None = None):
    text = _row_type_text(row_or_text).replace(" ", "")
    if "规格占比" in text or not _is_actual_price_row_type(text):
        return None
    quarter_text = str(quarter_label or "").strip().upper()
    text_upper = text.upper()
    if quarter_text and quarter_text in text_upper:
        return 0
    if re.search(r"Q[1-4]", text_upper):
        return 1
    if any(token in text for token in ("25年实际单价", "25年实际价格", "25实际单价", "25实际价格")):
        return 2
    return None


def first_match(rows_by_type: dict, include_text: str, exclude_text: str = ""):
    def as_list(v):
        return v if isinstance(v, list) else [v]

    def score_row(row):
        return _actual_price_row_score(row)

    for k, v in rows_by_type.items():
        if include_text in k and (exclude_text == "" or exclude_text not in k):
            rows = [r for r in as_list(v) if isinstance(r, dict)]
            if not rows:
                continue
            rows.sort(key=score_row, reverse=True)
            return rows[0]
    return None


def first_match_any(rows_by_type: dict, include_texts, exclude_texts=()):
    include_list = [str(t) for t in (include_texts or []) if str(t)]
    exclude_list = [str(t) for t in (exclude_texts or []) if str(t)]

    def as_list(v):
        return v if isinstance(v, list) else [v]

    def score_row(row):
        return _actual_price_row_score(row)

    for k, v in rows_by_type.items():
        if include_list and not any(text in k for text in include_list):
            continue
        if exclude_list and any(text in k for text in exclude_list):
            continue
        rows = [r for r in as_list(v) if isinstance(r, dict)]
        if not rows:
            continue
        rows.sort(key=score_row, reverse=True)
        return rows[0]
    return None


def _first_actual_price_match(rows_by_type: dict | None, must_include=(), include_any=(), exclude_texts=()):
    if not rows_by_type:
        return None
    must_list = [str(t) for t in (must_include or []) if str(t)]
    include_list = [str(t) for t in (include_any or []) if str(t)]
    exclude_list = [str(t) for t in (exclude_texts or []) if str(t)]

    def as_list(v):
        return v if isinstance(v, list) else [v]

    def score_row(row):
        return _actual_price_row_score(row)

    for k, v in rows_by_type.items():
        key = str(k)
        if must_list and not all(text in key for text in must_list):
            continue
        if include_list and not any(text in key for text in include_list):
            continue
        if exclude_list and any(text in key for text in exclude_list):
            continue
        if "实际单价" not in key and "实际价格" not in key:
            continue
        rows = [r for r in as_list(v) if isinstance(r, dict)]
        if not rows:
            continue
        rows.sort(key=score_row, reverse=True)
        return rows[0]
    return None


def _select_baseline_actual_row(sources, quarter_label: str):
    source_list = [source for source in (sources or []) if source]
    candidates = []

    def as_list(v):
        return v if isinstance(v, list) else [v]

    for source_idx, source in enumerate(source_list):
        for key_idx, (key, value) in enumerate(source.items()):
            rank = _baseline_actual_rank(key, quarter_label)
            if rank is None:
                continue
            rows = [r for r in as_list(value) if isinstance(r, dict)]
            for row_idx, row in enumerate(rows):
                candidates.append(
                    (
                        rank,
                        source_idx,
                        key_idx,
                        -_actual_price_row_score(row),
                        row_idx,
                        row,
                    )
                )
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[:-1])
    return candidates[0][-1]


def match_impact_row(rows_by_type: dict, prefer_scope: str = ""):
    def as_list(v):
        return v if isinstance(v, list) else [v]

    def pick(rows):
        if not rows:
            return None
        if prefer_scope:
            scoped = [r for r in rows if str(r.get("影响口径", "")).strip() == prefer_scope]
            if scoped:
                rows = scoped
        rows = [r for r in rows if isinstance(r, dict)]
        if not rows:
            return None
        rows.sort(
            key=lambda r: (
                to_num(r.get("半成品总成本")) is not None,
                to_num(r.get("修形前原料综合耗用单价")) is not None,
            ),
            reverse=True,
        )
        return rows[0]

    for k, v in rows_by_type.items():
        if "对半成品成本的影响" in k:
            row = pick(as_list(v))
            if row is not None:
                return row
    return None


def select_material_rows(md: dict, qd: dict, quarter_label: str):
    mrow = first_match_any(md, ("月实际单价", "月实际价格", "本月实际单价", "本月实际价格"), ("Q", "25"))
    qrow = _select_baseline_actual_row([qd, md], quarter_label)
    diff = first_match(md, "差异", "")
    return mrow, qrow, diff


def select_actual_rows(md: dict, qd: dict, quarter_label: str):
    current = first_match_any(md, ("月实际单价", "月实际价格", "本月实际单价", "本月实际价格"), ("Q", "25"))
    if current is None:
        current = first_match_any(md, ("实际单价", "实际价格"), ("Q", "25"))
    if current is None:
        current = first_match_any(md, ("实际单价", "实际价格"))

    previous = _select_baseline_actual_row([md, qd], quarter_label)
    return current, previous


def _resolve_effective_q_qty(qrow: dict | None, impact_row: dict | None):
    q_qty = to_num((qrow or {}).get("半成品入库量"))
    if q_qty in (None, 0):
        q_qty = to_num((impact_row or {}).get("半成品入库量"))
    return q_qty or 0.0


def _is_zero_market_suppressed_impact_row(row: dict | None):
    if not row:
        return False

    price = to_num(row.get("修形前原料综合耗用单价"))
    util = to_num(row.get("修形利用率"))
    loss = to_num(row.get("损耗率"))
    raw = to_num(row.get("半成品原料成本"))
    lab = to_num(row.get("半成品修形人工成本"))
    total = to_num(row.get("半成品总成本"))

    if raw is None and all(v is not None for v in (price, util, loss)):
        raw = price + util + loss
    if total is None and raw is not None and lab is not None:
        total = raw + lab

    normalized_values = tuple(0.0 if v is None else float(v) for v in (price, util, loss, raw, lab, total))
    return all(abs(v) < 1e-12 for v in normalized_values)

def _actual_metric_values(row: dict | None, kind: str):
    if not row:
        return {}

    def pick(primary_key: str, fallback_key: str):
        value = to_num(row.get(primary_key))
        if value is None:
            value = to_num(row.get(fallback_key))
        return value

    pre = pick("修形前原料综合耗用单价", "pre")
    util = pick("修形利用率", "util")
    loss = pick("损耗率", "loss")
    raw = pick("半成品原料成本", "raw")
    lab = pick("半成品修形人工成本", "lab")
    qty = pick("半成品入库量", "qty")
    if raw is None:
        raw = _calc_raw_cost(kind, pre, util, loss)
    return {
        "pre": pre,
        "util": util,
        "loss": loss,
        "raw": raw,
        "lab": lab,
        "qty": qty,
    }


def _has_total_metrics(row: dict | None):
    if not row:
        return False
    metric_keys = [
        "修形前原料综合耗用单价",
        "修形利用率",
        "损耗率",
        "半成品修形人工成本",
        "半成品总成本",
    ]
    return any(to_num(row.get(key)) is not None for key in metric_keys)


def _build_blank_total_impact_row(base_row: dict | None, qty=None):
    base_row = base_row or {}
    return {
        "产品族": base_row.get("产品族", ""),
        "修行后原料": base_row.get("修行后原料", ""),
        "使用半成品规格": base_row.get("使用半成品规格", ""),
        "行类型": "对半成品成本的影响",
        "影响口径": "总成本",
        "综合单价": None,
        "修形前原料综合耗用单价": None,
        "修形利用率": None,
        "损耗率": None,
        "半成品原料成本": None,
        "半成品修形人工成本": None,
        "半成品总成本": None,
        "半成品入库量": qty,
        "raw_share": {},
    }


def _scale_unit_impact_row(unit_row: dict, qty, base_row: dict | None = None):
    base_row = base_row or unit_row or {}
    scaled = _build_blank_total_impact_row(base_row, qty=qty)
    if qty in (None, 0):
        return scaled
    metric_keys = [
        "修形前原料综合耗用单价",
        "修形利用率",
        "损耗率",
        "半成品修形人工成本",
        "半成品总成本",
    ]
    for key in metric_keys:
        value = to_num((unit_row or {}).get(key))
        if value is not None:
            scaled[key] = value * qty
    return scaled


def resolve_material_impact(
    kind: str,
    md: dict,
    qd: dict,
    quarter_label: str,
    market_map: dict | None = None,
    fallback_actual_map: dict | None = None,
):
    market_map = market_map or {}
    fallback_actual_map = fallback_actual_map or {}

    impact = None
    if "对半成品成本的影响" in md:
        impact_rows = md.get("对半成品成本的影响") or []
        impact = match_impact_row({"对半成品成本的影响": impact_rows}, prefer_scope="总成本")
    if impact is None:
        impact = match_impact_row(md, prefer_scope="总成本")
    if impact is None and "对半成品成本的影响" in qd:
        impact_rows = qd.get("对半成品成本的影响") or []
        impact = match_impact_row({"对半成品成本的影响": impact_rows}, prefer_scope="总成本")
    if impact is None:
        impact = match_impact_row(qd, prefer_scope="总成本")
    if impact is not None and str(impact.get("影响口径", "")).strip() == "总成本":
        if _has_total_metrics(impact):
            return impact
        # Some Q files contain a blank placeholder total-impact row. Do not let it
        # suppress fallback calculation from actual rows.
        impact = None

    mrow, qrow, diff = select_material_rows(md, qd, quarter_label)
    current_actual, previous_actual = select_actual_rows(md, qd, quarter_label)
    base_row = current_actual or mrow or qrow or diff or previous_actual or {}
    if not base_row:
        return None

    qty = to_num((current_actual or {}).get("半成品入库量"))
    if qty is None:
        qty = to_num((mrow or {}).get("半成品入库量"))

    unit_impact = None
    if "对半成品成本的影响" in md:
        unit_impact = match_impact_row({"对半成品成本的影响": md.get("对半成品成本的影响") or []}, prefer_scope="单位成本")
    if unit_impact is None:
        unit_impact = match_impact_row(md, prefer_scope="单位成本")
    if unit_impact is None and "对半成品成本的影响" in qd:
        unit_impact = match_impact_row({"对半成品成本的影响": qd.get("对半成品成本的影响") or []}, prefer_scope="单位成本")
    if unit_impact is None:
        unit_impact = match_impact_row(qd, prefer_scope="单位成本")
    if unit_impact is not None and qty not in (None, 0):
        return _scale_unit_impact_row(unit_impact, qty, base_row=base_row)

    if current_actual and previous_actual:
        fallback_impact = _fallback_total_impact_from_actual_rows(kind, current_actual, previous_actual, qty=qty)
        if fallback_impact is not None:
            return fallback_impact

    impact = _build_blank_total_impact_row(base_row, qty=qty)
    mat = impact.get("修行后原料") or base_row.get("修行后原料")
    fallback_actual = (fallback_actual_map.get(mat) or {}) if mat else {}

    current_vals = _actual_metric_values(current_actual or mrow, kind)
    previous_vals = _actual_metric_values(previous_actual, kind)
    for key in ("util", "loss", "raw", "lab"):
        if to_num(previous_vals.get(key)) is None and to_num(fallback_actual.get(key)) is not None:
            previous_vals[key] = fallback_actual.get(key)

    previous_pre = to_num((previous_actual or {}).get("修形前原料综合耗用单价"))
    if previous_pre is None:
        previous_pre = to_num(fallback_actual.get("pre"))
    if previous_pre is None:
        previous_pre = to_num((market_map.get(mat) or {}).get("previous_pre"))

    impact = _fill_missing_total_metrics(impact, kind, current_vals, previous_vals, previous_pre)
    return impact if _has_total_metrics(impact) or base_row else None

def _validate_part_material_coverage(records, quarter_label, kind):
    issues = []
    for rec in records:
        plant = rec.get("plant", "")
        rec_q_label = rec.get("q_label") or quarter_label
        month_rows = rec.get("month_tsc_rows") or []
        q_rows = rec.get("q_tsc_rows") or []
        month_parts = rec.get("month_part_rows") or []
        q_parts = rec.get("q_part_rows") or []
        month_part_mats = {norm_code(part.get("修行后原料")) for part in month_parts if norm_code(part.get("修行后原料"))}
        q_part_mats = {norm_code(part.get("修行后原料")) for part in q_parts if norm_code(part.get("修行后原料"))}

        m_grp = defaultdict(lambda: defaultdict(list))
        q_grp = defaultdict(lambda: defaultdict(list))
        for row in month_rows:
            m_grp[row["修行后原料"]][row["行类型"]].append(row)
        for row in q_rows:
            q_grp[row["修行后原料"]][row["行类型"]].append(row)

        market_map = rec.get("market_impact_map") or {}
        fallback_actual_map = rec.get("fallback_actual_map") or {}
        allowed_mats = rec.get("allowed_mats")
        missing_month = []
        missing_q = []
        excluded_mats = set()
        for mat in sorted(set(m_grp.keys()) | set(q_grp.keys())):
            if allowed_mats is not None and mat not in allowed_mats:
                continue
            md = m_grp.get(mat, {})
            qd = q_grp.get(mat, {})
            impact = resolve_material_impact(
                kind,
                md,
                qd,
                rec_q_label,
                market_map=market_map,
                fallback_actual_map=fallback_actual_map,
            )
            if impact is None or impact.get("影响口径") != "总成本":
                continue

            current_actual, previous_actual = select_actual_rows(md, qd, rec_q_label)
            mrow, qrow, _ = select_material_rows(md, qd, rec_q_label)
            month_qty = to_num((current_actual or mrow or {}).get("半成品入库量")) or 0.0
            q_qty = to_num((previous_actual or qrow or {}).get("半成品入库量")) or 0.0

            missing_month_flag = month_qty > 0 and mat not in month_part_mats
            missing_q_flag = q_qty > 0 and mat not in q_part_mats
            if missing_month_flag:
                missing_month.append(mat)
            if missing_q_flag:
                missing_q.append(mat)
            if missing_month_flag or missing_q_flag:
                excluded_mats.add(mat)

        if excluded_mats:
            if allowed_mats is None:
                allowed_mats = set()
            rec["allowed_mats"] = set(allowed_mats) - excluded_mats

        if missing_month or missing_q:
            issues.append(
                {
                    "plant": plant,
                    "month": missing_month,
                    "quarter": missing_q,
                }
            )

    return None


def _validate_tsc_material_overlap(records, quarter_label, kind, min_overlap_ratio: float = 0.6):
    issues = []

    def _material_set(rows):
        mats = set()
        for row in rows or []:
            mat = norm_code(row.get("修行后原料"))
            if not mat or mat == "0":
                continue
            if not re.fullmatch(r"\d+", mat):
                continue
            mats.add(mat)
        return mats

    for rec in records:
        plant = rec.get("plant", "")
        month_rows = rec.get("source_month_tsc_rows") or rec.get("month_tsc_rows") or []
        q_rows = rec.get("q_tsc_rows") or []
        month_mats = _material_set(month_rows)
        q_mats = _material_set(q_rows)
        if not month_mats or not q_mats:
            continue
        overlap = month_mats & q_mats
        ratio = len(overlap) / min(len(month_mats), len(q_mats))
        if ratio >= min_overlap_ratio:
            continue
        issues.append(
            {
                "plant": plant,
                "ratio": ratio,
                "month_only": sorted(month_mats - q_mats),
                "quarter_only": sorted(q_mats - month_mats),
            }
        )

    if not issues:
        return

    parts = []
    for item in issues:
        month_only = ', '.join(item['month_only'][:5]) or '-'
        quarter_only = ', '.join(item['quarter_only'][:5]) or '-'
        parts.append(
            f"{item['plant']}：月/Q半成品重合度 {item['ratio']:.0%}，月独有[{month_only}]，Q独有[{quarter_only}]"
        )
    detail = ' | '.join(parts[:8])
    return f"{kind} 月文件与Q文件的半成品结构存在差异。{detail}。将按当前上传数据继续计算。"


def build_kind(records, month_label, quarter_label, kind, material_spec_profile=None):
    s1_rows = []
    m_fs = defaultdict(lambda: {"冻品": 0.0, "鲜品": 0.0})
    q_fs = defaultdict(lambda: {"冻品": 0.0, "鲜品": 0.0})
    m_spec = defaultdict(lambda: defaultdict(float))
    q_spec = defaultdict(lambda: defaultdict(float))
    spec_set = set()
    spec_group_fn = _spec_group_fn_for_kind(kind)

    for rec in records:
        plant = rec["plant"]
        month_rows = rec["month_tsc_rows"]
        q_rows = rec["q_tsc_rows"]
        rec_q_label = rec.get("q_label") or quarter_label
        code_to_spec = rec["code_to_spec"]
        month_parts = rec["month_part_rows"]
        q_parts = rec["q_part_rows"]

        m_grp = defaultdict(lambda: defaultdict(list))
        q_grp = defaultdict(lambda: defaultdict(list))
        for r in month_rows:
            m_grp[r["修行后原料"]][r["行类型"]].append(r)
        for r in q_rows:
            q_grp[r["修行后原料"]][r["行类型"]].append(r)
        market_map = rec.get("market_impact_map") or {}
        fallback_actual_map = rec.get("fallback_actual_map") or {}
        allowed_mats = rec.get("allowed_mats")

        agg = {
            "工厂": plant,
            "原料采购单价影响": 0.0,
            "行情影响": 0.0,
            "扣除行情后采购绩效": 0.0,
            "修形利用率影响": 0.0,
            "损耗率影响": 0.0,
            "修形人工成本影响": 0.0,
            "综合影响": 0.0,
            f"{month_label}月产量": 0.0,
            f"{quarter_label}月均产量": 0.0,
        }

        mats = sorted(set(m_grp.keys()) | set(q_grp.keys()))
        valid_mats = set()
        for mat in mats:
            if not _looks_like_semifinished_code(mat):
                continue
            if allowed_mats is not None and mat not in allowed_mats:
                continue
            md = m_grp.get(mat, {})
            qd = q_grp.get(mat, {})

            impact = resolve_material_impact(
                kind,
                md,
                qd,
                rec_q_label,
                market_map=market_map,
                fallback_actual_map=fallback_actual_map,
            )
            if impact is None or impact.get("影响口径") != "总成本":
                continue

            mrow, qrow, diff = select_material_rows(md, qd, rec_q_label)

            m_qty = to_num((mrow or {}).get("半成品入库量")) or 0.0
            q_qty = _resolve_effective_q_qty(qrow, impact)
            if m_qty <= 0:
                continue

            valid_mats.add(mat)
            price_impact = to_num(impact.get("修形前原料综合耗用单价")) or 0.0
            market_gap = _calc_market_gap(kind, price_impact, market_map, mat, m_qty, q_qty, impact)

            agg["原料采购单价影响"] += price_impact
            agg["行情影响"] += market_gap
            agg["修形利用率影响"] += to_num(impact.get("修形利用率")) or 0.0
            agg["损耗率影响"] += to_num(impact.get("损耗率")) or 0.0
            agg["修形人工成本影响"] += to_num(impact.get("半成品修形人工成本")) or 0.0
            agg[f"{month_label}月产量"] += m_qty
            agg[f"{quarter_label}月均产量"] += q_qty

        for p in month_parts:
            t, q = p["鲜冻"], p["数量"]
            if t in ("冻品", "鲜品"):
                m_fs[plant][t] += q
            raw_code = norm_code(p.get("原料号"))
            spec_text = code_to_spec.get(raw_code, "")
            raw_desc = str(p.get("原料描述") or "")
            spec = _resolve_spec_bucket(kind, raw_code, spec_text, raw_desc, spec_group_fn, material_spec_profile)
            if spec:
                spec_set.add(spec)
                m_spec[plant][spec] += q

        for p in q_parts:
            t, q = p["鲜冻"], p["数量"]
            if t in ("冻品", "鲜品"):
                q_fs[plant][t] += q
            raw_code = norm_code(p.get("原料号"))
            spec_text = code_to_spec.get(raw_code, "")
            raw_desc = str(p.get("原料描述") or "")
            spec = _resolve_spec_bucket(kind, raw_code, spec_text, raw_desc, spec_group_fn, material_spec_profile)
            if spec:
                spec_set.add(spec)
                q_spec[plant][spec] += q

        # Q月均产量兜底：来自Q明细“辅助”或数量汇总（按千单位）。
        # 仅统计已纳入影响计算的半成品物料，避免被无关物料放大。
        q_qty_fallback = 0.0
        if q_parts and valid_mats:
            aux_by_mat = defaultdict(float)
            qty_sum = 0.0
            for p in q_parts:
                mat = p.get("修行后原料")
                if mat not in valid_mats:
                    continue
                if p.get("辅助") is not None:
                    aux_by_mat[mat] = max(aux_by_mat[mat], float(p["辅助"]))
                qty_sum += float(p["数量"])
            q_qty_fallback = (sum(aux_by_mat.values()) / 1000.0) if aux_by_mat else (qty_sum / 1000.0)

        if agg[f"{quarter_label}月均产量"] == 0 and q_qty_fallback > 0:
            agg[f"{quarter_label}月均产量"] = q_qty_fallback

        s1_rows.append(agg)

    s1 = pd.DataFrame(s1_rows)
    if not s1.empty:
        s1["扣除行情后采购绩效"] = s1["原料采购单价影响"] - s1["行情影响"]
        s1["综合影响"] = (
            s1["扣除行情后采购绩效"]
            + s1["修形利用率影响"]
            + s1["损耗率影响"]
            + s1["修形人工成本影响"]
        )
        # 对齐目标模板口径：不展示完全为0的工厂行（避免空工厂干扰展示与汇总版式）
        value_cols = [
            "原料采购单价影响",
            "行情影响",
            "扣除行情后采购绩效",
            "修形利用率影响",
            "损耗率影响",
            "修形人工成本影响",
            "综合影响",
            f"{month_label}月产量",
            f"{quarter_label}月均产量",
        ]
        numeric_view = s1[value_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
        non_zero_mask = numeric_view.abs().sum(axis=1) > 0
        if non_zero_mask.any():
            s1 = s1.loc[non_zero_mask].reset_index(drop=True)

        ordered_cols = [
            "工厂",
            "原料采购单价影响",
            "行情影响",
            "扣除行情后采购绩效",
            "修形利用率影响",
            "损耗率影响",
            "修形人工成本影响",
            "综合影响",
            f"{month_label}月产量",
            f"{quarter_label}月均产量",
        ]
        s1 = s1[[c for c in ordered_cols if c in s1.columns]]
        total = {c: (s1[c].sum() if c != "工厂" else "合计") for c in s1.columns}
        s1 = pd.concat([s1, pd.DataFrame([total])], ignore_index=True)
        base = s1.loc[s1["工厂"] == "合计", "综合影响"].iloc[0] if "合计" in s1["工厂"].values else 1.0
        base = base if base not in (0, None) else 1.0
        impact = {"工厂": "影响"}
        impact["原料采购单价影响"] = None
        impact["行情影响"] = None
        for c in ["扣除行情后采购绩效", "修形利用率影响", "损耗率影响", "修形人工成本影响", "综合影响"]:
            v = s1.loc[s1["工厂"] == "合计", c].iloc[0] if "合计" in s1["工厂"].values else 0.0
            impact[c] = v / base
        for c in [f"{month_label}月产量", f"{quarter_label}月均产量"]:
            impact[c] = None
        s1 = pd.concat([s1, pd.DataFrame([impact])], ignore_index=True)

    plants = sorted(set(list(m_fs.keys()) + list(q_fs.keys())))
    s2_rows = []
    for p in plants:
        mf, mx = m_fs[p]["冻品"], m_fs[p]["鲜品"]
        qf, qx = q_fs[p]["冻品"], q_fs[p]["鲜品"]
        mt, qt = mf + mx, qf + qx
        s2_rows += [
            {"工厂": p, "月份": f"{month_label}月", "冻品": mf / mt if mt else None, "鲜品": mx / mt if mt else None, "合计": 1 if mt else None},
            {"工厂": "", "月份": quarter_label, "冻品": qf / qt if qt else None, "鲜品": qx / qt if qt else None, "合计": 1 if qt else None},
            {"工厂": "", "月份": "差异", "冻品": (mf / mt if mt else 0) - (qf / qt if qt else 0), "鲜品": (mx / mt if mt else 0) - (qx / qt if qt else 0), "合计": 0},
        ]
    if plants:
        mf = sum(m_fs[p]["冻品"] for p in plants)
        mx = sum(m_fs[p]["鲜品"] for p in plants)
        qf = sum(q_fs[p]["冻品"] for p in plants)
        qx = sum(q_fs[p]["鲜品"] for p in plants)
        mt, qt = mf + mx, qf + qx
        s2_rows += [
            {"工厂": "合计", "月份": f"{month_label}月", "冻品": mf / mt if mt else None, "鲜品": mx / mt if mt else None, "合计": 1 if mt else None},
            {"工厂": "", "月份": quarter_label, "冻品": qf / qt if qt else None, "鲜品": qx / qt if qt else None, "合计": 1 if qt else None},
            {"工厂": "", "月份": "差异", "冻品": (mf / mt if mt else 0) - (qf / qt if qt else 0), "鲜品": (mx / mt if mt else 0) - (qx / qt if qt else 0), "合计": 0},
        ]
    s2 = pd.DataFrame(s2_rows)

    spec_headers = _spec_headers_for_export(kind, material_spec_profile)
    extra_specs = sorted(sp for sp in spec_set if sp not in spec_headers)
    specs = list(spec_headers) + extra_specs
    spec_plant_slots = _plant_slots_for_kind(kind)
    spec_data_plants = sorted(set(list(m_spec.keys()) + list(q_spec.keys())))
    spec_plants = (spec_plant_slots + [p for p in spec_data_plants if p not in spec_plant_slots]) if spec_data_plants else []
    s3_rows = []
    for p in spec_plants + (["合计"] if spec_plants else []):
        mm = defaultdict(float)
        qq = defaultdict(float)
        if p == "合计":
            for pp in spec_plants:
                for k, v in m_spec[pp].items():
                    mm[k] += v
                for k, v in q_spec[pp].items():
                    qq[k] += v
        else:
            mm.update(m_spec[p])
            qq.update(q_spec[p])
        mt, qt = sum(mm.values()), sum(qq.values())
        r1 = {"工厂": p, "月份": f"{month_label}月"}
        r2 = {"工厂": "", "月份": quarter_label}
        r3 = {"工厂": "", "月份": "差异"}
        for sp in specs:
            mv = mm[sp] / mt if mt and mm[sp] else None
            qv = qq[sp] / qt if qt and qq[sp] else None
            r1[sp], r2[sp], r3[sp] = mv, qv, _diff_ratio(mv, qv)
        r1["合计"] = 1 if mt else None
        r2["合计"] = 1 if qt else None
        r3["合计"] = _diff_ratio(r1["合计"], r2["合计"])
        s3_rows += [r1, r2, r3]
    s3 = pd.DataFrame(s3_rows)

    return s1, s2, s3


def build_audit_detail(records, month_label, quarter_label, kind):
    month_qty_col = f"{month_label}月产量"
    q_qty_col = f"{quarter_label}月均产量"

    cols = [
        "工厂",
        "修行后原料",
        "产品族",
        "使用半成品规格",
        "是否纳入综合影响",
        "影响口径",
        "月行类型",
        "Q行类型",
        "差异行类型",
        "原料采购单价影响",
        "行情影响",
        "扣除行情后采购绩效",
        "修形利用率影响",
        "损耗率影响",
        "修形人工成本影响",
        "综合影响",
        month_qty_col,
        q_qty_col,
    ]

    rows = []
    for rec in records:
        plant = rec["plant"]
        month_rows = rec["month_tsc_rows"]
        q_rows = rec["q_tsc_rows"]
        rec_q_label = rec.get("q_label") or quarter_label
        q_parts = rec.get("q_part_rows") or []

        m_grp = defaultdict(lambda: defaultdict(list))
        q_grp = defaultdict(lambda: defaultdict(list))
        for r in month_rows:
            m_grp[r["修行后原料"]][r["行类型"]].append(r)
        for r in q_rows:
            q_grp[r["修行后原料"]][r["行类型"]].append(r)
        market_map = rec.get("market_impact_map") or {}
        fallback_actual_map = rec.get("fallback_actual_map") or {}
        product_family_map = rec.get("product_family_map") or {}
        allowed_mats = rec.get("allowed_mats")

        mats = sorted(set(m_grp.keys()) | set(q_grp.keys()))
        for mat in mats:
            if not _looks_like_semifinished_code(mat):
                continue
            if allowed_mats is not None and mat not in allowed_mats:
                continue
            md = m_grp.get(mat, {})
            qd = q_grp.get(mat, {})

            impact = resolve_material_impact(
                kind,
                md,
                qd,
                rec_q_label,
                market_map=market_map,
                fallback_actual_map=fallback_actual_map,
            )

            mrow, qrow, diff = select_material_rows(md, qd, rec_q_label)
            base_row = mrow or qrow or impact or {}

            m_qty = to_num((mrow or {}).get("半成品入库量")) or 0.0
            q_qty = _resolve_effective_q_qty(qrow, impact)
            impact_scope = str((impact or {}).get("影响口径", "")).strip()
            included = impact is not None and impact_scope == "总成本" and m_qty > 0

            price_impact = to_num((impact or {}).get("修形前原料综合耗用单价"))
            market_impact = _calc_market_gap(kind, price_impact, market_map, mat, m_qty, q_qty, impact)
            net_purchase_impact = None if price_impact is None else price_impact - market_impact
            total_impact = None
            if net_purchase_impact is not None:
                total_impact = (
                    net_purchase_impact
                    + (to_num((impact or {}).get("修形利用率")) or 0.0)
                    + (to_num((impact or {}).get("损耗率")) or 0.0)
                    + (to_num((impact or {}).get("半成品修形人工成本")) or 0.0)
                )

            row = {
                "工厂": plant,
                "修行后原料": mat,
                "产品族": product_family_map.get(mat) or base_row.get("产品族", ""),
                "使用半成品规格": base_row.get("使用半成品规格", ""),
                "是否纳入综合影响": "是" if included else "否",
                "影响口径": impact_scope,
                "月行类型": (mrow or {}).get("行类型", ""),
                "Q行类型": (qrow or {}).get("行类型", ""),
                "差异行类型": (diff or {}).get("行类型", ""),
                "原料采购单价影响": price_impact,
                "行情影响": market_impact,
                "扣除行情后采购绩效": net_purchase_impact,
                "修形利用率影响": to_num((impact or {}).get("修形利用率")),
                "损耗率影响": to_num((impact or {}).get("损耗率")),
                "修形人工成本影响": to_num((impact or {}).get("半成品修形人工成本")),
                "综合影响": total_impact,
                month_qty_col: m_qty,
                q_qty_col: q_qty,
            }
            rows.append(row)

    if not rows:
        return pd.DataFrame(columns=cols)

    out = pd.DataFrame(rows)
    out = out.reindex(columns=cols)
    out = out.sort_values(by=["工厂", "是否纳入综合影响", "修行后原料"], ascending=[True, False, True], kind="stable")
    return out


def _as_workbook_source(source):
    if hasattr(source, "seek"):
        source.seek(0)
    return source


def _source_cache_key(source):
    if isinstance(source, (str, Path)):
        path = Path(source)
        try:
            stat = path.stat()
            return ("path", str(path.resolve()), stat.st_mtime_ns, stat.st_size)
        except OSError:
            return ("path", str(path))

    name = getattr(source, "name", None)
    size = None
    if hasattr(source, "size"):
        size = getattr(source, "size")
    elif hasattr(source, "getbuffer"):
        try:
            size = len(source.getbuffer())
        except Exception:
            size = None
    return ("object", id(source), str(name or ""), size)


def _clear_source_caches():
    _WORKBOOK_CACHE.clear()
    _EXTRACT_CACHE.clear()


def _load_workbook_cached(source, data_only: bool = True):
    key = (_source_cache_key(source), bool(data_only))
    wb = _WORKBOOK_CACHE.get(key)
    if wb is not None:
        return wb
    wb = load_workbook(_as_workbook_source(source), data_only=data_only, read_only=False)
    if hasattr(wb, "_external_links"):
        wb._external_links = []
    _WORKBOOK_CACHE[key] = wb
    return wb

def _cached_extract(cache_name: str, source, kind: str, builder):
    key = (cache_name, _source_cache_key(source), kind)
    if key not in _EXTRACT_CACHE:
        _EXTRACT_CACHE[key] = builder()
    value = _EXTRACT_CACHE[key]
    if isinstance(value, dict):
        return {k: (v.copy() if isinstance(v, dict) else v) for k, v in value.items()}
    if isinstance(value, list):
        return [item.copy() if isinstance(item, dict) else item for item in value]
    return deepcopy(value)


def _norm_text(value):
    return "" if value is None else str(value).strip()


def _parse_formula_ref(value):
    text = _norm_text(value)
    if not text.startswith("="):
        return None
    m = re.match(r"^=(?:'(?P<sheet1>[^']+)'|(?P<sheet2>[^!]+))!\$?(?P<col>[A-Z]+)\$?(?P<row>\d+)$", text)
    if not m:
        return None
    sheet_name = m.group("sheet1") or m.group("sheet2") or ""
    col = m.group("col") or ""
    row = m.group("row") or ""
    if not sheet_name or not col or not row:
        return None
    return {"sheet_name": sheet_name, "col": col, "row": int(row)}


def _formula_ref_value(wb, value):
    ref = _parse_formula_ref(value)
    if not ref:
        return value
    sheet_name = ref.get("sheet_name")
    if sheet_name not in wb.sheetnames:
        return value
    return wb[sheet_name][f"{ref['col']}{ref['row']}"].value


def _resolve_tsc_source_info(wb, ws, current_row: int | None, prod_col: int | None, mat_col: int | None, spec_col: int | None):
    if not current_row:
        return {}

    prod_value = _formula_ref_value(wb, ws.cell(current_row, prod_col).value) if prod_col else None
    spec_value = _formula_ref_value(wb, ws.cell(current_row, spec_col).value) if spec_col else None

    mat_value = None
    if mat_col:
        for candidate_row in (current_row + 2, current_row + 1, current_row):
            if candidate_row > ws.max_row:
                continue
            candidate = _formula_ref_value(wb, ws.cell(candidate_row, mat_col).value)
            if _looks_like_semifinished_code(candidate):
                mat_value = candidate
                break

    return {
        "产品族": _norm_text(prod_value),
        "修行后原料": norm_code(mat_value),
        "使用半成品规格": _norm_text(spec_value),
    }

def _find_header_row_ws(ws, needle: str, max_rows: int = 6):
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, ws.max_column + 1):
            if _norm_text(ws.cell(r, c).value) == needle:
                return r
    return None


def _find_col_ws(ws, row: int | None, needle: str):
    if not row or row < 1:
        return None
    for c in range(1, ws.max_column + 1):
        if _norm_text(ws.cell(row, c).value) == needle:
            return c
    return None


def _first_col_ws(ws, rows, needles):
    for row in rows:
        for needle in needles:
            col = _find_col_ws(ws, row, needle)
            if col:
                return col
    return None


def _find_anchor_row(ws, current_row: int, mat_col: int | None, header_row: int):
    if not mat_col:
        return current_row
    for r in range(current_row, header_row, -1):
        if _looks_like_semifinished_code(ws.cell(r, mat_col).value):
            return r
    return current_row


def _iter_total_rows(ws, header_row: int):
    for r in range(header_row + 1, ws.max_row + 1):
        labels = {_norm_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 8) + 1)}
        if "总成本" in labels:
            yield r


def _detect_content_col_ws(ws, row_start: int, row_end: int, max_cols: int, tokens):
    best_col = None
    best_score = 0
    for c in range(1, min(ws.max_column, max_cols) + 1):
        score = 0
        for r in range(max(1, row_start), min(ws.max_row, row_end) + 1):
            text = _norm_text(ws.cell(r, c).value)
            if any(token in text for token in tokens):
                score += 1
        if score > best_score:
            best_col = c
            best_score = score
    return best_col if best_score > 0 else None


def _find_type_scope_cols_ws(ws, header_row: int):
    search_rows = [header_row, header_row - 1, header_row - 2]
    type_col = _first_col_ws(ws, search_rows, ["行类型"])
    if not type_col:
        type_col = _detect_content_col_ws(
            ws,
            header_row + 1,
            header_row + 12,
            10,
            ["实际单价", "实际价格", "规格占比", "差异", "对半成品成本的影响"],
        )
    scope_col = _first_col_ws(ws, search_rows, ["影响口径"])
    if not scope_col:
        scope_col = _detect_content_col_ws(ws, header_row + 1, header_row + 12, 10, ["单位成本", "总成本"])
    if not scope_col and type_col and type_col + 1 <= ws.max_column:
        scope_col = type_col + 1
    return type_col, scope_col


def _find_metric_col_ws(ws, header_row: int, needles):
    return _first_col_ws(ws, [header_row, header_row + 1, header_row - 1], needles)


def _find_actual_rows_before_total(ws, header_row: int, total_row: int, type_col: int | None):
    if not type_col:
        return None, None
    actual_rows = []
    for r in range(max(header_row + 1, total_row - 8), total_row):
        row_type = _norm_text(ws.cell(r, type_col).value)
        if _is_actual_price_row_type(row_type):
            actual_rows.append(r)
    current_row = next((r for r in actual_rows if _is_current_actual_price_row_type(_norm_text(ws.cell(r, type_col).value))), None)
    if current_row is None:
        current_row = actual_rows[0] if actual_rows else None
    previous_candidates = [r for r in actual_rows if r != current_row]
    ranked_previous = [
        (_baseline_actual_rank(_norm_text(ws.cell(r, type_col).value)), idx, r)
        for idx, r in enumerate(previous_candidates)
    ]
    ranked_previous = [item for item in ranked_previous if item[0] is not None]
    previous_row = sorted(ranked_previous, key=lambda item: (item[0], item[1]))[0][2] if ranked_previous else None
    if previous_row is None:
        previous_row = previous_candidates[0] if previous_candidates else None
    return current_row, previous_row


def _calc_raw_cost(kind: str, pre_price, util_rate, loss_rate):
    pre_value = to_num(pre_price)
    util_value = to_num(util_rate)
    loss_value = to_num(loss_rate)
    if pre_value is None or util_value in (None, 0) or loss_value is None:
        return None
    factor = 0.95 if kind == "腿肉" else 0.7
    return (pre_value - (1 - util_value - loss_value) * pre_value * factor) / util_value


def _is_reasonable_previous_metrics(util_rate, loss_rate, raw_cost, labor_cost):
    util_value = to_num(util_rate)
    loss_value = to_num(loss_rate)
    raw_value = to_num(raw_cost)
    labor_value = to_num(labor_cost)
    return (
        util_value is not None
        and 0 < util_value <= 1.5
        and loss_value is not None
        and -0.5 <= loss_value <= 0.5
        and raw_value is not None
        and raw_value > 0
        and labor_value is not None
    )


def _find_hidden_prev_pre_ws(ws, prev_row: int | None, total_col: int | None, current_pre):
    if not prev_row or not total_col:
        return None
    current_pre_num = to_num(current_pre)
    candidates = []
    for c in range(total_col + 1, ws.max_column + 1):
        value = to_num(ws.cell(prev_row, c).value)
        if value is None:
            continue
        if current_pre_num is not None and not (0 < value <= current_pre_num * 2.5):
            continue
        diff = abs(value - current_pre_num) if current_pre_num is not None else 0.0
        candidates.append((diff, c, value))
    return min(candidates)[2] if candidates else None


def _extract_market_context_map(source, kind: str):
    return _cached_extract("market_context", source, kind, lambda: _extract_market_context_map_uncached(source, kind))


def _extract_market_context_map_uncached(source, kind: str):
    wb = _load_workbook_cached(source, data_only=True)
    wb_formula = _load_workbook_cached(source, data_only=False)
    sheet_name = _pick_market_sheet_name(wb.sheetnames, kind)
    if not sheet_name:
        return {}
    ws = wb[sheet_name]
    ws_formula = wb_formula[sheet_name]
    header_row = _find_header_row_ws(ws, "修形前原料综合耗用单价")
    if not header_row:
        return {}

    search_rows = [header_row, header_row - 1, header_row - 2]
    prod_col = _first_col_ws(ws, search_rows, ["产品族"])
    mat_col = _first_col_ws(ws, search_rows, ["修行后原料", "修形后原料"])
    spec_col = _first_col_ws(ws, search_rows, ["使用半成品规格"])
    type_col, _ = _find_type_scope_cols_ws(ws, header_row)
    pre_col = _find_metric_col_ws(ws, header_row, ["修形前原料综合耗用单价"])
    util_col = _find_metric_col_ws(ws, header_row, ["修形利用率"])
    loss_col = _find_metric_col_ws(ws, header_row, ["损耗率"])
    raw_col = _find_metric_col_ws(ws, header_row, ["半成品原料成本"])
    labor_col = _find_metric_col_ws(ws, header_row, ["半成品修形人工成本", "人工", "半成品人工成本"])
    qty_col = _first_col_ws(ws, search_rows + [header_row + 1], ["半成品入库量"])
    total_col = _find_metric_col_ws(ws, header_row, ["半成品总成本"])

    out = {}
    for row_idx in _iter_total_rows(ws, header_row):
        current_row, previous_row = _find_actual_rows_before_total(ws, header_row, row_idx, type_col)
        anchor_row = _find_anchor_row(ws_formula, row_idx, mat_col, header_row)
        source_info = _resolve_tsc_source_info(wb_formula, ws_formula, current_row, prod_col, mat_col, spec_col)
        mat = norm_code(source_info.get("修行后原料")) or (norm_code(ws_formula.cell(anchor_row, mat_col).value) if mat_col else "")
        if not _looks_like_semifinished_code(mat):
            continue

        product_family = source_info.get("产品族") or (_norm_text(ws_formula.cell(anchor_row, prod_col).value) if prod_col else "")
        spec_name = source_info.get("使用半成品规格") or (_norm_text(ws_formula.cell(anchor_row, spec_col).value) if spec_col else "")
        current_pre = to_num(ws.cell(current_row, pre_col).value) if current_row and pre_col else None
        current_util = to_num(ws.cell(current_row, util_col).value) if current_row and util_col else None
        current_loss = to_num(ws.cell(current_row, loss_col).value) if current_row and loss_col else None
        current_raw = to_num(ws.cell(current_row, raw_col).value) if current_row and raw_col else None
        current_lab = to_num(ws.cell(current_row, labor_col).value) if current_row and labor_col else None
        current_qty = to_num(ws.cell(current_row, qty_col).value) if current_row and qty_col else None
        if current_qty is None and anchor_row and qty_col:
            current_qty = to_num(ws.cell(anchor_row, qty_col).value)
        previous_pre = to_num(ws.cell(previous_row, pre_col).value) if previous_row and pre_col else None
        if previous_pre is None:
            previous_pre = _find_hidden_prev_pre_ws(ws, previous_row, total_col, current_pre)
        previous_util = to_num(ws.cell(previous_row, util_col).value) if previous_row and util_col else None
        previous_loss = to_num(ws.cell(previous_row, loss_col).value) if previous_row and loss_col else None
        previous_raw = to_num(ws.cell(previous_row, raw_col).value) if previous_row and raw_col else None
        previous_lab = to_num(ws.cell(previous_row, labor_col).value) if previous_row and labor_col else None
        market_unit = None
        baseline_raw = _calc_raw_cost(kind, previous_pre, current_util, current_loss)
        if baseline_raw is not None and current_raw is not None:
            market_unit = current_raw - baseline_raw
        market_impact = to_num(ws.cell(row_idx, total_col).value) if total_col else None
        if market_impact is None and market_unit is not None and current_qty not in (None, 0):
            market_impact = market_unit * current_qty
        if market_impact is None and previous_raw is None and previous_pre is not None:
            previous_raw = _calc_raw_cost(kind, previous_pre, previous_util, previous_loss)
        if market_impact is None:
            impact_row = _fill_missing_total_metrics(
                {
                    "修形前原料综合耗用单价": None,
                    "修形利用率": None,
                    "损耗率": None,
                    "半成品修形人工成本": None,
                    "半成品总成本": None,
                },
                kind,
                {
                    "pre": current_pre,
                    "util": current_util,
                    "loss": current_loss,
                    "raw": current_raw,
                    "lab": current_lab,
                    "qty": current_qty,
                },
                {
                    "util": previous_util,
                    "loss": previous_loss,
                    "raw": previous_raw,
                    "lab": previous_lab,
                },
                previous_pre,
            )
            market_impact = to_num(impact_row.get("半成品总成本"))
        out[mat] = {
            "sheet_name": sheet_name,
            "产品族": product_family,
            "使用半成品规格": spec_name,
            "previous_pre": previous_pre,
            "market_unit_impact": market_unit,
            "行情影响": market_impact or 0.0,
        }
    return out

def _fill_missing_total_metrics(row: dict, kind: str, current_vals: dict, previous_vals: dict, previous_pre):
    cur_pre = to_num(current_vals.get("pre"))
    cur_util = to_num(current_vals.get("util"))
    cur_loss = to_num(current_vals.get("loss"))
    cur_raw = to_num(current_vals.get("raw"))
    cur_lab = to_num(current_vals.get("lab"))
    cur_qty = to_num(current_vals.get("qty"))
    prev_pre = to_num(previous_pre)
    prev_util = to_num(previous_vals.get("util"))
    prev_loss = to_num(previous_vals.get("loss"))
    prev_raw = to_num(previous_vals.get("raw"))
    prev_lab = to_num(previous_vals.get("lab"))

    if (
        cur_pre is None
        or cur_util in (None, 0)
        or cur_loss is None
        or cur_raw is None
        or cur_lab is None
        or cur_qty in (None, 0)
        or prev_pre is None
        or not _is_reasonable_previous_metrics(prev_util, prev_loss, prev_raw, prev_lab)
    ):
        return row

    prev_raw_with_cur = _calc_raw_cost(kind, prev_pre, cur_util, cur_loss)
    cur_raw_with_prev_loss = _calc_raw_cost(kind, cur_pre, cur_util, prev_loss)
    if prev_raw_with_cur is None or cur_raw_with_prev_loss is None:
        return row

    price_total = (cur_raw - prev_raw_with_cur) * cur_qty
    loss_total = (cur_raw - cur_raw_with_prev_loss) * cur_qty
    util_total = ((cur_raw - prev_raw) - (price_total / cur_qty) - (loss_total / cur_qty)) * cur_qty
    labor_total = (cur_lab - prev_lab) * cur_qty
    raw_total = price_total + util_total + loss_total
    total_total = raw_total + labor_total

    if row.get("修形前原料综合耗用单价") is None:
        row["修形前原料综合耗用单价"] = price_total
    if row.get("修形利用率") is None:
        row["修形利用率"] = util_total
    if row.get("损耗率") is None:
        row["损耗率"] = loss_total
    if row.get("半成品修形人工成本") is None:
        row["半成品修形人工成本"] = labor_total
    if row.get("半成品总成本") is None:
        row["半成品总成本"] = total_total
    return row


def _fallback_total_impact_from_actual_rows(kind: str, current_row: dict | None, previous_row: dict | None, qty=None):
    """Build total impact from current/previous actual rows when the uploaded month TSC
    only has actual price rows and lacks explicit 差异/对半成品成本的影响 rows.

    This keeps rows such as DL-其他 in the screen consistent with the exported workbook,
    instead of showing only volume with all impact values as 0.
    """
    current_vals = _actual_metric_values(current_row, kind)
    previous_vals = _actual_metric_values(previous_row, kind)
    qty_value = to_num(qty)
    if qty_value in (None, 0):
        qty_value = to_num(current_vals.get("qty"))
    base_row = current_row or previous_row or {}
    impact = _build_blank_total_impact_row(base_row, qty=qty_value)
    previous_pre = to_num((previous_row or {}).get("修形前原料综合耗用单价"))
    impact = _fill_missing_total_metrics(impact, kind, current_vals, previous_vals, previous_pre)
    return impact if _has_total_metrics(impact) else None


def _extract_total_impact_rows(source, kind: str):
    return _cached_extract("total_impact_rows", source, kind, lambda: _extract_total_impact_rows_uncached(source, kind))


def _extract_total_impact_rows_uncached(source, kind: str):
    wb = _load_workbook_cached(source, data_only=True)
    wb_formula = _load_workbook_cached(source, data_only=False)
    sheet_name = f"{kind}TSC"
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    ws_formula = wb_formula[sheet_name]
    header_row = _find_header_row_ws(ws, "修形前原料综合耗用单价")
    if not header_row:
        return []

    search_rows = [header_row, header_row - 1, header_row - 2]
    prod_col = _first_col_ws(ws, search_rows, ["产品族"])
    mat_col = _first_col_ws(ws, search_rows, ["修行后原料", "修形后原料"])
    spec_col = _first_col_ws(ws, search_rows, ["使用半成品规格"])
    qty_col = _first_col_ws(ws, search_rows, ["半成品入库量"])
    type_col, _ = _find_type_scope_cols_ws(ws, header_row)
    raw_col = _find_metric_col_ws(ws, header_row, ["半成品原料成本"])
    metric_cols = {
        "修形前原料综合耗用单价": _find_metric_col_ws(ws, header_row, ["修形前原料综合耗用单价"]),
        "修形利用率": _find_metric_col_ws(ws, header_row, ["修形利用率"]),
        "损耗率": _find_metric_col_ws(ws, header_row, ["损耗率"]),
        "半成品修形人工成本": _find_metric_col_ws(ws, header_row, ["半成品修形人工成本", "人工", "半成品人工成本"]),
        "半成品总成本": _find_metric_col_ws(ws, header_row, ["半成品总成本"]),
    }
    market_ctx = _extract_market_context_map(source, kind)

    out = []
    for row_idx in _iter_total_rows(ws, header_row):
        current_row, previous_row = _find_actual_rows_before_total(ws, header_row, row_idx, type_col)
        anchor_row = _find_anchor_row(ws_formula, row_idx, mat_col, header_row)
        source_info = _resolve_tsc_source_info(wb_formula, ws_formula, current_row, prod_col, mat_col, spec_col)
        mat = norm_code(source_info.get("修行后原料")) or (norm_code(ws_formula.cell(anchor_row, mat_col).value) if mat_col else "")
        if not _looks_like_semifinished_code(mat):
            continue

        product_family = source_info.get("产品族") or (_norm_text(ws_formula.cell(anchor_row, prod_col).value) if prod_col else "")
        spec_name = source_info.get("使用半成品规格") or (_norm_text(ws_formula.cell(anchor_row, spec_col).value) if spec_col else "")
        qty_value = to_num(ws.cell(current_row, qty_col).value) if current_row and qty_col else None
        if qty_value is None and anchor_row and qty_col:
            qty_value = to_num(ws.cell(anchor_row, qty_col).value)

        row = {
            "产品族": product_family,
            "修行后原料": mat,
            "使用半成品规格": spec_name,
            "行类型": "对半成品成本的影响",
            "影响口径": "总成本",
            "综合单价": None,
            "半成品入库量": qty_value,
            "raw_share": {},
        }
        for key, col in metric_cols.items():
            row[key] = to_num(ws.cell(row_idx, col).value) if col else None
        if current_row and previous_row:
            row = _fill_missing_total_metrics(
                row,
                kind,
                {
                    "pre": to_num(ws.cell(current_row, metric_cols["修形前原料综合耗用单价"]).value)
                    if metric_cols["修形前原料综合耗用单价"]
                    else None,
                    "util": to_num(ws.cell(current_row, metric_cols["修形利用率"]).value) if metric_cols["修形利用率"] else None,
                    "loss": to_num(ws.cell(current_row, metric_cols["损耗率"]).value) if metric_cols["损耗率"] else None,
                    "raw": to_num(ws.cell(current_row, raw_col).value) if raw_col else None,
                    "lab": to_num(ws.cell(current_row, metric_cols["半成品修形人工成本"]).value)
                    if metric_cols["半成品修形人工成本"]
                    else None,
                    "qty": qty_value,
                },
                {
                    "util": to_num(ws.cell(previous_row, metric_cols["修形利用率"]).value) if metric_cols["修形利用率"] else None,
                    "loss": to_num(ws.cell(previous_row, metric_cols["损耗率"]).value) if metric_cols["损耗率"] else None,
                    "raw": to_num(ws.cell(previous_row, raw_col).value) if raw_col else None,
                    "lab": to_num(ws.cell(previous_row, metric_cols["半成品修形人工成本"]).value)
                    if metric_cols["半成品修形人工成本"]
                    else None,
                },
                (market_ctx.get(mat) or {}).get("previous_pre"),
            )
        out.append(row)
    return out

def _replace_impact_rows(rows, total_rows):
    if not total_rows:
        return rows
    base_rows = [r for r in rows if "对半成品成本的影响" not in str(r.get("行类型", ""))]
    by_mat = {}
    for row in total_rows:
        mat = row.get("修行后原料")
        if mat:
            by_mat[mat] = row
    return base_rows + list(by_mat.values())


def _has_total_impact_rows(rows):
    return any(
        "对半成品成本的影响" in str(r.get("行类型", "")) and str(r.get("影响口径", "")).strip() == "总成本"
        for r in rows
    )


def _has_month_actual_rows(rows):
    return any(_is_current_actual_price_row_type(r) for r in rows)


def _pick_market_sheet_name(sheet_names, kind: str):
    kind_label = str(kind).strip()
    if kind_label in ("腿肉", "胸肉"):
        exact_name = f"{kind_label}行情-较季度"
        return exact_name if exact_name in sheet_names else None

    candidates = [f"{kind_label}行情-较季度", f"{kind_label}行情-季度", f"{kind_label}行情", f"行情情况-{kind_label}"]
    if kind_label == "其他":
        candidates += ["其他较季度", "其他行情情况", "其他-行情"]
    for name in candidates:
        if name in sheet_names:
            return name
    return None


def _extract_market_impact_map(source, kind: str):
    return _cached_extract("market_impact_map", source, kind, lambda: _extract_market_impact_map_uncached(source, kind))


def _extract_kind_market_impact_map(kind: str, month_source, quarter_source):
    kind_label = str(kind).strip()
    source = month_source if kind_label in ("腿肉", "胸肉") else quarter_source
    return _extract_market_impact_map(source, kind_label)


def _extract_market_impact_map_uncached(source, kind: str):
    market_ctx = _extract_market_context_map(source, kind)
    return {
        mat: {
            "sheet_name": payload.get("sheet_name"),
            "product_family": payload.get("\u4ea7\u54c1\u65cf", ""),
            "spec_name": payload.get("\u4f7f\u7528\u534a\u6210\u54c1\u89c4\u683c", ""),
            "market_unit_impact": to_num(payload.get("market_unit_impact")),
            "market_total_impact": to_num(payload.get("\u884c\u60c5\u5f71\u54cd")) or 0.0,
        }
        for mat, payload in market_ctx.items()
    }


def _resolve_market_impact_value(market_payload: dict | None, month_qty=None, q_qty=None):
    market_payload = market_payload or {}
    unit_impact = to_num(market_payload.get("market_unit_impact"))
    if unit_impact is not None:
        scale_qty = to_num(month_qty)
        if scale_qty in (None, 0):
            scale_qty = to_num(q_qty)
        if scale_qty not in (None, 0):
            return unit_impact * scale_qty
    return to_num(market_payload.get("market_total_impact")) or 0.0


def _calc_market_gap(kind: str, price_impact, market_map: dict, mat, month_qty, q_qty, impact_row):
    price_gap = to_num(price_impact) or 0.0
    if str(kind).strip() == "其他":
        return price_gap
    if q_qty > 0 and not _is_zero_market_suppressed_impact_row(impact_row):
        return _resolve_market_impact_value((market_map or {}).get(mat), month_qty, q_qty)
    return 0.0


def _manual_actual_score(payload: dict | None):
    payload = payload or {}
    score = 0
    for key in ("pre", "util", "loss", "raw", "lab"):
        if to_num(payload.get(key)) is not None:
            score += 10
    qty = to_num(payload.get("qty")) or 0.0
    return score + qty / 1000000.0


def _merge_manual_actual_row(existing: dict | None, candidate: dict | None):
    if not candidate:
        return existing or {}
    if not existing:
        return candidate
    return candidate if _manual_actual_score(candidate) >= _manual_actual_score(existing) else existing


def _extract_manual_actual_map(source, kind: str):
    return _cached_extract("manual_actual_map", source, kind, lambda: _extract_manual_actual_map_uncached(source, kind))


def _extract_manual_actual_map_uncached(source, kind: str):
    wb = _load_workbook_cached(source)
    out = {}

    if "校验" in wb.sheetnames:
        ws = wb["校验"]
        header_row = _find_header_row_ws(ws, "物料号") or _find_header_row_ws(ws, "物料号　　")
        if header_row:
            rows = [header_row, header_row + 1]
            mat_col = _first_col_ws(ws, rows, ["物料号", "物料号　　"])
            raw_code_col = _first_col_ws(ws, rows, ["原料号", "原料号　　"])
            qty_col = _first_col_ws(ws, rows, ["入库数量", "入库数量　", "入库数量　　"])
            aux_col = _first_col_ws(ws, rows, ["辅助"])
            amount_col = _first_col_ws(ws, rows, ["调整后实际额"])
            pre_col = _first_col_ws(ws, rows, ["修形前原料单价", "修形前原料综合耗用单价"])
            util_col = _first_col_ws(ws, rows, ["修形利用率"])
            loss_col = _first_col_ws(ws, rows, ["失水率", "损耗率"])
            for row_idx in range(header_row + 1, ws.max_row + 1):
                mat = norm_code(ws.cell(row_idx, mat_col).value) if mat_col else ""
                if not mat:
                    continue
                raw_code = _norm_text(ws.cell(row_idx, raw_code_col).value) if raw_code_col else ""
                if raw_code and "人工费用" not in raw_code:
                    continue
                pre = to_num(ws.cell(row_idx, pre_col).value) if pre_col else None
                util = to_num(ws.cell(row_idx, util_col).value) if util_col else None
                loss = to_num(ws.cell(row_idx, loss_col).value) if loss_col else None
                qty = to_num(ws.cell(row_idx, qty_col).value) if qty_col else None
                raw = None
                if pre is not None and util not in (None, 0) and loss is not None:
                    raw = _calc_raw_cost(kind, pre, util, loss)
                if raw is None:
                    amount = to_num(ws.cell(row_idx, amount_col).value) if amount_col else None
                    aux = to_num(ws.cell(row_idx, aux_col).value) if aux_col else None
                    if amount is not None and aux not in (None, 0):
                        raw = amount / aux
                candidate = {"pre": pre, "util": util, "loss": loss, "raw": raw, "qty": qty}
                out[mat] = _merge_manual_actual_row(out.get(mat), candidate)

    if "人工" in wb.sheetnames:
        ws = wb["人工"]
        header_row = _find_header_row_ws(ws, "物料号") or _find_header_row_ws(ws, "物料号　　")
        if header_row:
            rows = [header_row, header_row + 1]
            mat_col = _first_col_ws(ws, rows, ["物料号", "物料号　　"])
            raw_code_col = _first_col_ws(ws, rows, ["原料号", "原料号　　"])
            qty_col = _first_col_ws(ws, rows, ["入库数量", "入库数量　", "入库数量　　"])
            amount_col = _first_col_ws(ws, rows, ["实际金额"])
            for row_idx in range(header_row + 1, ws.max_row + 1):
                mat = norm_code(ws.cell(row_idx, mat_col).value) if mat_col else ""
                if not mat:
                    continue
                raw_code = _norm_text(ws.cell(row_idx, raw_code_col).value) if raw_code_col else ""
                if raw_code and "人工费用" not in raw_code:
                    continue
                qty = to_num(ws.cell(row_idx, qty_col).value) if qty_col else None
                amount = to_num(ws.cell(row_idx, amount_col).value) if amount_col else None
                labor = (amount / qty) if amount is not None and qty not in (None, 0) else None
                current = dict(out.get(mat) or {})
                if labor is not None:
                    current["lab"] = labor
                if qty is not None and (to_num(current.get("qty")) is None or qty > (to_num(current.get("qty")) or 0.0)):
                    current["qty"] = qty
                out[mat] = current

    return out


def _looks_like_product_family(value) -> bool:
    text = _norm_text(value)
    if not text or text in ("分类", "人工费用", "#N/A"):
        return False
    if to_num(text) is not None:
        return False
    return True


def _looks_like_semifinished_code(value) -> bool:
    code = norm_code(value)
    return code.isdigit() and len(code) >= 8 and not code.startswith("310")


def _extract_product_family_map(source, kind: str):
    return _cached_extract("product_family_map", source, kind, lambda: _extract_product_family_map_uncached(source, kind))


def _extract_product_family_map_uncached(source, kind: str):
    wb = _load_workbook_cached(source)
    sheet_name = None
    for sn in wb.sheetnames:
        if sn.endswith(kind) and not sn.endswith("TSC"):
            sheet_name = sn
            break
    if not sheet_name:
        return {}

    ws = wb[sheet_name]
    out = {}
    for row_idx in range(1, ws.max_row + 1):
        left_mat = norm_code(ws.cell(row_idx, 1).value)
        left_family = _norm_text(ws.cell(row_idx, 4).value)
        right_mat = norm_code(ws.cell(row_idx, 5).value)
        right_family = _norm_text(ws.cell(row_idx, 8).value)

        if _looks_like_semifinished_code(left_mat) and _looks_like_product_family(left_family):
            out[left_mat] = left_family
        if _looks_like_semifinished_code(right_mat) and _looks_like_product_family(right_family):
            out[right_mat] = right_family
    return out


def fmt_pct(x):
    if x is None or pd.isna(x):
        return "-"
    try:
        return f"{float(x):.0%}"
    except Exception:
        return str(x)


def fmt_pct_1(x):
    if x is None or pd.isna(x):
        return "-"
    try:
        return f"{float(x):.1%}"
    except Exception:
        return str(x)


def fmt_int_paren(x):
    if x is None or pd.isna(x):
        return "-"
    try:
        v = float(x)
    except Exception:
        return str(x)
    n = int(round(v))
    if n < 0:
        return f"({abs(n):,})"
    return f"{n:,}"


def _fill_display_blanks(df: pd.DataFrame):
    out = df.copy().astype(object)
    out = out.where(~out.isna(), "-")
    return out.replace("", "-")


def _center_display(df: pd.DataFrame):
    out = _fill_display_blanks(df)
    return out.style.set_properties(**{"text-align": "center"}).set_table_styles(
        [
            {"selector": "th", "props": [("text-align", "center")]},
            {"selector": "td", "props": [("text-align", "center")]},
        ]
    )


def _rename_q_avg_qty_column(df, quarter_label, baseline_label=None):
    out = df.copy()
    source_col = f"{quarter_label}月均产量"
    target_label = baseline_label or quarter_label
    target_col = f"{target_label}月均产量"
    if source_col != target_col and source_col in out.columns and target_col not in out.columns:
        out = out.rename(columns={source_col: target_col})
    return out


def _prepare_detail_export_df(df, quarter_label, baseline_label, market_col: str, kind_label: str):
    out = _rename_q_avg_qty_column(df, quarter_label, baseline_label)
    if "半成品总成本" in out.columns and market_col in out.columns:
        out = out.copy()
        out["半成品总成本"] = out[market_col]
    if str(kind_label).strip() not in ("腿肉", "胸肉"):
        return out

    price_col = "原料采购单价影响"
    if price_col not in out.columns or market_col not in out.columns:
        return out

    out = out.copy()
    out[price_col] = out[market_col]
    out = out.drop(columns=[market_col])
    return out.rename(columns={price_col: market_col})


def _rename_period_label_value(df, quarter_label, baseline_label=None):
    out = df.copy()
    target_label = baseline_label or quarter_label
    if out.empty or "月份" not in out.columns or target_label == quarter_label:
        return out
    out["月份"] = out["月份"].map(lambda value: target_label if str(value).strip() == str(quarter_label) else value)
    return out

def present_s1(df, month_label, quarter_label, baseline_label=None):
    out = _rename_q_avg_qty_column(df, quarter_label, baseline_label)
    if out.empty or "工厂" not in out.columns:
        return out
    out = out.copy().astype(object)
    market_col = "行情差异" if "行情差异" in out.columns else "行情影响"
    q_qty_col = f"{baseline_label or quarter_label}月均产量"
    val_cols = [
        "原料采购单价影响",
        market_col,
        "扣除行情后采购绩效",
        "修形利用率影响",
        "损耗率影响",
        "修形人工成本影响",
        "综合影响",
        f"{month_label}月产量",
        q_qty_col,
    ]
    impact_mask = out["工厂"].astype(str) == "影响"
    for c in val_cols:
        if c not in out.columns:
            continue
        out.loc[~impact_mask, c] = out.loc[~impact_mask, c].apply(fmt_int_paren)
    for c in ["扣除行情后采购绩效", "修形利用率影响", "损耗率影响", "修形人工成本影响", "综合影响"]:
        if c in out.columns:
            out.loc[impact_mask, c] = out.loc[impact_mask, c].apply(fmt_pct_1)
    for c in ["原料采购单价影响", market_col, f"{month_label}月产量", q_qty_col]:
        if c in out.columns:
            out.loc[impact_mask, c] = "-"
    return _fill_display_blanks(out)


def present(df, pct_cols=None):
    out = df.copy().astype(object)
    pct_cols = pct_cols or []
    for c in pct_cols:
        if c in out.columns:
            out[c] = out[c].apply(fmt_pct)
    return _fill_display_blanks(out)


def _calc_table_width(ws, header_row, start_col):
    c = start_col
    while c <= ws.max_column and ws.cell(header_row, c).value not in (None, ""):
        c += 1
    return c - start_col


def _calc_table_rows(ws, header_row, start_col):
    r = header_row + 1
    while r <= ws.max_row and ws.cell(r, start_col).value not in (None, ""):
        r += 1
    return max(0, r - header_row - 1)


def style_summary_sheet(ws, header_row=1, start_col=1, data_rows=None, data_cols=None):
    yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_fmt = '#,##0;[Red](#,##0)'
    pct_fmt = "0%"
    impact_pct_fmt = '0.0%;[Red](0.0%)'

    data_cols = data_cols if data_cols is not None else _calc_table_width(ws, header_row, start_col)
    data_rows = data_rows if data_rows is not None else _calc_table_rows(ws, header_row, start_col)
    if data_cols <= 0:
        return

    headers = {
        c: str(ws.cell(header_row, c).value or "").strip()
        for c in range(start_col, start_col + data_cols)
    }
    pct_cols = {
        c for c, text in headers.items()
        if text in {"冻品", "鲜品", "合计"}
    }
    impact_pct_cols = {
        c for c, text in headers.items()
        if text in {"扣除行情后采购绩效", "修形利用率影响", "损耗率影响", "修形人工成本影响", "综合影响"}
    }

    for r in range(header_row, header_row + data_rows + 1):
        label = str(ws.cell(r, start_col).value or "").strip()
        period_text = str(ws.cell(r, start_col + 1).value or "").strip() if data_cols >= 2 else ""
        is_emphasis_row = label in {"合计", "影响"} or period_text in {"差异", "影响"}
        for c in range(start_col, start_col + data_cols):
            cell = ws.cell(r, c)
            cell.alignment = center
            if r == header_row or is_emphasis_row:
                cell.fill = yellow
            if r > header_row and cell.value in (None, ""):
                cell.value = "-"
            if r > header_row and isinstance(cell.value, (int, float)):
                if label == "影响" and c in impact_pct_cols:
                    cell.number_format = impact_pct_fmt
                else:
                    cell.number_format = pct_fmt if c in pct_cols else num_fmt


def style_audit_sheet(ws, header_row=1, start_col=1, data_rows=None, data_cols=None):
    blue = PatternFill(fill_type="solid", fgColor="DCE6F1")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_fmt = '#,##0_);[Red](#,##0)'

    data_cols = data_cols if data_cols is not None else _calc_table_width(ws, header_row, start_col)
    data_rows = data_rows if data_rows is not None else _calc_table_rows(ws, header_row, start_col)
    if data_cols <= 0:
        return

    for c in range(start_col, start_col + data_cols):
        ws.cell(header_row, c).fill = blue
    for r in range(header_row, header_row + data_rows + 1):
        for c in range(start_col, start_col + data_cols):
            cell = ws.cell(r, c)
            cell.alignment = center
            if r > header_row and isinstance(cell.value, (int, float)):
                cell.number_format = num_fmt


def style_ratio_sheet(ws):
    yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
    center = Alignment(horizontal="center", vertical="center")
    pct_fmt = "0.0%"
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).fill = yellow
    for r in range(1, ws.max_row + 1):
        row_label = ws.cell(r, 1).value
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.alignment = center
            if row_label == "合计":
                cell.fill = yellow
            if r > 1 and c >= 3 and isinstance(cell.value, (int, float)):
                cell.number_format = pct_fmt


def _text_width(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        text = f"{value:,.3f}".rstrip("0").rstrip(".")
    else:
        text = str(value).strip()
    if not text:
        return 0
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in text)


def _autosize_columns(ws, start_row=1, end_row=None, min_width=8, max_width=28, padding=2):
    last_row = min(ws.max_row, end_row) if end_row else ws.max_row
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        best = 0
        for row_idx in range(start_row, last_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if value in (None, ""):
                continue
            best = max(best, _text_width(value))
        if best > 0:
            ws.column_dimensions[letter].width = min(max_width, max(min_width, best + padding))


def _apply_min_widths(ws, width_map: dict):
    for col, width in (width_map or {}).items():
        current = ws.column_dimensions[col].width or 0
        ws.column_dimensions[col].width = max(current, width)


def _set_row_heights(ws, row_ranges):
    for start_row, end_row, height in row_ranges:
        last_row = min(end_row or ws.max_row, ws.max_row)
        for row_idx in range(start_row, last_row + 1):
            ws.row_dimensions[row_idx].height = height


def _apply_output_sheet_layout(ws):
    title = str(ws.title or "")
    if title == "Sheet1":
        _autosize_columns(ws, start_row=4, end_row=12, min_width=10, max_width=18)
        _apply_min_widths(ws, {"A": 24, "B": 12, "C": 12, "D": 12, "E": 14, "F": 12, "G": 12})
        _set_row_heights(ws, [(1, 1, 22), (2, 2, 30), (4, 9, 22)])
        return

    if title in ("腿肉", "胸肉"):
        _autosize_columns(ws, start_row=3, end_row=200, min_width=8, max_width=28)
        _apply_min_widths(
            ws,
            {
                "A": 10, "B": 12, "C": 24, "D": 12, "E": 12,
                "F": 12, "G": 12, "H": 12, "I": 12, "J": 12, "K": 12, "L": 10, "M": 12,
                "N": 12, "O": 24, "P": 12, "Q": 12, "R": 12, "S": 12, "T": 12, "U": 12, "V": 12,
                "W": 12, "X": 24, "Y": 12, "Z": 12, "AA": 12, "AB": 12, "AC": 12, "AD": 12, "AE": 12, "AF": 12,
                "AQ": 10, "AR": 12, "AS": 12, "AT": 12, "AU": 12, "AV": 12, "AW": 12, "AX": 12, "AY": 12, "AZ": 12,
                "BA": 12, "BB": 12,
            },
        )
        _set_row_heights(ws, [(1, 1, 28), (2, 4, 22), (5, None, 20)])
        return

    if title.startswith("腿肉占比-") or title.startswith("胸肉占比-"):
        _autosize_columns(ws, start_row=3, end_row=250, min_width=8, max_width=30)
        _apply_min_widths(
            ws,
            {
                "A": 8, "B": 10, "C": 14, "D": 30, "E": 12, "F": 10, "G": 12, "H": 14,
                "I": 12, "J": 18, "K": 10, "L": 12, "M": 12, "N": 12, "O": 12, "P": 10,
                "Q": 12, "R": 12, "S": 12, "T": 10, "U": 12, "V": 12, "W": 12, "X": 12,
            },
        )
        _set_row_heights(ws, [(1, 3, 22), (4, None, 20)])
        return

    if title.startswith("腿肉分工厂-") or title.startswith("胸肉分工厂-"):
        _autosize_columns(ws, start_row=2, end_row=120, min_width=8, max_width=14)
        _apply_min_widths(
            ws,
            {
                "A": 8, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10, "H": 10,
                "I": 10, "J": 10, "K": 10, "L": 10, "M": 12, "N": 12, "O": 10, "P": 10,
                "Q": 10, "R": 10, "S": 10, "T": 10, "U": 10, "V": 10, "W": 12, "X": 12,
                "Y": 10, "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10, "AE": 10, "AF": 12,
                "AG": 10, "AH": 10, "AI": 10, "AJ": 10, "AK": 10, "AL": 10, "AM": 10, "AN": 10,
                "AO": 10, "AP": 10, "AQ": 10, "AR": 10,
            },
        )
        _set_row_heights(ws, [(1, 2, 22), (3, None, 20)])
        return

    if title == "原料规格":
        _autosize_columns(ws, start_row=1, end_row=400, min_width=10, max_width=30)
        _apply_min_widths(ws, {"A": 10, "B": 14, "C": 30, "D": 18})
        _set_row_heights(ws, [(1, 1, 22), (2, None, 20)])


def _apply_output_layout(wb):
    for ws in wb.worksheets:
        _apply_output_sheet_layout(ws)


def write_audit_summary_sheet(writer, sheet_name, audit_df, summary_df):
    # 右侧摘要起始列尽量对齐目标模板（W列），同时为左侧核对表保留间隔
    summary_startcol = max(22, len(audit_df.columns) + 2)  # 0-based
    audit_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0, startcol=0)
    summary_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0, startcol=summary_startcol)
    ws = writer.book[sheet_name]
    style_audit_sheet(ws, header_row=1, start_col=1, data_rows=len(audit_df), data_cols=len(audit_df.columns))
    style_summary_sheet(
        ws,
        header_row=1,
        start_col=summary_startcol + 1,
        data_rows=len(summary_df),
        data_cols=len(summary_df.columns),
    )


def build_overview_sheet_df(kind_summaries):
    metrics = [
        "扣除行情后采购绩效",
        "修形利用率影响",
        "损耗率影响",
        "修形人工成本影响",
    ]
    rows = []
    for label, df in kind_summaries:
        total = {}
        if not df.empty and "工厂" in df.columns:
            total_rows = df[df["工厂"].astype(str) == "合计"]
            if not total_rows.empty:
                total = total_rows.iloc[0].to_dict()
        row = {"项目": label}
        for metric in metrics:
            row[metric] = _safe_float(total.get(metric))
        row["预留"] = 0.0
        row["综合影响"] = sum(_safe_float(row.get(metric)) for metric in metrics)
        rows.append(row)
    if rows:
        total_row = {"项目": "合计"}
        for col in metrics + ["预留", "综合影响"]:
            total_row[col] = sum(_safe_float(row.get(col)) for row in rows)
        rows.append(total_row)
    columns = ["项目", *metrics, "预留", "综合影响"]
    return pd.DataFrame(rows, columns=columns)


def _is_display_export_sheet(title: str) -> bool:
    title = str(title or "")
    return (
        title in {"腿肉", "胸肉", "其他"}
        or title.startswith("腿肉占比-")
        or title.startswith("胸肉占比-")
        or title.startswith("其他占比-")
        or title.startswith("腿肉分工厂-")
        or title.startswith("胸肉分工厂-")
        or title.startswith("其他分工厂-")
    )


def _apply_export_column_widths(ws, header_row: int = 1):
    title = str(ws.title or "")
    spec_tokens = ("80g", "110g", "120", "170", "200", "220", "245", "260", "285", "300")

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(header_row, col_idx).value
        header_text = "" if header is None else str(header).strip()
        header_key = _norm_header_key(header_text)
        width = None

        if header_key == "工厂":
            width = 10
        elif header_key == "月份":
            width = 10
        elif header_key in {"产品族", "鲜冻", "规格", "原料规格", "对应部位"}:
            width = 14
        elif header_key in {"使用半成品规格", "原料描述", "半成品描述"}:
            width = 32
        elif header_key in {"原料号", "半成品", "物料号"}:
            width = 14
        elif header_key in {"冻品", "鲜品", "合计"}:
            width = 12
        elif header_key.endswith("产量"):
            width = 12
        elif any(token in header_text for token in ("影响", "绩效", "差异", "成本")):
            width = 16
        elif any(token in header_text.lower() for token in spec_tokens) or "/" in header_text:
            width = 12

        if title == "Sheet1":
            width = max(width or 0, 12 if col_idx == 1 else 14)
        elif title in {"腿肉", "胸肉", "其他"}:
            base = 10 if col_idx == 1 else 12
            width = max(width or 0, base)
        elif title.startswith(("腿肉占比-", "胸肉占比-", "其他占比-")):
            base = 10 if col_idx <= 2 else 12
            width = max(width or 0, base)
        elif title.startswith(("腿肉分工厂-", "胸肉分工厂-", "其他分工厂-")):
            base = 12 if col_idx <= 2 else 10
            width = max(width or 0, base)
        elif title.endswith("明细") or title.endswith("汇总") or title == "原料规格":
            width = max(width or 0, 12)

        if width is not None:
            letter = get_column_letter(col_idx)
            current = ws.column_dimensions[letter].width or 0
            ws.column_dimensions[letter].width = max(current, width)


def _export_cell_number_format(header_text: str, pct_cols=None):
    pct_cols = set(pct_cols or [])
    header_text = str(header_text or "").strip()
    header_key = _norm_header_key(header_text)
    text_header_keys = {
        "工厂",
        "月份",
        "产品族",
        "使用半成品规格",
        "原料号",
        "原料描述",
        "鲜冻",
        "原料规格",
        "对应部位",
        "半成品",
        "半成品描述",
        "规格",
        "是否纳入综合影响",
        "归类规格",
        "物料号",
    }
    if not header_text:
        return None
    if header_text in pct_cols:
        return "0%"
    if header_key in text_header_keys:
        return None
    return '#,##0_);[Red](#,##0)'


def _style_export_sheet(ws, header_row: int = 1, freeze_cell=None, pct_cols=None):
    title_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    pct_cols = set(pct_cols or [])
    display_sheet = _is_display_export_sheet(ws.title)
    left_header_keys = {
        "工厂",
        "月份",
        "产品族",
        "使用半成品规格",
        "原料号",
        "原料描述",
        "鲜冻",
        "原料规格",
        "对应部位",
        "半成品",
        "半成品描述",
        "规格",
        "是否纳入综合影响",
        "归类规格",
        "物料号",
    }
    header_text_by_col = {}

    if freeze_cell:
        ws.freeze_panes = freeze_cell

    if header_row > 1 and ws["A1"].value not in (None, ""):
        if ws.max_column > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        title_cell = ws["A1"]
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center
        ws.row_dimensions[1].height = 24

    ws.row_dimensions[header_row].height = 22
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col_idx)
        header_text = str(cell.value or "").strip()
        header_text_by_col[col_idx] = header_text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            header_text = header_text_by_col.get(col_idx, "")
            header_key = _norm_header_key(header_text)
            if display_sheet and cell.value in (None, ""):
                cell.value = "-"
            cell.alignment = center if display_sheet or header_key not in left_header_keys else left
            if isinstance(cell.value, (int, float)):
                fmt = _export_cell_number_format(header_text, pct_cols=pct_cols)
                if fmt:
                    cell.number_format = fmt

    _autosize_columns(ws, start_row=header_row, end_row=min(ws.max_row, 500), min_width=8, max_width=38)
    _apply_export_column_widths(ws, header_row=header_row)


def write_export_sheet(writer, sheet_name: str, df: pd.DataFrame, title: str | None = None, pct_cols=None):
    startrow = 2 if title else 0
    df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow)
    ws = writer.book[sheet_name]
    if title:
        ws["A1"] = title
    header_row = startrow + 1
    freeze_cell = f"A{header_row + 1}" if ws.max_row > header_row else None
    _style_export_sheet(ws, header_row=header_row, freeze_cell=freeze_cell, pct_cols=pct_cols)


def export_calculated_workbook(
    output_stream,
    month_label: str,
    quarter_label: str,
    baseline_label: str,
    leg_records,
    bre_records,
    other_records,
    leg_s1: pd.DataFrame,
    bre_s1: pd.DataFrame,
    other_s1: pd.DataFrame,
    leg_s2: pd.DataFrame,
    bre_s2: pd.DataFrame,
    other_s2: pd.DataFrame,
    leg_s3: pd.DataFrame,
    bre_s3: pd.DataFrame,
    other_s3: pd.DataFrame,
    leg_audit: pd.DataFrame,
    bre_audit: pd.DataFrame,
    other_audit: pd.DataFrame,
    material_spec_profile=None,
):
    leg_detail, leg_market_col = _build_kind_detail_df(leg_audit, month_label, quarter_label)
    bre_detail, bre_market_col = _build_kind_detail_df(bre_audit, month_label, quarter_label)
    other_detail, other_market_col = _build_kind_detail_df(other_audit, month_label, quarter_label)
    leg_grouped = _build_grouped_kind_summary(leg_detail, month_label, quarter_label, leg_market_col)
    bre_grouped = _build_grouped_kind_summary(bre_detail, month_label, quarter_label, bre_market_col)
    other_grouped = _build_grouped_kind_summary(other_detail, month_label, quarter_label, other_market_col)
    leg_detail_export = _prepare_detail_export_df(leg_detail, quarter_label, baseline_label, leg_market_col, "腿肉")
    bre_detail_export = _prepare_detail_export_df(bre_detail, quarter_label, baseline_label, bre_market_col, "胸肉")
    other_detail_export = _prepare_detail_export_df(other_detail, quarter_label, baseline_label, other_market_col, "其他")
    leg_grouped_export = _rename_q_avg_qty_column(leg_grouped, quarter_label, baseline_label)
    bre_grouped_export = _rename_q_avg_qty_column(bre_grouped, quarter_label, baseline_label)
    other_grouped_export = _rename_q_avg_qty_column(other_grouped, quarter_label, baseline_label)
    leg_audit_export = _rename_q_avg_qty_column(leg_audit, quarter_label, baseline_label)
    bre_audit_export = _rename_q_avg_qty_column(bre_audit, quarter_label, baseline_label)
    other_audit_export = _rename_q_avg_qty_column(other_audit, quarter_label, baseline_label)
    leg_parts = _build_part_detail_df(leg_records)
    bre_parts = _build_part_detail_df(bre_records)
    other_parts = _build_part_detail_df(other_records)
    leg_month_spec, leg_q_spec = _build_spec_amounts(
        leg_records,
        _leg_spec_group,
        kind="腿肉",
        material_spec_profile=material_spec_profile,
    )
    bre_month_spec, bre_q_spec = _build_spec_amounts(
        bre_records,
        _bre_spec_group,
        kind="胸肉",
        material_spec_profile=material_spec_profile,
    )
    other_month_spec, other_q_spec = _build_spec_amounts(
        other_records,
        _other_spec_group,
        kind="其他",
        material_spec_profile=material_spec_profile,
    )
    leg_plant_amount_df = _build_spec_amount_export_df(
        leg_month_spec,
        leg_q_spec,
        _spec_headers_for_export("腿肉", material_spec_profile),
        _plant_slots_for_kind("\u817f\u8089"),
        month_label,
        baseline_label,
    )
    bre_plant_amount_df = _build_spec_amount_export_df(
        bre_month_spec,
        bre_q_spec,
        _spec_headers_for_export("胸肉", material_spec_profile),
        _plant_slots_for_kind("\u80f8\u8089"),
        month_label,
        baseline_label,
    )
    other_plant_amount_df = _build_spec_amount_export_df(
        other_month_spec,
        other_q_spec,
        _spec_headers_for_export("其他", material_spec_profile),
        _plant_slots_for_kind("其他"),
        month_label,
        baseline_label,
    )
    material_specs = _build_material_spec_df(leg_records, bre_records, other_records, material_spec_profile=material_spec_profile)
    overview_df = build_overview_sheet_df([("腿肉", leg_s1), ("胸肉", bre_s1), ("其他", other_s1)])
    leg_summary_export = _rename_q_avg_qty_column(leg_s1, quarter_label, baseline_label)
    bre_summary_export = _rename_q_avg_qty_column(bre_s1, quarter_label, baseline_label)
    other_summary_export = _rename_q_avg_qty_column(other_s1, quarter_label, baseline_label)
    leg_ratio_pct_cols = ["冻品", "鲜品", "合计"]
    bre_ratio_pct_cols = ["冻品", "鲜品", "合计"]
    other_ratio_pct_cols = ["冻品", "鲜品", "合计"]
    leg_plant_pct_cols = [c for c in leg_s3.columns if c not in ("工厂", "月份")]
    bre_plant_pct_cols = [c for c in bre_s3.columns if c not in ("工厂", "月份")]
    other_plant_pct_cols = [c for c in other_s3.columns if c not in ("工厂", "月份")]
    leg_ratio_df = _rename_period_label_value(leg_s2, quarter_label, baseline_label)
    bre_ratio_df = _rename_period_label_value(bre_s2, quarter_label, baseline_label)
    other_ratio_df = _rename_period_label_value(other_s2, quarter_label, baseline_label)
    leg_plant_df = _rename_period_label_value(leg_s3, quarter_label, baseline_label)
    bre_plant_df = _rename_period_label_value(bre_s3, quarter_label, baseline_label)
    other_plant_df = _rename_period_label_value(other_s3, quarter_label, baseline_label)

    with pd.ExcelWriter(output_stream, engine="openpyxl") as writer:
        write_export_sheet(writer, "Sheet1", overview_df, title=f"系统成本的综合影响（{month_label}月VS{baseline_label}）")
        write_export_sheet(writer, "腿肉", leg_summary_export, title=f"腿肉综合影响（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "胸肉", bre_summary_export, title=f"胸肉综合影响（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "其他", other_summary_export, title=f"其他综合影响（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "腿肉明细", leg_detail_export, title=f"腿肉明细（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "胸肉明细", bre_detail_export, title=f"胸肉明细（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "其他明细", other_detail_export, title=f"其他明细（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "腿肉产品族汇总", leg_grouped_export, title=f"腿肉产品族汇总（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "胸肉产品族汇总", bre_grouped_export, title=f"胸肉产品族汇总（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "其他产品族汇总", other_grouped_export, title=f"其他产品族汇总（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, f"腿肉占比-{month_label}", leg_ratio_df, title=f"腿肉鲜冻品占比（{month_label}月 vs {baseline_label}）", pct_cols=leg_ratio_pct_cols)
        write_export_sheet(writer, f"胸肉占比-{month_label}", bre_ratio_df, title=f"胸肉鲜冻品占比（{month_label}月 vs {baseline_label}）", pct_cols=bre_ratio_pct_cols)
        write_export_sheet(writer, f"其他占比-{month_label}", other_ratio_df, title=f"其他鲜冻品占比（{month_label}月 vs {baseline_label}）", pct_cols=other_ratio_pct_cols)
        write_export_sheet(writer, f"腿肉分工厂-{baseline_label}", leg_plant_df, title=f"腿肉分工厂占比（{month_label}月 vs {baseline_label}）", pct_cols=leg_plant_pct_cols)
        write_export_sheet(writer, f"胸肉分工厂-{baseline_label}", bre_plant_df, title=f"胸肉分工厂占比（{month_label}月 vs {baseline_label}）", pct_cols=bre_plant_pct_cols)
        write_export_sheet(writer, f"其他分工厂-{baseline_label}", other_plant_df, title=f"其他分工厂占比（{month_label}月 vs {baseline_label}）", pct_cols=other_plant_pct_cols)
        write_export_sheet(writer, f"\u817f\u8089\u5206\u5de5\u5382\u5e95\u8868-{baseline_label}", leg_plant_amount_df, title=f"\u817f\u8089\u5206\u5de5\u5382\u5e95\u8868\uff08{month_label}\u6708 vs {baseline_label}\uff09")
        write_export_sheet(writer, f"\u80f8\u8089\u5206\u5de5\u5382\u5e95\u8868-{baseline_label}", bre_plant_amount_df, title=f"\u80f8\u8089\u5206\u5de5\u5382\u5e95\u8868\uff08{month_label}\u6708 vs {baseline_label}\uff09")
        write_export_sheet(writer, f"其他分工厂底表-{baseline_label}", other_plant_amount_df, title=f"其他分工厂底表（{month_label}月 vs {baseline_label}）")
        write_export_sheet(writer, "腿肉核对", leg_audit_export, title="腿肉核对明细")
        write_export_sheet(writer, "胸肉核对", bre_audit_export, title="胸肉核对明细")
        write_export_sheet(writer, "其他核对", other_audit_export, title="其他核对明细")
        write_export_sheet(writer, "腿肉原料明细", leg_parts, title=f"腿肉原料明细（{month_label}月）")
        write_export_sheet(writer, "胸肉原料明细", bre_parts, title=f"胸肉原料明细（{month_label}月）")
        write_export_sheet(writer, "其他原料明细", other_parts, title=f"其他原料明细（{month_label}月）")
        write_export_sheet(writer, "原料规格", material_specs, title="原料规格映射")

        style_summary_sheet(writer.book["腿肉"], header_row=3)
        style_summary_sheet(writer.book["胸肉"], header_row=3)
        style_summary_sheet(writer.book["其他"], header_row=3)


        visible_sheet_names = {
            "腿肉",
            "胸肉",
            "其他",
            f"腿肉占比-{month_label}",
            f"胸肉占比-{month_label}",
            f"其他占比-{month_label}",
            f"腿肉分工厂-{baseline_label}",
            f"胸肉分工厂-{baseline_label}",
            f"其他分工厂-{baseline_label}",
        }
        for ws in writer.book.worksheets:
            ws.sheet_state = "visible" if ws.title in visible_sheet_names else "hidden"
        if "腿肉" in writer.book.sheetnames:
            writer.book.active = writer.book.sheetnames.index("腿肉")


def _input_name(source):
    if hasattr(source, "name"):
        return str(source.name)
    return Path(str(source)).name


def _display_baseline_label(q_files, quarter_label: str) -> str:
    names = " ".join(_input_name(f) for f in q_files)
    if "25Q2-Q4" in names or "Q2-Q4" in names or "25年" in names:
        return "25年"
    return quarter_label


def _month_text(month_label: str) -> str:
    return f"{month_label}月"


def _safe_float(value) -> float:
    num = to_num(value)
    return float(num) if num is not None else 0.0


def _display_total_cost_as_market(row: dict, market_col: str) -> float:
    market_value = to_num((row or {}).get(market_col))
    if market_value is not None:
        return float(market_value)
    return _safe_float((row or {}).get("半成品总成本"))


def _clear_range(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    max_row = min(max_row, ws.max_row)
    max_col = min(max_col, ws.max_column)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _set_cell(ws, coord: str, value):
    if not isinstance(ws[coord], MergedCell):
        ws[coord] = value


def _norm_header_key(value) -> str:
    return str(value or "").strip().replace(" ", "").lower()


def _sort_plants(df: pd.DataFrame, order: list[str]) -> pd.DataFrame:
    if df.empty or "工厂" not in df.columns:
        return df
    rank = {name: idx for idx, name in enumerate(order)}
    out = df.copy()
    out["_rank"] = out["工厂"].map(lambda x: rank.get(str(x), len(rank)))
    out = out.sort_values(by=["_rank"], kind="stable").drop(columns="_rank")
    return out


def _build_kind_detail_df(audit_df: pd.DataFrame, month_label: str, quarter_label: str):
    month_qty_col = f"{month_label}月产量"
    q_qty_col = f"{quarter_label}月均产量"
    market_col = "行情差异" if "行情差异" in audit_df.columns else "行情影响"
    if audit_df.empty:
        columns = [
            "工厂",
            "产品族",
            "使用半成品规格",
            month_qty_col,
            q_qty_col,
            "原料采购单价影响",
            "修形利用率影响",
            "损耗率影响",
            "半成品原料成本",
            "修形人工成本影响",
            "半成品总成本",
            market_col,
            "综合影响",
        ]
        return pd.DataFrame(columns=columns), market_col

    detail = audit_df.copy()
    detail = detail[detail["是否纳入综合影响"].astype(str).str.strip() == "是"].copy()
    numeric_cols = [
        "原料采购单价影响",
        "修形利用率影响",
        "损耗率影响",
        "修形人工成本影响",
        "综合影响",
        month_qty_col,
        q_qty_col,
        market_col,
    ]
    for col in numeric_cols:
        if col in detail.columns:
            detail[col] = pd.to_numeric(detail[col], errors="coerce").fillna(0.0)
    detail["半成品原料成本"] = (
        detail["原料采购单价影响"].fillna(0.0)
        + detail["修形利用率影响"].fillna(0.0)
        + detail["损耗率影响"].fillna(0.0)
    )
    detail["半成品总成本"] = detail["半成品原料成本"] + detail["修形人工成本影响"].fillna(0.0)
    cols = [
        "工厂",
        "产品族",
        "使用半成品规格",
        month_qty_col,
        q_qty_col,
        "原料采购单价影响",
        "修形利用率影响",
        "损耗率影响",
        "半成品原料成本",
        "修形人工成本影响",
        "半成品总成本",
        market_col,
        "综合影响",
    ]
    detail = detail.reindex(columns=cols)
    if not detail.empty:
        detail["_impact_abs"] = detail["综合影响"].abs()
        detail = detail.sort_values(
            by=["工厂", month_qty_col, "_impact_abs"],
            ascending=[True, False, False],
            kind="stable",
        ).drop(columns="_impact_abs")
    return detail, market_col


def _build_grouped_kind_summary(detail_df: pd.DataFrame, month_label: str, quarter_label: str, market_col: str):
    month_qty_col = f"{month_label}月产量"
    q_qty_col = f"{quarter_label}月均产量"
    if detail_df.empty:
        return pd.DataFrame(
            columns=[
                "产品族",
                "使用半成品规格",
                "原料采购单价影响",
                "修形利用率影响",
                "损耗率影响",
                "修形人工成本影响",
                "半成品总成本",
                market_col,
                month_qty_col,
                q_qty_col,
            ]
        )
    grouped = (
        detail_df.groupby(["产品族", "使用半成品规格"], as_index=False, dropna=False)[
            [
                "原料采购单价影响",
                "修形利用率影响",
                "损耗率影响",
                "修形人工成本影响",
                "半成品总成本",
                market_col,
                month_qty_col,
                q_qty_col,
            ]
        ]
        .sum()
    )
    grouped["_sort"] = grouped["半成品总成本"].abs()
    grouped = grouped.sort_values(by=["_sort", month_qty_col], ascending=[False, False], kind="stable").drop(
        columns="_sort"
    )
    return grouped.reset_index(drop=True)


def _build_part_detail_df(records):
    rows = []
    for rec in records:
        plant = rec.get("plant", "")
        code_to_spec = rec.get("code_to_spec") or {}
        for part in rec.get("month_part_rows") or []:
            raw_code = norm_code(part.get("原料号"))
            qty = _safe_float(part.get("数量"))
            if not raw_code or qty <= 0:
                continue
            rows.append(
                {
                    "工厂": plant,
                    "原料号": raw_code,
                    "原料描述": str(part.get("原料描述") or "").strip(),
                    "数量": qty,
                    "鲜冻": str(part.get("鲜冻") or "").strip(),
                    "规格": str(code_to_spec.get(raw_code, "") or "").strip(),
                }
            )
    if not rows:
        return pd.DataFrame(columns=["工厂", "原料号", "原料描述", "数量", "鲜冻", "规格"])
    out = pd.DataFrame(rows)
    out = out.sort_values(by=["工厂", "数量"], ascending=[True, False], kind="stable").reset_index(drop=True)
    return out


def _base_spec_text(spec_text: str) -> str:
    spec = str(spec_text or "").strip()
    if not spec:
        return ""
    return spec.split("/", 1)[0].strip()


def _load_material_spec_profile(source):
    if source is None:
        return None

    stream = io.BytesIO(source.getvalue()) if hasattr(source, "getvalue") else source
    xls = pd.ExcelFile(stream)
    sheet_name = "原料规格" if "原料规格" in xls.sheet_names else xls.sheet_names[-1]
    df = xls.parse(sheet_name)
    if df is None or df.empty:
        raise ValueError("上传的原料规格文件没有可用数据。")

    cols = list(df.columns)

    def pick_col(*names, fallback=None):
        for n in names:
            for c in cols:
                if str(c).strip() == n:
                    return c
        if fallback is not None and 0 <= fallback < len(cols):
            return cols[fallback]
        return None

    c_kind = pick_col("分类", fallback=0)
    c_code = pick_col("物料号", fallback=1)
    c_desc = pick_col("原料描述", fallback=2)
    c_spec = pick_col("规格", fallback=3)

    rows = []
    supported_kinds = ("腿肉", "胸肉", "其他")
    headers = {kind: [] for kind in supported_kinds}
    lookup = {kind: {} for kind in supported_kinds}
    for _, row in df.iterrows():
        kind = str(row.get(c_kind) or "").strip()
        if kind not in supported_kinds:
            continue
        raw_code = norm_code(row.get(c_code))
        spec = str(row.get(c_spec) or "").strip()
        if not raw_code or not spec:
            continue
        desc = "" if c_desc is None or pd.isna(row.get(c_desc)) else str(row.get(c_desc)).strip()
        if raw_code not in lookup[kind]:
            lookup[kind][raw_code] = spec
        if spec not in headers[kind]:
            headers[kind].append(spec)
        rows.append({"分类": kind, "物料号": raw_code, "原料描述": desc, "规格": spec})

    if not rows:
        raise ValueError("上传的原料规格文件未识别到腿肉、胸肉或其他物料规格。")

    clean_df = pd.DataFrame(rows).drop_duplicates(subset=["分类", "物料号"], keep="first")
    clean_df = clean_df.sort_values(by=["分类", "物料号"], kind="stable").reset_index(drop=True)
    return {"df": clean_df, "lookup": lookup, "headers": headers}



def _filter_parts_by_material_spec(kind: str, parts, material_spec_profile=None):
    parts = list(parts or [])
    if kind == "其他":
        allowed_mats = {norm_code(part.get("修行后原料")) for part in parts if norm_code(part.get("修行后原料"))}
        return parts, allowed_mats

    if not material_spec_profile:
        allowed_mats = {norm_code(part.get("修行后原料")) for part in parts if norm_code(part.get("修行后原料"))}
        return parts, allowed_mats

    # 规则：
    # 1) 白名单匹配对象是“展开后的原料号”，不是半成品号；
    # 2) 命中白名单的原料行保留，未命中的原料行过滤；
    # 3) 同一半成品下面只要命中任意 1 个白名单原料号，该半成品就允许纳入后续综合影响汇总。
    kind_lookup = (material_spec_profile.get("lookup") or {}).get(kind) or {}
    allowed_raw_codes = {norm_code(code) for code in kind_lookup.keys() if norm_code(code)}

    filtered = []
    matched_mats = set()
    for part in parts:
        raw_code = norm_code(part.get("原料号"))
        if raw_code not in allowed_raw_codes:
            continue
        filtered.append(part)
        mat = norm_code(part.get("修行后原料"))
        if mat:
            matched_mats.add(mat)

    return filtered, matched_mats

def _load_market_price_profile(source):
    if source is None:
        return {"df": pd.DataFrame(columns=MARKET_PRICE_TEMPLATE_COLUMNS), "lookup": {}}

    stream = io.BytesIO(source.getvalue()) if hasattr(source, "getvalue") else source
    xls = pd.ExcelFile(stream)
    sheet_name = MARKET_PRICE_TEMPLATE_SHEET if MARKET_PRICE_TEMPLATE_SHEET in xls.sheet_names else xls.sheet_names[0]
    df = xls.parse(sheet_name)
    if df is None or df.empty:
        return {"df": pd.DataFrame(columns=MARKET_PRICE_TEMPLATE_COLUMNS), "lookup": {}}

    cols = list(df.columns)

    def pick_col(*names, fallback=None):
        for name in names:
            c = find_col(cols, name)
            if c is not None:
                return c
        if fallback is not None and 0 <= fallback < len(cols):
            return cols[fallback]
        return None

    c_plant = pick_col("工厂", fallback=0)
    c_kind = pick_col("分类", fallback=1)
    c_current = pick_col("当前行情价", "当前价格", "当前价", fallback=2)
    c_previous = pick_col("基期行情价", "基期价格", "基期价", fallback=3)
    c_note = pick_col("备注", fallback=4)

    rows = []
    lookup = {}
    for _, row in df.iterrows():
        plant_raw = row.get(c_plant) if c_plant is not None else ""
        plant = "" if pd.isna(plant_raw) else normalize_plant_code(plant_raw)
        kind_raw = row.get(c_kind) if c_kind is not None else ""
        kind = "" if pd.isna(kind_raw) else str(kind_raw).strip()
        if kind not in ("腿肉", "胸肉", "其他") or not plant:
            continue
        current_pre = to_num(row.get(c_current)) if c_current is not None else None
        previous_pre = to_num(row.get(c_previous)) if c_previous is not None else None
        if current_pre is None and previous_pre is None:
            continue
        note = "" if c_note is None or pd.isna(row.get(c_note)) else str(row.get(c_note)).strip()
        payload = {"current_pre": current_pre, "previous_pre": previous_pre, "备注": note}
        lookup[(plant, kind)] = payload
        rows.append({"工厂": plant, "分类": kind, "当前行情价": current_pre, "基期行情价": previous_pre, "备注": note})

    clean_df = pd.DataFrame(rows, columns=MARKET_PRICE_TEMPLATE_COLUMNS)
    if not clean_df.empty:
        clean_df = clean_df.drop_duplicates(subset=["工厂", "分类"], keep="last").reset_index(drop=True)
    return {"df": clean_df, "lookup": lookup}


def _market_price_override_for(plant_code: str, kind: str, market_price_profile=None):
    lookup = (market_price_profile or {}).get("lookup") or {}
    return lookup.get((normalize_plant_code(plant_code), kind), {})


def _ordered_actual_rows(rows, quarter_label=None):
    actual_rows = [r for r in (rows or []) if _is_actual_price_row_type(r)]
    current_row = next((r for r in actual_rows if _is_current_actual_price_row_type(r)), None)
    if current_row is None:
        current_row = actual_rows[0] if actual_rows else None
    previous_candidates = [r for r in actual_rows if r is not current_row]
    ranked_previous = [
        (_baseline_actual_rank(r, quarter_label), idx, r)
        for idx, r in enumerate(previous_candidates)
    ]
    ranked_previous = [item for item in ranked_previous if item[0] is not None]
    previous_row = sorted(ranked_previous, key=lambda item: (item[0], item[1]))[0][2] if ranked_previous else None
    if previous_row is None:
        previous_row = previous_candidates[0] if previous_candidates else None
    return current_row, previous_row


def _build_override_market_impact_map(kind: str, month_rows, plant_code: str, market_price_profile=None):
    override = _market_price_override_for(plant_code, kind, market_price_profile)
    current_pre_override = to_num(override.get("current_pre"))
    previous_pre_override = to_num(override.get("previous_pre"))
    if current_pre_override is None and previous_pre_override is None:
        return {}

    by_mat = defaultdict(list)
    for row in month_rows or []:
        mat = norm_code((row or {}).get("修行后原料"))
        if mat:
            by_mat[mat].append(row)

    out = {}
    for mat, rows in by_mat.items():
        current_row, previous_row = _ordered_actual_rows(rows)
        current_vals = _actual_metric_values(current_row, kind)
        previous_vals = _actual_metric_values(previous_row, kind)
        cur_pre = current_pre_override if current_pre_override is not None else to_num(current_vals.get("pre"))
        prev_pre = previous_pre_override if previous_pre_override is not None else to_num(previous_vals.get("pre"))
        cur_util = to_num(current_vals.get("util"))
        cur_loss = to_num(current_vals.get("loss"))
        qty = to_num(current_vals.get("qty"))
        if cur_pre is None or prev_pre is None or cur_util in (None, 0) or cur_loss is None:
            continue
        current_raw = _calc_raw_cost(kind, cur_pre, cur_util, cur_loss)
        baseline_raw = _calc_raw_cost(kind, prev_pre, cur_util, cur_loss)
        if current_raw is None or baseline_raw is None:
            continue
        market_unit = current_raw - baseline_raw
        market_total = market_unit * qty if qty not in (None, 0) else None
        base_row = current_row or previous_row or {}
        out[mat] = {
            "sheet_name": MARKET_PRICE_TEMPLATE_SHEET,
            "product_family": base_row.get("产品族", ""),
            "spec_name": base_row.get("使用半成品规格", ""),
            "current_pre": cur_pre,
            "previous_pre": prev_pre,
            "market_unit_impact": market_unit,
            "market_total_impact": market_total if market_total is not None else 0.0,
        }
    return out


def _resolve_spec_bucket(kind: str, raw_code: str, spec_text: str, raw_desc: str, spec_group_fn, material_spec_profile=None):
    if material_spec_profile:
        kind_lookup = (material_spec_profile.get("lookup") or {}).get(kind, {})
        mapped_spec = str(kind_lookup.get(raw_code) or "").strip()
        if not mapped_spec:
            return None
        return spec_group_fn(mapped_spec, raw_desc)
    return spec_group_fn(spec_text, raw_desc)


def _spec_headers_for_export(kind: str, material_spec_profile=None):
    base_headers = _spec_headers_for_kind(kind)
    if not material_spec_profile:
        return base_headers

    spec_group_fn = _spec_group_fn_for_kind(kind)
    raw_headers = (material_spec_profile.get("headers") or {}).get(kind) or []
    resolved_headers = []
    for raw_header in raw_headers:
        header = spec_group_fn(str(raw_header or "").strip(), "")
        if header and header not in resolved_headers:
            resolved_headers.append(header)

    extras = [header for header in resolved_headers if header not in base_headers]
    return list(base_headers) + extras


def _build_material_spec_df(leg_records, bre_records, other_records=None, material_spec_profile=None):
    if material_spec_profile:
        profile_df = material_spec_profile.get("df")
        if isinstance(profile_df, pd.DataFrame):
            return profile_df.copy()
        return pd.DataFrame(columns=["分类", "物料号", "原料描述", "规格"])

    rows = []
    for kind, records in (("腿肉", leg_records), ("胸肉", bre_records), ("其他", other_records or [])):
        desc_map = {}
        spec_map = {}
        for rec in records:
            spec_map.update(rec.get("code_to_spec") or {})
            for part in (rec.get("month_part_rows") or []) + (rec.get("q_part_rows") or []):
                raw_code = norm_code(part.get("原料号"))
                if raw_code and raw_code not in desc_map:
                    desc_map[raw_code] = str(part.get("原料描述") or "").strip()
        for raw_code, spec in spec_map.items():
            rows.append(
                {
                    "分类": kind,
                    "物料号": raw_code,
                    "原料描述": desc_map.get(raw_code, ""),
                    "规格": _base_spec_text(spec) or str(spec or "").strip(),
                }
            )
    if not rows:
        return pd.DataFrame(columns=["分类", "物料号", "原料描述", "规格"])
    out = pd.DataFrame(rows).drop_duplicates(subset=["分类", "物料号"], keep="first")
    return out.sort_values(by=["分类", "物料号"], kind="stable").reset_index(drop=True)


def _build_fresh_frozen_amounts(records):
    month_amounts = defaultdict(lambda: {"冻品": 0.0, "鲜品": 0.0})
    q_amounts = defaultdict(lambda: {"冻品": 0.0, "鲜品": 0.0})
    for rec in records:
        plant = rec.get("plant", "")
        for part in rec.get("month_part_rows") or []:
            kind = str(part.get("鲜冻") or "").strip()
            if kind in ("冻品", "鲜品"):
                month_amounts[plant][kind] += _safe_float(part.get("数量")) / 1000.0
        for part in rec.get("q_part_rows") or []:
            kind = str(part.get("鲜冻") or "").strip()
            if kind in ("冻品", "鲜品"):
                q_amounts[plant][kind] += _safe_float(part.get("数量")) / 1000.0

    def with_total(source):
        total = {"冻品": 0.0, "鲜品": 0.0}
        out = {plant: dict(values) for plant, values in source.items()}
        for values in out.values():
            total["冻品"] += values.get("冻品", 0.0)
            total["鲜品"] += values.get("鲜品", 0.0)
        out["合计"] = total
        return out

    return with_total(month_amounts), with_total(q_amounts)


def _build_spec_amount_export_df(
    month_amounts: dict,
    q_amounts: dict,
    headers: list[str],
    plants: list[str],
    month_label: str,
    baseline_label: str,
):
    columns = ["期间", "工厂", *headers, "合计"]
    rows = []

    def append_period(period_label: str, amount_map: dict):
        for plant in [*plants, "合计"]:
            values = amount_map.get(plant, {})
            row = {"期间": period_label, "工厂": plant}
            total = 0.0
            for header in headers:
                value = _safe_float(values.get(header))
                row[header] = value
                total += value
            row["合计"] = total
            rows.append(row)

    append_period(_month_text(month_label), month_amounts)
    rows.append({key: None for key in columns})
    append_period(baseline_label, q_amounts)

    df = pd.DataFrame(rows, columns=columns)
    num_cols = [*headers, "合计"]
    for col in num_cols:
        df[col] = df[col].map(lambda v: round(v, 3) if isinstance(v, (int, float)) and pd.notna(v) else v)
    return df


def _spec_group_fn_for_kind(kind: str):
    if kind == "腿肉":
        return _leg_spec_group
    if kind == "胸肉":
        return _bre_spec_group
    return _other_spec_group


def _spec_headers_for_kind(kind: str):
    if kind == "腿肉":
        return ["无规格", "80g", "110G", "120g以上", "120-170g", "170-220", "200-300"]
    if kind == "胸肉":
        return ["无规格", "120G以上", "200g", "220g以上", "260g", "170-220g", "220-260g", "200-300", "220-300", "260-300g", "300g"]
    return ["无规格"]


def _plant_slots_for_kind(kind: str):
    if kind == "腿肉":
        return ["BB1", "BB2", "LY", "DL"]
    if kind == "胸肉":
        return ["BB1", "BB2", "TJ", "LY", "YZ"]
    return ["BB1", "BB2", "TJ", "LY", "YZ", "DL"]


def _leg_spec_group(spec_text: str, raw_desc: str = ""):
    base = _norm_header_key(_base_spec_text(spec_text) or raw_desc)
    if not base:
        return None
    if "无规格" in base:
        return "无规格"
    if "80g" in base:
        return "80g"
    if "110g" in base:
        return "110G"
    if "120-170" in base:
        return "120-170g"
    if "170-220" in base:
        return "170-220"
    if "245-285" in base or "200-300" in base:
        return "200-300"
    if "120g" in base:
        return "120g以上"
    return None


def _bre_spec_group(spec_text: str, raw_desc: str = ""):
    base = _norm_header_key(_base_spec_text(spec_text) or raw_desc)
    if not base:
        return None
    if "无规格" in base or "内采" in base:
        return "无规格"
    if "170-220" in base:
        return "170-220g"
    if "220-260" in base:
        return "220-260g"
    if "220-300" in base:
        return "220-300"
    if "260-300" in base:
        return "260-300g"
    if "200-300" in base:
        return "200-300"
    if "300g" in base:
        return "300g"
    if "260g" in base:
        return "260g"
    if "220g" in base:
        return "220g以上"
    if "200g" in base:
        return "200g"
    if "120" in base:
        return "120G以上"
    return None


def _other_spec_group(spec_text: str, raw_desc: str = ""):
    base = _base_spec_text(spec_text) or _base_spec_text(raw_desc)
    base = str(base or "").strip()
    if not base:
        return None
    if "无规格" in base or "内采" in base:
        return "无规格"
    return base


def _build_spec_amounts(records, spec_group_fn, *, kind=None, material_spec_profile=None):
    month_amounts = defaultdict(lambda: defaultdict(float))
    q_amounts = defaultdict(lambda: defaultdict(float))
    for rec in records:
        plant = rec.get("plant", "")
        code_to_spec = rec.get("code_to_spec") or {}
        for bucket, source in ((month_amounts, rec.get("month_part_rows") or []), (q_amounts, rec.get("q_part_rows") or [])):
            for part in source:
                raw_code = norm_code(part.get("原料号"))
                spec_text = code_to_spec.get(raw_code, "")
                header = _resolve_spec_bucket(
                    kind or "",
                    raw_code,
                    spec_text,
                    str(part.get("原料描述") or ""),
                    spec_group_fn,
                    material_spec_profile,
                )
                if not header:
                    continue
                bucket[plant][header] += _safe_float(part.get("数量")) / 1000.0

    def with_total(source):
        total = defaultdict(float)
        out = {plant: dict(values) for plant, values in source.items()}
        for values in out.values():
            for header, value in values.items():
                total[header] += value
        out["合计"] = dict(total)
        return out

    return with_total(month_amounts), with_total(q_amounts)


def _ratio_from_amounts(amount_map, headers):
    out = {}
    for plant, values in amount_map.items():
        total = sum(values.get(header, 0.0) for header in headers)
        ratios = {header: (values.get(header, 0.0) / total if total else 0.0) for header in headers}
        ratios["合计"] = 1.0 if total else 0.0
        out[plant] = ratios
    return out


def _clear_rows_for_ratio_comparison(ws, start_row: int, plant_count: int, start_col: int, end_col: int):
    end_row = start_row + plant_count * 3 - 1
    _clear_range(ws, start_row, end_row, start_col, end_col)


def _fill_material_spec_sheet(ws, material_df: pd.DataFrame):
    _clear_range(ws, 2, max(ws.max_row, 200), 1, 4)
    for idx, row in material_df.iterrows():
        r = idx + 2
        ws[f"A{r}"] = row["分类"]
        ws[f"B{r}"] = row["物料号"]
        ws[f"C{r}"] = row["原料描述"]
        ws[f"D{r}"] = row["规格"]


def _fill_sheet1_summary(ws, leg_s1: pd.DataFrame, bre_s1: pd.DataFrame, month_label: str, baseline_label: str):
    def pick_total(df):
        if df.empty:
            return {}
        total_rows = df[df["工厂"].astype(str) == "合计"]
        return total_rows.iloc[0].to_dict() if not total_rows.empty else {}

    leg_total = pick_total(leg_s1)
    bre_total = pick_total(bre_s1)
    _set_cell(ws, "A2", f"系统成本的综合影响（{month_label}月VS{baseline_label}）")

    for row_idx, total in ((5, leg_total), (6, bre_total)):
        ws[f"B{row_idx}"] = total.get("扣除行情后采购绩效")
        ws[f"C{row_idx}"] = total.get("修形利用率影响")
        ws[f"D{row_idx}"] = total.get("损耗率影响")
        ws[f"E{row_idx}"] = total.get("修形人工成本影响")
        ws[f"F{row_idx}"] = 0
        ws[f"G{row_idx}"] = sum(
            _safe_float(total.get(col))
            for col in ("扣除行情后采购绩效", "修形利用率影响", "损耗率影响", "修形人工成本影响")
        )

    for row_idx in (7, 8):
        for col in ("B", "C", "D", "E", "F", "G"):
            ws[f"{col}{row_idx}"] = 0

    for col in ("B", "C", "D", "E", "F", "G"):
        ws[f"{col}9"] = _safe_float(ws[f"{col}5"].value) + _safe_float(ws[f"{col}6"].value)


def _ratio_tuple(values: dict):
    frozen = _safe_float((values or {}).get("冻品"))
    fresh = _safe_float((values or {}).get("鲜品"))
    total = frozen + fresh
    if total > 0:
        return frozen, fresh, total, frozen / total, fresh / total, 1.0
    return frozen, fresh, total, None, None, None


def _diff_ratio(month_ratio, q_ratio):
    if month_ratio is None and q_ratio is None:
        return None
    return (month_ratio or 0.0) - (q_ratio or 0.0)


def _fill_kind_sheet(
    ws,
    detail_df: pd.DataFrame,
    grouped_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    month_label: str,
    quarter_label: str,
    baseline_label: str,
    kind_label: str,
):
    month_qty_col = f"{month_label}月产量"
    q_qty_col = f"{quarter_label}月均产量"
    market_col = "行情差异" if "行情差异" in summary_df.columns else "行情影响"

    _set_cell(ws, "W1", f"主原料系统成本的综合影响（{_month_text(month_label)}&{baseline_label}）-{kind_label}")
    _set_cell(ws, "AQ1", f"主原料系统成本的综合影响（{_month_text(month_label)}&{baseline_label}）-{kind_label}")
    _set_cell(ws, "AD2", f"{_month_text(month_label)}行情-{baseline_label}行情")
    _set_cell(ws, "D4", f"{month_label}月产量")
    _set_cell(ws, "E4", f"{baseline_label}月均产量")
    _set_cell(ws, "AE2", f"{month_label}月产量")
    _set_cell(ws, "AF2", f"{baseline_label}月均产量")
    _set_cell(ws, "AY2", f"{month_label}月产量")
    _set_cell(ws, "AZ2", f"{baseline_label}月均产量")
    if kind_label == "胸肉":
        _set_cell(ws, "M4", f"{baseline_label}行情差异")
    else:
        _set_cell(ws, "M4", "行情差异")

    _clear_range(ws, 5, max(ws.max_row, 200), 1, 13)
    _clear_range(ws, 3, max(ws.max_row, 200), 14, 22)
    _clear_range(ws, 3, max(ws.max_row, 200), 23, 32)
    _clear_range(ws, 3, max(ws.max_row, 200), 43, 54)

    detail_capacity = max(0, ws.max_row - 4)
    detail_rows = detail_df.head(detail_capacity).to_dict("records")
    for idx, row in enumerate(detail_rows, start=5):
        price = _safe_float(row.get("原料采购单价影响"))
        util = _safe_float(row.get("修形利用率影响"))
        loss = _safe_float(row.get("损耗率影响"))
        raw_cost = _safe_float(row.get("半成品原料成本"))
        labor = _safe_float(row.get("修形人工成本影响"))
        market = _safe_float(row.get(market_col))
        total = _display_total_cost_as_market(row, market_col)
        ws[f"A{idx}"] = row.get("工厂")
        ws[f"B{idx}"] = row.get("产品族")
        ws[f"C{idx}"] = row.get("使用半成品规格")
        ws[f"D{idx}"] = row.get(month_qty_col)
        ws[f"E{idx}"] = row.get(q_qty_col)
        ws[f"F{idx}"] = price
        ws[f"G{idx}"] = util
        ws[f"H{idx}"] = loss
        ws[f"I{idx}"] = raw_cost
        ws[f"J{idx}"] = labor
        ws[f"K{idx}"] = total
        ws[f"L{idx}"] = None
        ws[f"M{idx}"] = market

    group_capacity = max(0, ws.max_row - 2)
    grouped_rows = grouped_df.head(group_capacity).to_dict("records")
    for idx, row in enumerate(grouped_rows, start=3):
        price = _safe_float(row.get("原料采购单价影响"))
        util = _safe_float(row.get("修形利用率影响"))
        loss = _safe_float(row.get("损耗率影响"))
        labor = _safe_float(row.get("修形人工成本影响"))
        market = _safe_float(row.get(market_col))
        total = _display_total_cost_as_market(row, market_col)
        month_qty = _safe_float(row.get(month_qty_col))
        q_qty = _safe_float(row.get(q_qty_col))
        _set_cell(ws, f"N{idx}", row.get("产品族"))
        _set_cell(ws, f"O{idx}", row.get("使用半成品规格"))
        _set_cell(ws, f"P{idx}", price)
        _set_cell(ws, f"Q{idx}", util)
        _set_cell(ws, f"R{idx}", loss)
        _set_cell(ws, f"S{idx}", labor)
        _set_cell(ws, f"T{idx}", total)
        _set_cell(ws, f"U{idx}", market)
        _set_cell(ws, f"V{idx}", month_qty)
        _set_cell(ws, f"W{idx}", row.get("产品族"))
        _set_cell(ws, f"X{idx}", row.get("使用半成品规格"))
        _set_cell(ws, f"Y{idx}", price)
        _set_cell(ws, f"Z{idx}", util)
        _set_cell(ws, f"AA{idx}", loss)
        _set_cell(ws, f"AB{idx}", labor)
        _set_cell(ws, f"AC{idx}", total)
        _set_cell(ws, f"AD{idx}", market)
        _set_cell(ws, f"AE{idx}", month_qty)
        _set_cell(ws, f"AF{idx}", q_qty)

    summary_capacity = max(0, ws.max_row - 2)
    summary_rows = summary_df.head(summary_capacity).to_dict("records")
    for idx, row in enumerate(summary_rows, start=3):
        _set_cell(ws, f"AQ{idx}", row.get("工厂"))
        _set_cell(ws, f"AR{idx}", row.get("原料采购单价影响"))
        _set_cell(ws, f"AS{idx}", row.get(market_col))
        _set_cell(ws, f"AT{idx}", row.get("扣除行情后采购绩效"))
        _set_cell(ws, f"AU{idx}", row.get("修形利用率影响"))
        _set_cell(ws, f"AV{idx}", row.get("损耗率影响"))
        _set_cell(ws, f"AW{idx}", row.get("修形人工成本影响"))
        _set_cell(ws, f"AX{idx}", row.get("综合影响"))
        _set_cell(ws, f"AY{idx}", row.get(month_qty_col))
        _set_cell(ws, f"AZ{idx}", row.get(q_qty_col))


def _sorted_part_rows(part_df: pd.DataFrame, plant_slots: list[str]) -> pd.DataFrame:
    raw = part_df.copy()
    if raw.empty:
        return raw
    plant_rank = {name: idx for idx, name in enumerate(plant_slots)}
    raw["_rank"] = raw["工厂"].map(lambda x: plant_rank.get(str(x), len(plant_rank)))
    return raw.sort_values(by=["_rank", "数量"], ascending=[True, False], kind="stable").drop(columns="_rank")


def _write_ratio_compare_rows(
    ws,
    start_row: int,
    plant_col: str,
    label_col: str,
    value_cols: tuple[str, str, str],
    plant: str,
    month_label_text: str,
    baseline_label: str,
    month_values: tuple[float | None, float | None, float | None],
    q_values: tuple[float | None, float | None, float | None],
):
    ws[f"{plant_col}{start_row}"] = plant
    ws[f"{label_col}{start_row}"] = month_label_text
    ws[f"{label_col}{start_row + 1}"] = baseline_label
    ws[f"{label_col}{start_row + 2}"] = "差异"
    for col, month_value, q_value in zip(value_cols, month_values, q_values):
        ws[f"{col}{start_row}"] = month_value
        ws[f"{col}{start_row + 1}"] = q_value
        ws[f"{col}{start_row + 2}"] = _diff_ratio(month_value, q_value)


def _write_spec_compare_rows(
    ws,
    start_row: int,
    plant_col: str,
    label_col: str,
    value_start_col: int,
    total_col: str,
    headers: list[str],
    plant: str,
    month_label_text: str,
    baseline_label: str,
    month_ratio: dict,
    q_ratio: dict,
):
    ws[f"{plant_col}{start_row}"] = plant
    ws[f"{label_col}{start_row}"] = month_label_text
    ws[f"{label_col}{start_row + 1}"] = baseline_label
    ws[f"{label_col}{start_row + 2}"] = "差异"
    for col_idx, header in enumerate(headers, start=value_start_col):
        month_value = month_ratio.get(header)
        q_value = q_ratio.get(header)
        ws.cell(start_row, col_idx).value = month_value
        ws.cell(start_row + 1, col_idx).value = q_value
        ws.cell(start_row + 2, col_idx).value = _diff_ratio(month_value, q_value)
    ws[f"{total_col}{start_row}"] = month_ratio.get("合计")
    ws[f"{total_col}{start_row + 1}"] = q_ratio.get("合计")
    ws[f"{total_col}{start_row + 2}"] = _diff_ratio(month_ratio.get("合计"), q_ratio.get("合计"))


def _fill_leg_ratio_sheet(ws, part_df: pd.DataFrame, month_amounts: dict, q_amounts: dict, month_label: str, baseline_label: str):
    plant_slots = ["BB2", "BB1", "LY", "DL", "合计"]
    month_label_text = _month_text(month_label)

    _set_cell(ws, "K1", month_label_text)
    _set_cell(ws, "K2", baseline_label)
    _clear_range(ws, 3, max(ws.max_row, 200), 2, 21)

    raw = _sorted_part_rows(part_df, plant_slots)
    raw_capacity = max(0, ws.max_row - 2)
    for idx, row in enumerate(raw.head(raw_capacity).to_dict("records"), start=3):
        qty = _safe_float(row.get("数量"))
        ws[f"B{idx}"] = row.get("工厂")
        ws[f"C{idx}"] = row.get("原料号")
        ws[f"D{idx}"] = row.get("原料描述")
        ws[f"E{idx}"] = qty
        ws[f"F{idx}"] = qty if qty > 0 else None
        ws[f"G{idx}"] = qty
        ws[f"H{idx}"] = row.get("鲜冻")
        ws[f"I{idx}"] = qty if qty > 0 else None
        ws[f"J{idx}"] = _base_spec_text(row.get("规格"))

    for idx, plant in enumerate(plant_slots):
        month_f, month_x, month_total, month_ratio_f, month_ratio_x, month_ratio_total = _ratio_tuple(month_amounts.get(plant, {}))
        q_f, q_x, q_total, q_ratio_f, q_ratio_x, q_ratio_total = _ratio_tuple(q_amounts.get(plant, {}))
        amount_row = 4 + idx * 2
        ratio_row = amount_row + 1
        ws[f"K{amount_row}"] = plant
        ws[f"L{amount_row}"] = month_f if month_total else None
        ws[f"M{amount_row}"] = month_x if month_total else None
        ws[f"N{amount_row}"] = month_total if month_total else None
        ws[f"K{ratio_row}"] = "鲜冻品占比"
        ws[f"L{ratio_row}"] = month_ratio_f
        ws[f"M{ratio_row}"] = month_ratio_x
        ws[f"N{ratio_row}"] = month_ratio_total

        summary_row = 4 + idx
        ws[f"P{summary_row}"] = plant
        ws[f"Q{summary_row}"] = month_ratio_f
        ws[f"R{summary_row}"] = month_ratio_x
        ws[f"S{summary_row}"] = month_ratio_total

        month_values = (month_ratio_f, month_ratio_x, month_ratio_total)
        q_values = (q_ratio_f, q_ratio_x, q_ratio_total)
        _write_ratio_compare_rows(ws, 16 + idx * 3, "K", "L", ("M", "N", "O"), plant, month_label_text, baseline_label, month_values, q_values)
        _write_ratio_compare_rows(ws, 15 + idx * 3, "Q", "R", ("S", "T", "U"), plant, month_label_text, baseline_label, month_values, q_values)


def _fill_bre_ratio_sheet(ws, part_df: pd.DataFrame, month_amounts: dict, q_amounts: dict, month_label: str, baseline_label: str):
    plant_slots = ["BB2", "BB1", "LY", "TJ", "YZ", "合计"]
    month_label_text = _month_text(month_label)

    _clear_range(ws, 4, max(ws.max_row, 250), 2, 24)

    raw = _sorted_part_rows(part_df, plant_slots)
    raw_capacity = max(0, ws.max_row - 3)
    for idx, row in enumerate(raw.head(raw_capacity).to_dict("records"), start=4):
        qty = _safe_float(row.get("数量"))
        ws[f"B{idx}"] = row.get("工厂")
        ws[f"C{idx}"] = row.get("原料号")
        ws[f"D{idx}"] = row.get("原料描述")
        ws[f"E{idx}"] = qty
        ws[f"F{idx}"] = row.get("鲜冻")
        ws[f"G{idx}"] = qty if qty > 0 else None
        ws[f"H{idx}"] = _base_spec_text(row.get("规格"))

    for idx, plant in enumerate(plant_slots):
        month_f, month_x, month_total, month_ratio_f, month_ratio_x, month_ratio_total = _ratio_tuple(month_amounts.get(plant, {}))
        q_f, q_x, q_total, q_ratio_f, q_ratio_x, q_ratio_total = _ratio_tuple(q_amounts.get(plant, {}))
        amount_row = 4 + idx * 2
        ratio_row = amount_row + 1
        ws[f"I{amount_row}"] = plant
        ws[f"J{amount_row}"] = month_f if month_total else None
        ws[f"K{amount_row}"] = month_x if month_total else None
        ws[f"L{amount_row}"] = month_total if month_total else None
        ws[f"I{ratio_row}"] = "鲜冻品占比"
        ws[f"J{ratio_row}"] = month_ratio_f
        ws[f"K{ratio_row}"] = month_ratio_x
        ws[f"L{ratio_row}"] = month_ratio_total

        month_values = (month_ratio_f, month_ratio_x, month_ratio_total)
        q_values = (q_ratio_f, q_ratio_x, q_ratio_total)
        _write_ratio_compare_rows(ws, 4 + idx * 3, "N", "O", ("P", "Q", "R"), plant, month_label_text, baseline_label, month_values, q_values)
        _write_ratio_compare_rows(ws, 4 + idx * 3, "T", "U", ("V", "W", "X"), plant, month_label_text, baseline_label, month_values, q_values)


def _fill_leg_plant_sheet(ws, month_spec_amounts: dict, q_spec_amounts: dict, month_label: str, baseline_label: str):
    headers = ["无规格", "80g", "110G", "120g以上", "120-170g", "170-220", "200-300"]
    plant_slots = ["BB1", "BB2", "LY", "DL", "合计"]
    month_ratios = _ratio_from_amounts(month_spec_amounts, headers)
    q_ratios = _ratio_from_amounts(q_spec_amounts, headers)
    month_label_text = _month_text(month_label)

    _set_cell(ws, "L1", month_label_text)
    _set_cell(ws, "N1", baseline_label)
    _clear_range(ws, 3, max(ws.max_row, 100), 2, 30)

    for idx, plant in enumerate(plant_slots, start=3):
        values = month_spec_amounts.get(plant, {})
        ws[f"B{idx}"] = plant
        total = 0.0
        for offset, header in enumerate(headers, start=3):
            value = _safe_float(values.get(header))
            ws.cell(idx, offset).value = value if value else None
            total += value
        ws[f"J{idx}"] = total if total else None

    for slot_idx, plant in enumerate(plant_slots):
        month_ratio = month_ratios.get(plant, {})
        q_ratio = q_ratios.get(plant, {})
        start = 3 + slot_idx * 3
        _write_spec_compare_rows(ws, start, "L", "M", 14, "T", headers, plant, month_label_text, baseline_label, month_ratio, q_ratio)
        _write_spec_compare_rows(ws, start, "V", "W", 24, "AD", headers, plant, month_label_text, baseline_label, month_ratio, q_ratio)


def _fill_bre_plant_sheet(ws, month_spec_amounts: dict, q_spec_amounts: dict, month_label: str, baseline_label: str):
    headers = ["无规格", "120G以上", "200g", "220g以上", "260g", "170-220g", "220-260g", "200-300", "220-300", "260-300g", "300g"]
    plant_slots = ["BB1", "BB2", "TJ", "LY", "YZ", "合计"]
    month_ratios = _ratio_from_amounts(month_spec_amounts, headers)
    q_ratios = _ratio_from_amounts(q_spec_amounts, headers)
    month_label_text = _month_text(month_label)

    _clear_range(ws, 3, max(ws.max_row, 100), 2, 44)

    for idx, plant in enumerate(plant_slots, start=3):
        values = month_spec_amounts.get(plant, {})
        ws[f"B{idx}"] = plant
        total = 0.0
        for offset, header in enumerate(headers, start=3):
            value = _safe_float(values.get(header))
            ws.cell(idx, offset).value = value if value else None
            total += value
        ws[f"N{idx}"] = total if total else None

    for slot_idx, plant in enumerate(plant_slots):
        month_ratio = month_ratios.get(plant, {})
        q_ratio = q_ratios.get(plant, {})
        start = 3 + slot_idx * 3
        _write_spec_compare_rows(ws, start, "P", "Q", 18, "AC", headers, plant, month_label_text, baseline_label, month_ratio, q_ratio)
        _write_spec_compare_rows(ws, start, "AE", "AF", 33, "AR", headers, plant, month_label_text, baseline_label, month_ratio, q_ratio)


def _ensure_sheet(wb, sheet_name: str, *, index=None):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(title=sheet_name, index=index)


def export_template_workbook(
    template_path,
    month_label: str,
    quarter_label: str,
    baseline_label: str,
    leg_records,
    bre_records,
    leg_s1: pd.DataFrame,
    bre_s1: pd.DataFrame,
    leg_audit: pd.DataFrame,
    bre_audit: pd.DataFrame,
    material_spec_profile=None,
):
    wb = load_workbook(template_path, keep_links=False)
    if hasattr(wb, "_external_links"):
        wb._external_links = []
    leg_detail, leg_market_col = _build_kind_detail_df(leg_audit, month_label, quarter_label)
    bre_detail, bre_market_col = _build_kind_detail_df(bre_audit, month_label, quarter_label)
    leg_grouped = _build_grouped_kind_summary(leg_detail, month_label, quarter_label, leg_market_col)
    bre_grouped = _build_grouped_kind_summary(bre_detail, month_label, quarter_label, bre_market_col)

    leg_parts = _build_part_detail_df(leg_records)
    bre_parts = _build_part_detail_df(bre_records)
    leg_month_fs, leg_q_fs = _build_fresh_frozen_amounts(leg_records)
    bre_month_fs, bre_q_fs = _build_fresh_frozen_amounts(bre_records)
    leg_month_spec, leg_q_spec = _build_spec_amounts(
        leg_records,
        _leg_spec_group,
        kind="腿肉",
        material_spec_profile=material_spec_profile,
    )
    bre_month_spec, bre_q_spec = _build_spec_amounts(
        bre_records,
        _bre_spec_group,
        kind="胸肉",
        material_spec_profile=material_spec_profile,
    )
    material_specs = _build_material_spec_df(leg_records, bre_records, material_spec_profile=material_spec_profile)

    _fill_sheet1_summary(_ensure_sheet(wb, "Sheet1", index=0), leg_s1, bre_s1, month_label, baseline_label)
    _fill_kind_sheet(wb["腿肉"], leg_detail, leg_grouped, leg_s1, month_label, quarter_label, baseline_label, "腿肉")
    _fill_kind_sheet(wb["胸肉"], bre_detail, bre_grouped, bre_s1, month_label, quarter_label, baseline_label, "胸肉")
    _fill_leg_ratio_sheet(wb["腿肉占比-1"], leg_parts, leg_month_fs, leg_q_fs, month_label, baseline_label)
    _fill_bre_ratio_sheet(wb["胸肉占比-12"], bre_parts, bre_month_fs, bre_q_fs, month_label, baseline_label)
    _fill_leg_plant_sheet(wb["腿肉分工厂-Q3"], leg_month_spec, leg_q_spec, month_label, baseline_label)
    _fill_bre_plant_sheet(wb["胸肉分工厂-Q3"], bre_month_spec, bre_q_spec, month_label, baseline_label)
    _fill_material_spec_sheet(wb["原料规格"], material_specs)
    _apply_output_layout(wb)
    return wb



if SKIP_STREAMLIT_UI:
    month_files = []
    q_files = []
    material_spec_file = None
    market_price_file = None
else:
    month_files = st.file_uploader("上传 x月系统成本文件", type=["xlsx"], accept_multiple_files=True)
    q_files = st.file_uploader("上传 Qx系统成本文件", type=["xlsx"], accept_multiple_files=True)
    material_spec_file = st.file_uploader("上传原料规格文件", type=["xlsx"], accept_multiple_files=False)
    market_price_file = st.file_uploader("上传行情价覆盖文件（可选）", type=["xlsx"], accept_multiple_files=False)


if month_files and q_files:
    _clear_source_caches()
    run_started = datetime.now()
    run_timer = perf_counter()
    run_meta = {
        "run_id": run_started.strftime("%Y%m%d_%H%M%S"),
        "started_at": run_started.isoformat(timespec="seconds"),
        "status": "running",
        "month_files": [f.name for f in month_files],
        "q_files": [f.name for f in q_files],
    }
    if material_spec_file is not None:
        run_meta["material_spec_file"] = material_spec_file.name
    if market_price_file is not None:
        run_meta["market_price_file"] = market_price_file.name

    try:
        if not SKIP_STREAMLIT_UI:
            st.info("处理中，请等待 1-3 分钟...")
        month_label = month_from_name(month_files[0].name)
        quarter_label = quarter_from_files(q_files)

        month_map = {plant_from_name(f.name): f for f in month_files}
        q_map = {plant_from_name(f.name): f for f in q_files}
        plants = sorted(set(month_map.keys()) & set(q_map.keys()))
        plant_codes = [normalize_plant_code(p) for p in plants]
        run_meta["month_plant_count"] = len(month_map)
        run_meta["q_plant_count"] = len(q_map)
        run_meta["matched_plant_count"] = len(plants)
        run_meta["matched_plants"] = plant_codes
        if not plants:
            raise ValueError("月文件与Q文件未匹配到相同工厂。")

        if material_spec_file is None:
            raise ValueError("请上传原料规格文件，物料规格将严格按该文件计算。")
        material_spec_profile = _load_material_spec_profile(material_spec_file)
        market_price_profile = _load_market_price_profile(market_price_file)
        run_meta["market_price_override_rows"] = len((market_price_profile or {}).get("df", pd.DataFrame()))

        leg_records = []
        bre_records = []
        other_records = []
        for plant in plants:
            plant_code = normalize_plant_code(plant)
            mf, qf = month_map[plant], q_map[plant]
            mx, qx = pd.ExcelFile(mf), pd.ExcelFile(qf)

            m_leg_tsc = read_tsc_df(mx, "腿肉")
            m_bre_tsc = read_tsc_df(mx, "胸肉")
            m_other_tsc = read_tsc_df(mx, "其他")
            q_leg_tsc = read_tsc_df(qx, "腿肉")
            q_bre_tsc = read_tsc_df(qx, "胸肉")
            q_other_tsc = read_tsc_df(qx, "其他")

            m_leg_part = read_part_df(mx, "腿肉")
            m_bre_part = read_part_df(mx, "胸肉")
            m_other_part = read_part_df(mx, "其他")
            q_leg_part = read_part_df(qx, "腿肉")
            q_bre_part = read_part_df(qx, "胸肉")
            q_other_part = read_part_df(qx, "其他")

            if m_leg_tsc is not None or q_leg_tsc is not None:
                m_rows_raw, m_map_code = parse_tsc(m_leg_tsc) if m_leg_tsc is not None else ([], {})
                q_rows_raw, q_map_code = parse_tsc(q_leg_tsc) if q_leg_tsc is not None else ([], {})
                m_rows = _replace_impact_rows(m_rows_raw, _extract_total_impact_rows(mf, "腿肉"))
                q_rows = _replace_impact_rows(q_rows_raw, _extract_total_impact_rows(qf, "腿肉"))
                month_rows = m_rows if m_rows else q_rows
                code_map = dict(m_map_code)
                code_map.update(q_map_code)
                leg_market_map = dict(_extract_kind_market_impact_map("腿肉", mf, qf))
                leg_market_map.update(
                    _build_override_market_impact_map(
                        "腿肉",
                        m_rows if _has_month_actual_rows(m_rows) else month_rows,
                        plant_code,
                        market_price_profile=market_price_profile,
                    )
                )
                leg_month_parts = parse_part(m_leg_part) if m_leg_part is not None else parse_part(q_leg_part)
                leg_q_parts = parse_part(q_leg_part)
                leg_month_parts, leg_month_mats = _filter_parts_by_material_spec("腿肉", leg_month_parts, material_spec_profile)
                leg_q_parts, leg_q_mats = _filter_parts_by_material_spec("腿肉", leg_q_parts, material_spec_profile)
                leg_records.append(
                    {
                        "plant": plant_code,
                        "source_month_tsc_rows": m_rows,
                        "month_tsc_rows": month_rows,
                        "q_tsc_rows": q_rows,
                        "code_to_spec": code_map,
                        "month_part_rows": leg_month_parts,
                        "q_part_rows": leg_q_parts,
                        "allowed_mats": leg_month_mats | leg_q_mats,
                        "q_label": quarter_from_name(qf.name),
                        "market_impact_map": leg_market_map,
                        "fallback_actual_map": _extract_manual_actual_map(qf, "腿肉"),
                        "product_family_map": _extract_product_family_map(qf, "腿肉"),
                    }
                )

            if m_bre_tsc is not None or q_bre_tsc is not None:
                m_rows_raw, m_map_code = parse_tsc(m_bre_tsc) if m_bre_tsc is not None else ([], {})
                q_rows_raw, q_map_code = parse_tsc(q_bre_tsc) if q_bre_tsc is not None else ([], {})
                m_rows = _replace_impact_rows(m_rows_raw, _extract_total_impact_rows(mf, "胸肉"))
                q_rows = _replace_impact_rows(q_rows_raw, _extract_total_impact_rows(qf, "胸肉"))
                month_rows = m_rows if m_rows else q_rows
                code_map = dict(m_map_code)
                code_map.update(q_map_code)
                bre_market_map = dict(_extract_kind_market_impact_map("胸肉", mf, qf))
                bre_market_map.update(
                    _build_override_market_impact_map(
                        "胸肉",
                        m_rows if _has_month_actual_rows(m_rows) else month_rows,
                        plant_code,
                        market_price_profile=market_price_profile,
                    )
                )
                bre_month_parts = parse_part(m_bre_part) if m_bre_part is not None else parse_part(q_bre_part)
                bre_q_parts = parse_part(q_bre_part)
                bre_month_parts, bre_month_mats = _filter_parts_by_material_spec("胸肉", bre_month_parts, material_spec_profile)
                bre_q_parts, bre_q_mats = _filter_parts_by_material_spec("胸肉", bre_q_parts, material_spec_profile)
                bre_records.append(
                    {
                        "plant": plant_code,
                        "source_month_tsc_rows": m_rows,
                        "month_tsc_rows": month_rows,
                        "q_tsc_rows": q_rows,
                        "code_to_spec": code_map,
                        "month_part_rows": bre_month_parts,
                        "q_part_rows": bre_q_parts,
                        "allowed_mats": bre_month_mats | bre_q_mats,
                        "q_label": quarter_from_name(qf.name),
                        "market_impact_map": bre_market_map,
                        "fallback_actual_map": _extract_manual_actual_map(qf, "胸肉"),
                        "product_family_map": _extract_product_family_map(qf, "胸肉"),
                    }
                )

            if m_other_tsc is not None or q_other_tsc is not None:
                m_rows_raw, m_map_code = parse_tsc(m_other_tsc) if m_other_tsc is not None else ([], {})
                q_rows_raw, q_map_code = parse_tsc(q_other_tsc) if q_other_tsc is not None else ([], {})
                m_rows = _replace_impact_rows(m_rows_raw, _extract_total_impact_rows(mf, "其他"))
                q_rows = _replace_impact_rows(q_rows_raw, _extract_total_impact_rows(qf, "其他"))
                month_rows = m_rows if m_rows else q_rows
                code_map = dict(m_map_code)
                code_map.update(q_map_code)
                other_market_map = dict(_extract_market_impact_map(qf, "其他"))
                other_market_map.update(
                    _build_override_market_impact_map(
                        "其他",
                        m_rows if _has_month_actual_rows(m_rows) else month_rows,
                        plant_code,
                        market_price_profile=market_price_profile,
                    )
                )
                other_month_parts = parse_part(m_other_part) if m_other_part is not None else parse_part(q_other_part)
                other_q_parts = parse_part(q_other_part)
                other_month_parts, other_month_mats = _filter_parts_by_material_spec("其他", other_month_parts, material_spec_profile)
                other_q_parts, other_q_mats = _filter_parts_by_material_spec("其他", other_q_parts, material_spec_profile)
                other_records.append(
                    {
                        "plant": plant_code,
                        "source_month_tsc_rows": m_rows,
                        "month_tsc_rows": month_rows,
                        "q_tsc_rows": q_rows,
                        "code_to_spec": code_map,
                        "month_part_rows": other_month_parts,
                        "q_part_rows": other_q_parts,
                        "allowed_mats": other_month_mats | other_q_mats,
                        "q_label": quarter_from_name(qf.name),
                        "market_impact_map": other_market_map,
                        "fallback_actual_map": _extract_manual_actual_map(qf, "其他"),
                        "product_family_map": _extract_product_family_map(qf, "其他"),
                    }
                )

        _validate_part_material_coverage(leg_records, quarter_label, "腿肉")
        _validate_part_material_coverage(bre_records, quarter_label, "胸肉")
        _validate_part_material_coverage(other_records, quarter_label, "其他")

        leg_s1, leg_s2, leg_s3 = build_kind(
            leg_records,
            month_label,
            quarter_label,
            "腿肉",
            material_spec_profile=material_spec_profile,
        )
        bre_s1, bre_s2, bre_s3 = build_kind(
            bre_records,
            month_label,
            quarter_label,
            "胸肉",
            material_spec_profile=material_spec_profile,
        )
        other_s1, other_s2, other_s3 = build_kind(
            other_records,
            month_label,
            quarter_label,
            "其他",
            material_spec_profile=material_spec_profile,
        )
        leg_audit = build_audit_detail(leg_records, month_label, quarter_label, "腿肉")
        bre_audit = build_audit_detail(bre_records, month_label, quarter_label, "胸肉")
        other_audit = build_audit_detail(other_records, month_label, quarter_label, "其他")
        bre_s1 = bre_s1.rename(columns={"行情影响": "行情差异"})
        run_meta["leg_record_count"] = len(leg_records)
        run_meta["bre_record_count"] = len(bre_records)
        run_meta["other_record_count"] = len(other_records)
        run_meta["leg_rows"] = {"s1": int(len(leg_s1)), "s2": int(len(leg_s2)), "s3": int(len(leg_s3))}
        run_meta["bre_rows"] = {"s1": int(len(bre_s1)), "s2": int(len(bre_s2)), "s3": int(len(bre_s3))}
        run_meta["other_rows"] = {"s1": int(len(other_s1)), "s2": int(len(other_s2)), "s3": int(len(other_s3))}
        run_meta["audit_rows"] = {"leg": int(len(leg_audit)), "bre": int(len(bre_audit)), "other": int(len(other_audit))}

        baseline_label = _display_baseline_label(q_files, quarter_label)
        leg_ratio_view = _rename_period_label_value(
            present(leg_s2, pct_cols=["冻品", "鲜品", "合计"]),
            quarter_label,
            baseline_label,
        )
        bre_ratio_view = _rename_period_label_value(
            present(bre_s2, pct_cols=["冻品", "鲜品", "合计"]),
            quarter_label,
            baseline_label,
        )
        leg_plant_view = _rename_period_label_value(
            present(leg_s3, pct_cols=[c for c in leg_s3.columns if c not in ("工厂", "月份")]),
            quarter_label,
            baseline_label,
        )
        bre_plant_view = _rename_period_label_value(
            present(bre_s3, pct_cols=[c for c in bre_s3.columns if c not in ("工厂", "月份")]),
            quarter_label,
            baseline_label,
        )
        other_ratio_view = _rename_period_label_value(
            present(other_s2, pct_cols=["冻品", "鲜品", "合计"]),
            quarter_label,
            baseline_label,
        )
        other_plant_view = _rename_period_label_value(
            present(other_s3, pct_cols=[c for c in other_s3.columns if c not in ("工厂", "月份")]),
            quarter_label,
            baseline_label,
        )
        t1, t2, t3, t4, t5, t6, t7, t8, t9 = st.tabs(
            [
                f"腿肉-{month_label}月",
                f"胸肉-{month_label}月",
                f"其他-{month_label}月",
                f"腿肉占比-{month_label}",
                f"胸肉占比-{month_label}",
                f"其他占比-{month_label}",
                f"腿肉分工厂-{baseline_label}",
                f"胸肉分工厂-{baseline_label}",
                f"其他分工厂-{baseline_label}",
            ]
        )
        with t1:
            st.dataframe(_center_display(present_s1(leg_s1, month_label, quarter_label, baseline_label)), width="stretch")
        with t2:
            st.dataframe(_center_display(present_s1(bre_s1, month_label, quarter_label, baseline_label)), width="stretch")
        with t3:
            st.dataframe(_center_display(present_s1(other_s1, month_label, quarter_label, baseline_label)), width="stretch")
        with t4:
            st.dataframe(_center_display(leg_ratio_view), width="stretch")
        with t5:
            st.dataframe(_center_display(bre_ratio_view), width="stretch")
        with t6:
            st.dataframe(_center_display(other_ratio_view), width="stretch")
        with t7:
            st.dataframe(_center_display(leg_plant_view), width="stretch")
        with t8:
            st.dataframe(_center_display(bre_plant_view), width="stretch")
        with t9:
            st.dataframe(_center_display(other_plant_view), width="stretch")

        out = io.BytesIO()
        export_calculated_workbook(
            output_stream=out,
            month_label=month_label,
            quarter_label=quarter_label,
            baseline_label=baseline_label,
            leg_records=leg_records,
            bre_records=bre_records,
            other_records=other_records,
            leg_s1=leg_s1,
            bre_s1=bre_s1,
            other_s1=other_s1,
            leg_s2=leg_s2,
            bre_s2=bre_s2,
            other_s2=other_s2,
            leg_s3=leg_s3,
            bre_s3=bre_s3,
            other_s3=other_s3,
            leg_audit=leg_audit,
            bre_audit=bre_audit,
            other_audit=other_audit,
            material_spec_profile=material_spec_profile,
        )
        out_name = f"鲜冻品占比-{month_label}.xlsx"
        out_bytes = out.getvalue()

        st.download_button(
            "下载结果",
            data=out_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        run_meta["status"] = "success"
        run_meta["elapsed_sec"] = round(perf_counter() - run_timer, 3)
        run_meta["month_label"] = month_label
        run_meta["quarter_label"] = quarter_label
        append_run_log(run_meta)

        with st.expander("Run Monitor", expanded=False):
            st.json(run_meta)
    except Exception as e:
        run_meta["status"] = "failed"
        run_meta["elapsed_sec"] = round(perf_counter() - run_timer, 3)
        run_meta["error"] = str(e)
        try:
            append_run_log(run_meta)
        except Exception:
            pass
        st.error(f"处理失败: {e}")
    finally:
        _clear_source_caches()






























