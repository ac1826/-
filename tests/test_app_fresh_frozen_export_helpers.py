import importlib.util
import io
import os
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = next(path for path in ROOT.glob("app_*.py") if path.name != "app.py")

os.environ["FRESH_FROZEN_SKIP_UI"] = "1"
SPEC = importlib.util.spec_from_file_location("fresh_frozen_app_under_test", MODULE_PATH)
APP = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(APP)


class FreshFrozenExportHelperTests(unittest.TestCase):
    def _market_source_workbook(self, sheet_name: str, total_impact: float):
        buf = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append([""] * 14 + ["\u4fee\u5f62\u540e\u6210\u672c", "", "", "", "", "\u884c\u60c5\u4ef7", "\u884c\u60c5\u4ef7"])
        ws.append(
            [
                "\u4ea7\u54c1\u65cf",
                "\u4fee\u884c\u540e\u539f\u6599",
                "\u4f7f\u7528\u534a\u6210\u54c1\u89c4\u683c",
                "\u884c\u7c7b\u578b",
                "\u5f71\u54cd\u53e3\u5f84",
                "\u7efc\u5408\u5355\u4ef7",
                "\u4fee\u5f62\u524d\u539f\u6599\u7efc\u5408\u8017\u7528\u5355\u4ef7",
                "\u4fee\u5f62\u5229\u7528\u7387",
                "\u635f\u8017\u7387",
                "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c",
                "\u534a\u6210\u54c1\u4fee\u5f62\u4eba\u5de5\u6210\u672c",
                "\u534a\u6210\u54c1\u603b\u6210\u672c",
                "\u534a\u6210\u54c1\u5165\u5e93\u91cf",
                "BOM",
                "BOM\u5360\u6bd4",
                "\u57fa\u671f\u884c\u60c5\u4ef7",
                "\u5f53\u524d\u884c\u60c5\u4ef7",
            ]
        )
        ws.append([""] * 17)
        ws.append([""] * 17)
        ws.append(["A", "39000181", "\u89c4\u683c", "4\u6708\u5b9e\u9645\u5355\u4ef7", "", 10, 10, 0.6, 0.05, 12, 1, 13, 4, "", "", 9, 10])
        ws.append(["", "39000181", "", "25\u5e74Q4\u5b9e\u9645\u5355\u4ef7", "", 9, 9, 0.6, 0.05, 11, 1, 12, "", "", "", 9, 10])
        ws.append(["", "39000181", "", "4\u6708\u89c4\u683c\u5360\u6bd4", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.append(["", "39000181", "", "\u89c4\u683c\u5360\u6bd4", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.append(["", "39000181", "", "\u5dee\u5f02", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws.append(["", "39000181", "", "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd", "\u5355\u4f4d\u6210\u672c", "", 1, 0, 0, 1, 0, 1, "", "", "", "", ""])
        ws.append(["", "39000181", "", "\u5bf9\u534a\u6210\u54c1\u6210\u672c\u7684\u5f71\u54cd", "\u603b\u6210\u672c", "", total_impact, 0, 0, total_impact, 0, total_impact, "", "", "", "", ""])
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_leg_and_breast_market_maps_are_extracted_from_month_source(self):
        for kind, sheet_name in [
            ("\u817f\u8089", "\u817f\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6"),
            ("\u80f8\u8089", "\u80f8\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6"),
        ]:
            with self.subTest(kind=kind):
                month_source = self._market_source_workbook(sheet_name, 11.0)
                q_source = self._market_source_workbook(sheet_name, 99.0)

                market_map = APP._extract_kind_market_impact_map(kind, month_source, q_source)

                self.assertEqual(market_map["39000181"]["market_total_impact"], 11.0)

    def test_other_market_map_still_uses_quarter_source(self):
        month_source = self._market_source_workbook("\u5176\u4ed6\u884c\u60c5", 11.0)
        q_source = self._market_source_workbook("\u5176\u4ed6\u884c\u60c5", 99.0)

        market_map = APP._extract_kind_market_impact_map("\u5176\u4ed6", month_source, q_source)

        self.assertEqual(market_map["39000181"]["market_total_impact"], 99.0)

    def test_baseline_actual_rows_prefer_quarter_label_over_25_year_rows(self):
        md = {
            "25\u5e74\u5b9e\u9645\u4ef7\u683c": [{"\u884c\u7c7b\u578b": "25\u5e74\u5b9e\u9645\u4ef7\u683c", "\u7efc\u5408\u5355\u4ef7": 25.0}],
            "25Q4\u5b9e\u9645\u4ef7\u683c": [{"\u884c\u7c7b\u578b": "25Q4\u5b9e\u9645\u4ef7\u683c", "\u7efc\u5408\u5355\u4ef7": 4.0}],
        }
        _, previous = APP.select_actual_rows(md, {}, "Q4")
        _, qrow, _ = APP.select_material_rows(md, {}, "Q4")

        self.assertEqual(previous["\u7efc\u5408\u5355\u4ef7"], 4.0)
        self.assertEqual(qrow["\u7efc\u5408\u5355\u4ef7"], 4.0)

    def test_baseline_actual_rows_prefer_25q_unit_price_over_25_year_rows(self):
        md = {
            "25\u5e74\u5b9e\u9645\u5355\u4ef7": [{"\u884c\u7c7b\u578b": "25\u5e74\u5b9e\u9645\u5355\u4ef7", "\u7efc\u5408\u5355\u4ef7": 25.0}],
            "25Q4\u5b9e\u9645\u5355\u4ef7": [{"\u884c\u7c7b\u578b": "25Q4\u5b9e\u9645\u5355\u4ef7", "\u7efc\u5408\u5355\u4ef7": 4.0}],
        }
        _, previous = APP.select_actual_rows(md, {}, "Q4")
        _, qrow, _ = APP.select_material_rows(md, {}, "Q4")

        self.assertEqual(previous["\u7efc\u5408\u5355\u4ef7"], 4.0)
        self.assertEqual(qrow["\u7efc\u5408\u5355\u4ef7"], 4.0)

    def test_baseline_actual_rows_fall_back_to_any_q_before_25_year(self):
        md = {
            "25\u5e74\u5b9e\u9645\u4ef7\u683c": [{"\u884c\u7c7b\u578b": "25\u5e74\u5b9e\u9645\u4ef7\u683c", "\u7efc\u5408\u5355\u4ef7": 25.0}],
            "Q3\u5b9e\u9645\u4ef7\u683c": [{"\u884c\u7c7b\u578b": "Q3\u5b9e\u9645\u4ef7\u683c", "\u7efc\u5408\u5355\u4ef7": 3.0}],
        }
        _, previous = APP.select_actual_rows(md, {}, "Q4")
        _, qrow, _ = APP.select_material_rows(md, {}, "Q4")

        self.assertEqual(previous["\u7efc\u5408\u5355\u4ef7"], 3.0)
        self.assertEqual(qrow["\u7efc\u5408\u5355\u4ef7"], 3.0)

    def test_baseline_actual_rows_fall_back_to_25_year_when_quarter_missing(self):
        md = {
            "25\u5e74\u5b9e\u9645\u5355\u4ef7": [{"\u884c\u7c7b\u578b": "25\u5e74\u5b9e\u9645\u5355\u4ef7", "\u7efc\u5408\u5355\u4ef7": 25.0}],
        }
        _, previous = APP.select_actual_rows(md, {}, "Q4")
        _, qrow, _ = APP.select_material_rows(md, {}, "Q4")

        self.assertEqual(previous["\u7efc\u5408\u5355\u4ef7"], 25.0)
        self.assertEqual(qrow["\u7efc\u5408\u5355\u4ef7"], 25.0)

    def test_baseline_actual_rows_accept_price_and_unit_price_labels_from_quarter_source(self):
        for label in ("Q4\u5b9e\u9645\u4ef7\u683c", "Q4\u5b9e\u9645\u5355\u4ef7"):
            with self.subTest(label=label):
                qd = {
                    "25\u5e74\u5b9e\u9645\u4ef7\u683c": [{"\u884c\u7c7b\u578b": "25\u5e74\u5b9e\u9645\u4ef7\u683c", "\u7efc\u5408\u5355\u4ef7": 25.0}],
                    label: [{"\u884c\u7c7b\u578b": label, "\u7efc\u5408\u5355\u4ef7": 4.0}],
                }
                _, previous = APP.select_actual_rows({}, qd, "Q4")
                _, qrow, _ = APP.select_material_rows({}, qd, "Q4")

                self.assertEqual(previous["\u7efc\u5408\u5355\u4ef7"], 4.0)
                self.assertEqual(qrow["\u7efc\u5408\u5355\u4ef7"], 4.0)

    def test_ordered_actual_rows_uses_same_baseline_priority(self):
        rows = [
            {"\u884c\u7c7b\u578b": "4\u6708\u5b9e\u9645\u5355\u4ef7", "\u7efc\u5408\u5355\u4ef7": 10.0},
            {"\u884c\u7c7b\u578b": "25\u5e74\u5b9e\u9645\u4ef7\u683c", "\u7efc\u5408\u5355\u4ef7": 25.0},
            {"\u884c\u7c7b\u578b": "25Q4\u5b9e\u9645\u4ef7\u683c", "\u7efc\u5408\u5355\u4ef7": 4.0},
        ]

        current, previous = APP._ordered_actual_rows(rows, "Q4")

        self.assertEqual(current["\u7efc\u5408\u5355\u4ef7"], 10.0)
        self.assertEqual(previous["\u7efc\u5408\u5355\u4ef7"], 4.0)

    def test_leg_and_breast_market_sheet_names_must_match_exact_quarter_compare_sheet(self):
        self.assertEqual(
            APP._pick_market_sheet_name(
                ["\u817f\u8089\u884c\u60c5", "\u817f\u8089\u884c\u60c5-\u5b63\u5ea6"],
                "\u817f\u8089",
            ),
            None,
        )
        self.assertEqual(
            APP._pick_market_sheet_name(
                ["\u80f8\u8089\u884c\u60c5", "\u80f8\u8089\u884c\u60c5-\u5b63\u5ea6"],
                "\u80f8\u8089",
            ),
            None,
        )
        self.assertEqual(
            APP._pick_market_sheet_name(["\u817f\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6"], "\u817f\u8089"),
            "\u817f\u8089\u884c\u60c5-\u8f83\u5b63\u5ea6",
        )

    def test_leg_and_breast_detail_exports_frontload_market_impact(self):
        detail = pd.DataFrame(
            [
                {
                    "\u5de5\u5382": "BB2",
                    "\u4ea7\u54c1\u65cf": "\u7c73\u6751\u817f\u8089\u5757",
                    "\u4f7f\u7528\u534a\u6210\u54c1\u89c4\u683c": "\u9e21\u53bb\u76ae\u817f\u8089\u5757/7-15g/\u81ea\u4fee\u5f62",
                    "4\u6708\u4ea7\u91cf": 295.532,
                    "Q4\u6708\u5747\u4ea7\u91cf": 992.368,
                    "\u539f\u6599\u91c7\u8d2d\u5355\u4ef7\u5f71\u54cd": -142.102892472746,
                    "\u4fee\u5f62\u5229\u7528\u7387\u5f71\u54cd": -2.055105745941991,
                    "\u635f\u8017\u7387\u5f71\u54cd": -0.03307565489866227,
                    "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c": -144.1910738735866,
                    "\u4fee\u5f62\u4eba\u5de5\u6210\u672c\u5f71\u54cd": -3.792534525539016,
                    "\u534a\u6210\u54c1\u603b\u6210\u672c": -147.9836083991257,
                    "\u884c\u60c5\u5f71\u54cd": -160.0262507645254,
                    "\u7efc\u5408\u5f71\u54cd": 12.04264236539977,
                }
            ]
        )

        for kind in ("\u817f\u8089", "\u80f8\u8089"):
            with self.subTest(kind=kind):
                out = APP._prepare_detail_export_df(detail, "Q4", "Q4", "\u884c\u60c5\u5f71\u54cd", kind)

                self.assertEqual(list(out.columns)[5], "\u884c\u60c5\u5f71\u54cd")
                self.assertNotIn("\u539f\u6599\u91c7\u8d2d\u5355\u4ef7\u5f71\u54cd", out.columns)
                self.assertEqual(list(out.columns).count("\u884c\u60c5\u5f71\u54cd"), 1)
                self.assertAlmostEqual(out.iloc[0, 5], -160.0262507645254)
                self.assertAlmostEqual(out.iloc[0]["\u534a\u6210\u54c1\u603b\u6210\u672c"], -160.0262507645254)

    def test_breast_detail_export_uses_market_difference_for_total_cost_display(self):
        detail = pd.DataFrame(
            [
                {
                    "\u5de5\u5382": "BB2",
                    "\u4ea7\u54c1\u65cf": "\u4e1c\u65b9\u7504\u9009\u9e21\u6392",
                    "\u4f7f\u7528\u534a\u6210\u54c1\u89c4\u683c": "\u9e21\u5f00\u7247\u5927\u80f8/65-78g/\u81ea\u4fee\u5f62",
                    "4\u6708\u4ea7\u91cf": 487.01,
                    "Q4\u6708\u5747\u4ea7\u91cf": 258.07,
                    "\u539f\u6599\u91c7\u8d2d\u5355\u4ef7\u5f71\u54cd": 522.462019213437,
                    "\u4fee\u5f62\u5229\u7528\u7387\u5f71\u54cd": 108.338723671265,
                    "\u635f\u8017\u7387\u5f71\u54cd": -74.2096212803364,
                    "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c": 556.591121604366,
                    "\u4fee\u5f62\u4eba\u5de5\u6210\u672c\u5f71\u54cd": 36.0661343685807,
                    "\u534a\u6210\u54c1\u603b\u6210\u672c": 592.657255972947,
                    "\u884c\u60c5\u5dee\u5f02": 421.750260152972,
                    "\u7efc\u5408\u5f71\u54cd": 170.906995819975,
                }
            ]
        )

        out = APP._prepare_detail_export_df(detail, "Q4", "Q4", "\u884c\u60c5\u5dee\u5f02", "\u80f8\u8089")

        self.assertEqual(list(out.columns)[5], "\u884c\u60c5\u5dee\u5f02")
        self.assertAlmostEqual(out.iloc[0, 5], 421.750260152972)
        self.assertAlmostEqual(out.iloc[0]["\u534a\u6210\u54c1\u603b\u6210\u672c"], 421.750260152972)
        self.assertAlmostEqual(detail.iloc[0]["\u534a\u6210\u54c1\u603b\u6210\u672c"], 592.657255972947)

    def test_other_detail_export_keeps_purchase_impact_front_column(self):
        detail = pd.DataFrame(
            [
                {
                    "\u5de5\u5382": "BB2",
                    "\u4ea7\u54c1\u65cf": "\u5176\u4ed6",
                    "\u4f7f\u7528\u534a\u6210\u54c1\u89c4\u683c": "\u89c4\u683c",
                    "4\u6708\u4ea7\u91cf": 10,
                    "Q4\u6708\u5747\u4ea7\u91cf": 20,
                    "\u539f\u6599\u91c7\u8d2d\u5355\u4ef7\u5f71\u54cd": 12.0,
                    "\u534a\u6210\u54c1\u603b\u6210\u672c": 19.0,
                    "\u884c\u60c5\u5f71\u54cd": 7.0,
                    "\u7efc\u5408\u5f71\u54cd": 5.0,
                }
            ]
        )

        out = APP._prepare_detail_export_df(detail, "Q4", "Q4", "\u884c\u60c5\u5f71\u54cd", "\u5176\u4ed6")

        self.assertEqual(list(out.columns)[5], "\u539f\u6599\u91c7\u8d2d\u5355\u4ef7\u5f71\u54cd")
        self.assertEqual(out.iloc[0, 5], 12.0)
        self.assertIn("\u884c\u60c5\u5f71\u54cd", out.columns)
        self.assertAlmostEqual(out.iloc[0]["\u534a\u6210\u54c1\u603b\u6210\u672c"], 7.0)
        self.assertAlmostEqual(detail.iloc[0]["\u534a\u6210\u54c1\u603b\u6210\u672c"], 19.0)

    def test_template_kind_sheets_display_market_impact_in_total_cost_cells(self):
        for kind, market_col in [
            ("\u817f\u8089", "\u884c\u60c5\u5f71\u54cd"),
            ("\u80f8\u8089", "\u884c\u60c5\u5dee\u5f02"),
            ("\u5176\u4ed6", "\u884c\u60c5\u5f71\u54cd"),
        ]:
            with self.subTest(kind=kind):
                detail = pd.DataFrame(
                    [
                        {
                            "\u5de5\u5382": "BB2",
                            "\u4ea7\u54c1\u65cf": "\u4e1c\u65b9\u7504\u9009\u9e21\u6392",
                            "\u4f7f\u7528\u534a\u6210\u54c1\u89c4\u683c": "\u9e21\u5f00\u7247\u5927\u80f8/65-78g/\u81ea\u4fee\u5f62",
                            "4\u6708\u4ea7\u91cf": 487.01,
                            "Q4\u6708\u5747\u4ea7\u91cf": 258.07,
                            "\u539f\u6599\u91c7\u8d2d\u5355\u4ef7\u5f71\u54cd": 522.462019213437,
                            "\u4fee\u5f62\u5229\u7528\u7387\u5f71\u54cd": 108.338723671265,
                            "\u635f\u8017\u7387\u5f71\u54cd": -74.2096212803364,
                            "\u534a\u6210\u54c1\u539f\u6599\u6210\u672c": 556.591121604366,
                            "\u4fee\u5f62\u4eba\u5de5\u6210\u672c\u5f71\u54cd": 36.0661343685807,
                            "\u534a\u6210\u54c1\u603b\u6210\u672c": 592.657255972947,
                            market_col: 421.750260152972,
                            "\u7efc\u5408\u5f71\u54cd": 170.906995819975,
                        }
                    ]
                )
                grouped = pd.DataFrame(
                    [
                        {
                            "\u4ea7\u54c1\u65cf": "\u4e1c\u65b9\u7504\u9009\u9e21\u6392",
                            "\u4f7f\u7528\u534a\u6210\u54c1\u89c4\u683c": "\u9e21\u5f00\u7247\u5927\u80f8/65-78g/\u81ea\u4fee\u5f62",
                            "\u539f\u6599\u91c7\u8d2d\u5355\u4ef7\u5f71\u54cd": 90.0,
                            "\u4fee\u5f62\u5229\u7528\u7387\u5f71\u54cd": 5.0,
                            "\u635f\u8017\u7387\u5f71\u54cd": 4.0,
                            "\u4fee\u5f62\u4eba\u5de5\u6210\u672c\u5f71\u54cd": 3.0,
                            "\u534a\u6210\u54c1\u603b\u6210\u672c": 102.0,
                            market_col: 70.0,
                            "4\u6708\u4ea7\u91cf": 10.0,
                            "Q4\u6708\u5747\u4ea7\u91cf": 20.0,
                        }
                    ]
                )
                summary = pd.DataFrame(
                    [
                        {
                            "\u5de5\u5382": "BB2",
                            "\u539f\u6599\u91c7\u8d2d\u5355\u4ef7\u5f71\u54cd": 90.0,
                            market_col: 70.0,
                            "\u6263\u9664\u884c\u60c5\u540e\u91c7\u8d2d\u7ee9\u6548": 20.0,
                            "\u4fee\u5f62\u5229\u7528\u7387\u5f71\u54cd": 5.0,
                            "\u635f\u8017\u7387\u5f71\u54cd": 4.0,
                            "\u4fee\u5f62\u4eba\u5de5\u6210\u672c\u5f71\u54cd": 3.0,
                            "\u7efc\u5408\u5f71\u54cd": 32.0,
                            "4\u6708\u4ea7\u91cf": 10.0,
                            "Q4\u6708\u5747\u4ea7\u91cf": 20.0,
                        }
                    ]
                )
                ws = Workbook().active
                ws["AZ50"] = "size marker"

                APP._fill_kind_sheet(ws, detail, grouped, summary, "4", "Q4", "Q4", kind)

                self.assertAlmostEqual(ws["K5"].value, 421.750260152972)
                self.assertAlmostEqual(ws["T3"].value, 70.0)
                self.assertAlmostEqual(ws["AC3"].value, 70.0)
                self.assertAlmostEqual(detail.iloc[0]["\u534a\u6210\u54c1\u603b\u6210\u672c"], 592.657255972947)
                self.assertAlmostEqual(grouped.iloc[0]["\u534a\u6210\u54c1\u603b\u6210\u672c"], 102.0)

    def assertCellFloat(self, ws, coord: str, expected: float):
        self.assertIsNotNone(ws[coord].value, coord)
        self.assertAlmostEqual(ws[coord].value, expected, places=6, msg=coord)

    def test_fill_leg_ratio_sheet_preserves_sorting_and_summary_blocks(self):
        part_df = pd.DataFrame(
            [
                {"工厂": "LY", "原料号": "M3", "原料描述": "d3", "数量": 5, "鲜冻": "鲜品", "规格": " 200-300 "},
                {"工厂": "BB2", "原料号": "M1", "原料描述": "d1", "数量": 10, "鲜冻": "冻品", "规格": None},
                {"工厂": "BB1", "原料号": "M2", "原料描述": "d2", "数量": -2, "鲜冻": "鲜品", "规格": "80g"},
            ]
        )
        month_amounts = {
            "BB2": {"鲜品": 30, "冻品": 70},
            "BB1": {"鲜品": 100, "冻品": 0},
            "LY": {"鲜品": 10, "冻品": 10},
            "DL": {"鲜品": 0, "冻品": 0},
            "合计": {"鲜品": 140, "冻品": 80},
        }
        q_amounts = {
            "BB2": {"鲜品": 40, "冻品": 60},
            "BB1": {"鲜品": 50, "冻品": 50},
            "LY": {"鲜品": 5, "冻品": 15},
            "DL": {"鲜品": 20, "冻品": 0},
            "合计": {"鲜品": 115, "冻品": 125},
        }

        ws = Workbook().active
        ws["U200"] = "stale"

        APP._fill_leg_ratio_sheet(ws, part_df, month_amounts, q_amounts, "3", "Q1")

        self.assertEqual(ws["K1"].value, "3月")
        self.assertEqual(ws["K2"].value, "Q1")
        self.assertEqual(ws["B3"].value, "BB2")
        self.assertEqual(ws["C3"].value, "M1")
        self.assertCellFloat(ws, "E3", 10.0)
        self.assertCellFloat(ws, "F3", 10.0)
        self.assertCellFloat(ws, "G3", 10.0)
        self.assertEqual(ws["H3"].value, "冻品")
        self.assertCellFloat(ws, "I3", 10.0)
        self.assertEqual(ws["J3"].value, "")

        self.assertEqual(ws["B4"].value, "BB1")
        self.assertCellFloat(ws, "E4", -2.0)
        self.assertIsNone(ws["F4"].value)
        self.assertIsNone(ws["I4"].value)
        self.assertEqual(ws["J4"].value, "80g")

        self.assertEqual(ws["B5"].value, "LY")
        self.assertEqual(ws["J5"].value, "200-300")

        self.assertEqual(ws["K4"].value, "BB2")
        self.assertCellFloat(ws, "L4", 70.0)
        self.assertCellFloat(ws, "M4", 30.0)
        self.assertCellFloat(ws, "N4", 100.0)
        self.assertEqual(ws["P4"].value, "BB2")
        self.assertCellFloat(ws, "Q4", 0.7)
        self.assertCellFloat(ws, "R4", 0.3)
        self.assertCellFloat(ws, "S4", 1.0)

        self.assertEqual(ws["K16"].value, "BB2")
        self.assertEqual(ws["L16"].value, "3月")
        self.assertCellFloat(ws, "M16", 0.7)
        self.assertCellFloat(ws, "N16", 0.3)
        self.assertCellFloat(ws, "O16", 1.0)
        self.assertEqual(ws["Q15"].value, "BB2")
        self.assertEqual(ws["R15"].value, "3月")
        self.assertCellFloat(ws, "S15", 0.7)
        self.assertCellFloat(ws, "T15", 0.3)
        self.assertCellFloat(ws, "U15", 1.0)
        self.assertIsNone(ws["U200"].value)

    def test_fill_bre_ratio_sheet_preserves_sorting_and_compare_blocks(self):
        part_df = pd.DataFrame(
            [
                {"工厂": "YZ", "原料号": "M5", "原料描述": "d5", "数量": 3, "鲜冻": "鲜品", "规格": "300g"},
                {"工厂": "BB2", "原料号": "M1", "原料描述": "d1", "数量": 10, "鲜冻": "冻品", "规格": None},
                {"工厂": "TJ", "原料号": "M2", "原料描述": "d2", "数量": -2, "鲜冻": "鲜品", "规格": "220-260g"},
            ]
        )
        month_amounts = {
            "BB2": {"鲜品": 30, "冻品": 70},
            "TJ": {"鲜品": 100, "冻品": 0},
            "YZ": {"鲜品": 10, "冻品": 10},
            "合计": {"鲜品": 140, "冻品": 80},
        }
        q_amounts = {
            "BB2": {"鲜品": 40, "冻品": 60},
            "TJ": {"鲜品": 50, "冻品": 50},
            "YZ": {"鲜品": 5, "冻品": 15},
            "合计": {"鲜品": 115, "冻品": 125},
        }

        ws = Workbook().active
        ws["X250"] = "stale"

        APP._fill_bre_ratio_sheet(ws, part_df, month_amounts, q_amounts, "4", "Q2")

        self.assertEqual(ws["B4"].value, "BB2")
        self.assertEqual(ws["C4"].value, "M1")
        self.assertCellFloat(ws, "E4", 10.0)
        self.assertEqual(ws["F4"].value, "冻品")
        self.assertCellFloat(ws, "G4", 10.0)
        self.assertEqual(ws["H4"].value, "")

        self.assertEqual(ws["B5"].value, "TJ")
        self.assertCellFloat(ws, "E5", -2.0)
        self.assertIsNone(ws["G5"].value)
        self.assertEqual(ws["H5"].value, "220-260g")

        self.assertEqual(ws["I4"].value, "BB2")
        self.assertCellFloat(ws, "J4", 70.0)
        self.assertCellFloat(ws, "K4", 30.0)
        self.assertCellFloat(ws, "L4", 100.0)

        self.assertEqual(ws["N4"].value, "BB2")
        self.assertEqual(ws["O4"].value, "4月")
        self.assertCellFloat(ws, "P4", 0.7)
        self.assertCellFloat(ws, "Q4", 0.3)
        self.assertCellFloat(ws, "R4", 1.0)

        self.assertEqual(ws["T4"].value, "BB2")
        self.assertEqual(ws["U4"].value, "4月")
        self.assertCellFloat(ws, "V4", 0.7)
        self.assertCellFloat(ws, "W4", 0.3)
        self.assertCellFloat(ws, "X4", 1.0)
        self.assertIsNone(ws["X250"].value)

    def test_fill_leg_plant_sheet_preserves_amount_and_diff_sections(self):
        month_spec_amounts = {
            "BB1": {"80g": 3, "200-300": 2},
            "BB2": {"无规格": 1},
            "LY": {"110G": 4},
            "DL": {},
            "合计": {"无规格": 1, "80g": 3, "110G": 4, "200-300": 2},
        }
        q_spec_amounts = {
            "BB1": {"80g": 1, "200-300": 4},
            "BB2": {"无规格": 1, "110G": 1},
            "LY": {"110G": 2},
            "DL": {},
            "合计": {"无规格": 1, "80g": 1, "110G": 3, "200-300": 4},
        }

        ws = Workbook().active
        ws["AD100"] = "stale"

        APP._fill_leg_plant_sheet(ws, month_spec_amounts, q_spec_amounts, "3", "Q1")

        self.assertEqual(ws["L1"].value, "3月")
        self.assertEqual(ws["N1"].value, "Q1")
        self.assertEqual(ws["B3"].value, "BB1")
        self.assertIsNone(ws["C3"].value)
        self.assertCellFloat(ws, "D3", 3.0)
        self.assertCellFloat(ws, "I3", 2.0)
        self.assertCellFloat(ws, "J3", 5.0)

        self.assertEqual(ws["B4"].value, "BB2")
        self.assertCellFloat(ws, "C4", 1.0)
        self.assertCellFloat(ws, "J4", 1.0)

        self.assertEqual(ws["L3"].value, "BB1")
        self.assertEqual(ws["M3"].value, "3月")
        self.assertCellFloat(ws, "N3", 0.0)
        self.assertCellFloat(ws, "O3", 0.6)
        self.assertCellFloat(ws, "T3", 1.0)

        self.assertEqual(ws["M4"].value, "Q1")
        self.assertCellFloat(ws, "O4", 0.2)
        self.assertCellFloat(ws, "T4", 1.0)

        self.assertEqual(ws["M5"].value, "差异")
        self.assertCellFloat(ws, "O5", 0.4)
        self.assertCellFloat(ws, "T5", 0.0)

        self.assertEqual(ws["V3"].value, "BB1")
        self.assertEqual(ws["W3"].value, "3月")
        self.assertCellFloat(ws, "X3", 0.0)
        self.assertCellFloat(ws, "Y3", 0.6)
        self.assertCellFloat(ws, "AD3", 1.0)
        self.assertIsNone(ws["AD100"].value)

    def test_fill_bre_plant_sheet_preserves_amount_and_diff_sections(self):
        month_spec_amounts = {
            "BB1": {"120G以上": 3, "300g": 2},
            "BB2": {"无规格": 1},
            "TJ": {"200g": 4},
            "LY": {},
            "YZ": {"220-300": 2},
            "合计": {"无规格": 1, "120G以上": 3, "200g": 4, "220-300": 2, "300g": 2},
        }
        q_spec_amounts = {
            "BB1": {"120G以上": 1, "300g": 4},
            "BB2": {"无规格": 1, "200g": 1},
            "TJ": {"200g": 2},
            "LY": {},
            "YZ": {"220-300": 1},
            "合计": {"无规格": 1, "120G以上": 1, "200g": 3, "220-300": 1, "300g": 4},
        }

        ws = Workbook().active
        ws["AR100"] = "stale"

        APP._fill_bre_plant_sheet(ws, month_spec_amounts, q_spec_amounts, "4", "Q2")

        self.assertEqual(ws["B3"].value, "BB1")
        self.assertIsNone(ws["C3"].value)
        self.assertCellFloat(ws, "D3", 3.0)
        self.assertCellFloat(ws, "M3", 2.0)
        self.assertCellFloat(ws, "N3", 5.0)

        self.assertEqual(ws["P3"].value, "BB1")
        self.assertEqual(ws["Q3"].value, "4月")
        self.assertCellFloat(ws, "R3", 0.0)
        self.assertCellFloat(ws, "S3", 0.6)
        self.assertCellFloat(ws, "AC3", 1.0)

        self.assertEqual(ws["Q4"].value, "Q2")
        self.assertCellFloat(ws, "S4", 0.2)
        self.assertCellFloat(ws, "AC4", 1.0)

        self.assertEqual(ws["Q5"].value, "差异")
        self.assertCellFloat(ws, "S5", 0.4)
        self.assertCellFloat(ws, "AC5", 0.0)

        self.assertEqual(ws["AE3"].value, "BB1")
        self.assertEqual(ws["AF3"].value, "4月")
        self.assertCellFloat(ws, "AG3", 0.0)
        self.assertCellFloat(ws, "AH3", 0.6)
        self.assertCellFloat(ws, "AR3", 1.0)
        self.assertIsNone(ws["AR100"].value)


if __name__ == "__main__":
    unittest.main()
