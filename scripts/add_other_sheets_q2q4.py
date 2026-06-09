from __future__ import annotations

import argparse
import math
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(r"F:\llqdocument\大成文件\熟食专案")
Q2Q4_DIR = ROOT_DIR / "成本半成品25Q2-Q4"
TARGET_FILES = [
    Q2Q4_DIR / "蚌埠二厂系统成本-Q2-Q4.xlsx",
    Q2Q4_DIR / "蚌埠一厂系统成本-Q2-Q4.xlsx",
    Q2Q4_DIR / "大连系统成本-Q2-Q4.xlsx",
    Q2Q4_DIR / "辽阳系统成本-Q2-Q4.xlsx",
    Q2Q4_DIR / "天津系统成本-Q2-Q4.xlsx",
    Q2Q4_DIR / "兖州系统成本-Q2-Q4.xlsx",
]

THIN = Side(border_style="thin", color="000000")
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def normalize_mat(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".")
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def clean_text(value) -> str:
    return str(value or "").replace("\u3000", "").strip()


def to_float(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def detect_detail_sheet(sheetnames: list[str], suffix: str) -> str:
    matches = []
    for name in sheetnames:
        if not name.endswith(suffix):
            continue
        if name.endswith("TSC"):
            continue
        if "行情" in name:
            continue
        if name in {"半成品", "人工"}:
            continue
        matches.append(name)
    if len(matches) != 1:
        raise ValueError(f"无法唯一识别 {suffix} 工作表: {matches}")
    return matches[0]


def detect_header_row(ws, key: str = "物料号") -> int:
    for r in range(1, min(ws.max_row, 8) + 1):
        headers = [clean_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if key in headers:
            return r
    raise ValueError(f"{ws.title} 未找到表头行")


def build_header_map(ws, header_row: int) -> dict[str, int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        name = clean_text(ws.cell(header_row, c).value)
        if name and name not in headers:
            headers[name] = c
    return headers


def collect_half_materials(ws_values) -> set[str]:
    header_row = detect_header_row(ws_values)
    mats = set()
    for r in range(header_row + 1, ws_values.max_row + 1):
        mat = normalize_mat(ws_values.cell(r, 1).value)
        if mat.startswith("390"):
            mats.add(mat)
    return mats


def collect_detail_materials(ws_values) -> set[str]:
    mats = set()
    for r in range(3, ws_values.max_row + 1):
        first = clean_text(ws_values.cell(r, 1).value)
        if first in {"人工费用", "半成品"}:
            break
        mat = normalize_mat(ws_values.cell(r, 1).value)
        if mat.startswith("390"):
            mats.add(mat)
    return mats


def build_series_map(ws_values) -> dict[str, str]:
    series_map = {}
    for r in range(2, ws_values.max_row + 1):
        semi = normalize_mat(ws_values.cell(r, 4).value)
        series = clean_text(ws_values.cell(r, 9).value)
        if semi and semi not in series_map and series:
            series_map[semi] = series
    return series_map


def build_labor_unit_map(ws_values) -> dict[str, float]:
    header_row = detect_header_row(ws_values)
    col_map = build_header_map(ws_values, header_row)
    mat_col = col_map.get("物料号")
    qty_col = col_map.get("入库数量")
    amt_col = col_map.get("实际金额")
    if not (mat_col and qty_col and amt_col):
        return {}

    sums = defaultdict(lambda: {"qty": 0.0, "amt": 0.0})
    for r in range(header_row + 1, ws_values.max_row + 1):
        mat = normalize_mat(ws_values.cell(r, mat_col).value)
        if not mat.startswith("390"):
            continue
        qty = to_float(ws_values.cell(r, qty_col).value) or 0.0
        amt = to_float(ws_values.cell(r, amt_col).value) or 0.0
        sums[mat]["qty"] += qty
        sums[mat]["amt"] += amt

    result = {}
    for mat, agg in sums.items():
        if agg["qty"]:
            result[mat] = agg["amt"] / agg["qty"]
    return result


def scrap_factor_for_other_desc(desc: str) -> float:
    text = clean_text(desc)
    if "翅中" in text:
        return 0.50
    if "翅根" in text:
        return 0.65
    if "琵琶" in text:
        return 0.65
    if "牛" in text:
        return 0.80
    if "猪" in text:
        return 1.00
    return 1.00


def set_summary_formulas(ws, last_row: int, use_subtotal: bool) -> None:
    if last_row < 3:
        for col in (3, 6, 7, 8, 9, 11, 12, 20):
            if col <= ws.max_column:
                ws.cell(1, col, 0)
        return

    def agg_formula(col_letter: str, divide_by_two: bool = False) -> str:
        if use_subtotal:
            inner = f"SUBTOTAL(9,{col_letter}3:{col_letter}{last_row})"
        else:
            inner = f"SUM({col_letter}3:{col_letter}{last_row})"
        return f"={inner}/2" if divide_by_two else f"={inner}"

    formulas = {
        3: agg_formula("C"),
        6: agg_formula("F"),
        7: agg_formula("G"),
        8: agg_formula("H"),
        9: agg_formula("I", divide_by_two=True),
        11: agg_formula("K", divide_by_two=True),
        12: agg_formula("L", divide_by_two=True),
        20: agg_formula("T", divide_by_two=True),
    }
    for col, formula in formulas.items():
        if col <= ws.max_column:
            ws.cell(1, col, formula)


def rewrite_detail_formulas(ws, start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        if ws.max_column >= 13:
            ws.cell(r, 13, f"=IFERROR(I{r}/J{r},0)")
        if ws.max_column >= 14:
            ws.cell(r, 14, f'=IFERROR(K{r}/I{r},"")')
        if ws.max_column >= 15:
            ws.cell(r, 15, f"=-L{r}/J{r}")
        if ws.max_column >= 16:
            ws.cell(r, 16, f'=IF(D{r}="人工费用",C{r}/J{r},0)')
        if ws.max_column >= 17:
            ws.cell(r, 17, f"=IF(P{r}>0,1-P{r}-O{r},0)")
        if ws.max_column >= 18:
            ws.cell(r, 18, f'=IF(H{r}>0,E{r},"")')
        if ws.max_column >= 19:
            ws.cell(r, 19, f'=IF(H{r}>0,A{r},"")')
        if ws.max_column >= 20:
            ws.cell(r, 20, f"=IF(H{r}>0,H{r},0)")
        if ws.max_column >= 21:
            ws.cell(r, 21, f'=IF(D{r}="人工费用",C{r}/T{r},0)')


def apply_detail_row_styles(ws, start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        raw_no = clean_text(ws.cell(r, 4).value)
        if raw_no == "人工费用":
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = YELLOW_FILL
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = GRID_BORDER
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).border = GRID_BORDER
        ws.cell(2, c).border = GRID_BORDER
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 18


def create_other_detail_sheet(wb, wb_values, other_sheet_name: str, other_materials: set[str], template_detail_name: str) -> int:
    if other_sheet_name in wb.sheetnames:
        wb.remove(wb[other_sheet_name])

    base = wb.copy_worksheet(wb["半成品"])
    base.title = other_sheet_name
    base.delete_rows(1, 1)
    base.cell(2, 2, wb[template_detail_name].cell(2, 2).value)

    for r in range(base.max_row, 2, -1):
        mat = normalize_mat(base.cell(r, 1).value)
        if not mat.startswith("390") or mat not in other_materials:
            base.delete_rows(r, 1)

    use_subtotal = "SUBTOTAL" in str(wb[template_detail_name].cell(1, 3).value or "").upper()
    last_row = base.max_row
    set_summary_formulas(base, last_row, use_subtotal=use_subtotal)
    if last_row >= 3:
        rewrite_detail_formulas(base, 3, last_row)
        apply_detail_row_styles(base, 3, last_row)
    return last_row


def build_other_metrics(ws_values, other_materials: set[str], series_map: dict[str, str], labor_unit_map: dict[str, float]):
    header_row = detect_header_row(ws_values)
    rows_by_mat = defaultdict(list)
    for r in range(header_row + 1, ws_values.max_row + 1):
        mat = normalize_mat(ws_values.cell(r, 1).value)
        if mat in other_materials:
            rows_by_mat[mat].append({
                "物料号": mat,
                "物料描述": clean_text(ws_values.cell(r, 2).value),
                "入库数量": to_float(ws_values.cell(r, 3).value),
                "原料号": normalize_mat(ws_values.cell(r, 4).value),
                "原料描述": clean_text(ws_values.cell(r, 5).value),
                "实际数量": to_float(ws_values.cell(r, 6).value),
                "实际金额": to_float(ws_values.cell(r, 7).value),
                "配方数量": to_float(ws_values.cell(r, 8).value),
                "调整后实际量": to_float(ws_values.cell(r, 9).value),
                "辅助": to_float(ws_values.cell(r, 10).value),
                "调整后实际额": to_float(ws_values.cell(r, 11).value),
                "碎肉量": to_float(ws_values.cell(r, 12).value),
            })

    raw_desc_map = {}
    result = []
    for mat, rows in rows_by_mat.items():
        desc = next((row["物料描述"] for row in rows if row["物料描述"]), "")
        labor_row = next((row for row in rows if row["原料号"] == "人工费用"), None)
        in_qty = None
        aux = None
        adj_amt = None
        scrap_qty = None
        if labor_row:
            in_qty = labor_row["入库数量"]
            aux = labor_row["辅助"] or labor_row["调整后实际量"]
            adj_amt = labor_row["调整后实际额"]
            scrap_qty = labor_row["碎肉量"]
        if in_qty is None:
            in_qty = max((row["入库数量"] or 0.0 for row in rows), default=0.0)
        if aux is None:
            aux = max((row["辅助"] or row["调整后实际量"] or 0.0 for row in rows), default=0.0)
        if adj_amt is None:
            adj_amt = sum((row["调整后实际额"] or 0.0) for row in rows if row["原料号"] != "人工费用")
        if scrap_qty is None:
            scrap_qty = sum((row["碎肉量"] or 0.0) for row in rows)

        month_unit = (adj_amt / aux) if aux else None
        util = (in_qty / aux) if aux else None
        scrap_ratio = (abs(scrap_qty) / aux) if aux else None
        loss = (1 - util - scrap_ratio) if (util is not None and scrap_ratio is not None) else None
        factor = scrap_factor_for_other_desc(desc)
        raw_cost = None
        if month_unit is not None and util not in (None, 0) and loss is not None:
            raw_cost = (month_unit - (1 - util - loss) * month_unit * factor) / util
        labor_unit = labor_unit_map.get(mat)
        total_cost = (raw_cost + labor_unit) if (raw_cost is not None and labor_unit is not None) else None

        raw_qty = defaultdict(float)
        raw_amt = defaultdict(float)
        for row in rows:
            raw_no = row["原料号"]
            qty = row["实际数量"] or 0.0
            amt = row["实际金额"] or 0.0
            if raw_no and raw_no != "人工费用" and qty > 0:
                raw_qty[raw_no] += qty
                raw_amt[raw_no] += amt
                if raw_no not in raw_desc_map and row["原料描述"]:
                    raw_desc_map[raw_no] = row["原料描述"]

        raw_units = {}
        raw_ratios = {}
        for raw_no in sorted(raw_qty):
            qty = raw_qty[raw_no]
            amt = raw_amt[raw_no]
            raw_units[raw_no] = (amt / qty) if qty else None
            raw_ratios[raw_no] = (qty / aux) if aux else None

        result.append({
            "物料号": mat,
            "物料描述": desc,
            "产品族": series_map.get(mat, ""),
            "半成品入库量": (in_qty / 1000.0) if in_qty not in (None, 0) else None,
            "综合单价": month_unit,
            "修形前原料综合耗用单价": month_unit,
            "修形利用率": util,
            "损耗率": loss,
            "半成品原料成本": raw_cost,
            "半成品修形人工成本": labor_unit,
            "半成品总成本": total_cost,
            "raw_units": raw_units,
            "raw_ratios": raw_ratios,
        })

    result.sort(key=lambda item: ((item["半成品入库量"] or 0.0), item["物料号"]), reverse=True)
    return result, raw_desc_map


def find_tsc_labels(template_ws) -> list[str]:
    defaults = ["25年实际单价", "Q2实际单价", "规格占比", "Q2规格占比", "差异", "对半成品成本的影响", "对半成品成本的影响"]
    labels = []
    for idx, row in enumerate(range(4, 11)):
        value = clean_text(template_ws.cell(row, 5).value)
        labels.append(value or defaults[idx])
    return labels


def style_tsc_header(ws, start_col: int, end_col: int) -> None:
    for r in range(1, 4):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = GRID_BORDER


def create_other_tsc_sheet(wb, tsc_sheet_name: str, metrics: list[dict], raw_desc_map: dict[str, str], template_tsc_name: str) -> int:
    if tsc_sheet_name in wb.sheetnames:
        wb.remove(wb[tsc_sheet_name])
    ws = wb.create_sheet(title=tsc_sheet_name)
    template_ws = wb[template_tsc_name]
    row_labels = find_tsc_labels(template_ws)

    raw_codes = sorted({code for item in metrics for code in item["raw_units"].keys()})
    raw_start = 7
    comp_col = raw_start + len(raw_codes)
    cost_start = comp_col + 1
    cost_headers = [
        "修形前原料综合耗用单价",
        "修形利用率",
        "损耗率",
        "半成品原料成本",
        "半成品修形人工成本",
        "半成品总成本",
    ]
    qty_col = cost_start + len(cost_headers)
    bom_col = qty_col + 1
    bom_ratio_col = qty_col + 2
    max_col = bom_ratio_col

    ws.cell(1, 2, "产品族")
    ws.cell(1, 3, "修行后原料")
    ws.cell(1, 4, "使用半成品规格")
    ws.cell(1, 5, "其他")

    if raw_codes:
        ws.merge_cells(start_row=1, start_column=raw_start, end_row=1, end_column=comp_col)
        ws.cell(1, raw_start, "修形前原料耗用")
    else:
        ws.cell(1, comp_col, "修形前原料耗用")

    ws.merge_cells(start_row=1, start_column=cost_start, end_row=1, end_column=cost_start + len(cost_headers) - 1)
    ws.cell(1, cost_start, "修形后成本")
    ws.cell(1, qty_col, "半成品入库量")
    ws.cell(1, bom_col, "BOM")
    ws.cell(1, bom_ratio_col, "BOM占比")

    for idx, raw_code in enumerate(raw_codes, start=raw_start):
        raw_code_value = int(raw_code) if raw_code.isdigit() else raw_code
        ws.cell(2, idx, raw_code_value)
        ws.cell(3, idx, raw_desc_map.get(raw_code, ""))
    ws.cell(3, comp_col, "综合单价")
    for idx, name in enumerate(cost_headers, start=cost_start):
        ws.cell(2, idx, name)

    style_tsc_header(ws, 1, max_col)

    data_row = 4
    unit_fmt = "0.00;[Red](0.00);-"
    qty_fmt = "#,##0;[Red](#,##0);-"
    pct_fmt = "0%;[Red](0%);-"
    pct1_fmt = "0.0%;[Red](0.0%);-"

    for seq, item in enumerate(metrics, start=1):
        base = data_row
        block_rows = list(range(base, base + 7))
        merges = [1, 2, 3, 4, qty_col, bom_col, bom_ratio_col]
        ws.cell(base, 1, seq)
        ws.cell(base, 2, item["产品族"])
        ws.cell(base, 3, int(item["物料号"]) if item["物料号"].isdigit() else item["物料号"])
        ws.cell(base, 4, item["物料描述"])
        if item["半成品入库量"] is not None:
            ws.cell(base, qty_col, item["半成品入库量"])
            ws.cell(base, qty_col).number_format = qty_fmt
        for col in merges:
            ws.merge_cells(start_row=base, start_column=col, end_row=base + 6, end_column=col)

        for offset, row in enumerate(block_rows):
            ws.cell(row, 5, row_labels[offset])
            impact_type = ""
            if offset == 5:
                impact_type = "单位成本"
            elif offset == 6:
                impact_type = "总成本"
            ws.cell(row, 6, impact_type)

            if offset == 0:
                for idx, raw_code in enumerate(raw_codes, start=raw_start):
                    value = item["raw_units"].get(raw_code)
                    if value is not None:
                        ws.cell(row, idx, value)
                        ws.cell(row, idx).number_format = unit_fmt
                if item["综合单价"] is not None:
                    ws.cell(row, comp_col, item["综合单价"])
                    ws.cell(row, comp_col).number_format = unit_fmt
                values = [
                    item["修形前原料综合耗用单价"],
                    item["修形利用率"],
                    item["损耗率"],
                    item["半成品原料成本"],
                    item["半成品修形人工成本"],
                    item["半成品总成本"],
                ]
                for col, value in zip(range(cost_start, cost_start + len(cost_headers)), values):
                    if value is None:
                        continue
                    ws.cell(row, col, value)
                    if col == cost_start + 1:
                        ws.cell(row, col).number_format = pct_fmt
                    elif col == cost_start + 2:
                        ws.cell(row, col).number_format = pct1_fmt
                    else:
                        ws.cell(row, col).number_format = unit_fmt
            elif offset == 2:
                ratio_sum = 0.0
                has_ratio = False
                for idx, raw_code in enumerate(raw_codes, start=raw_start):
                    value = item["raw_ratios"].get(raw_code)
                    if value is not None:
                        ws.cell(row, idx, value)
                        ws.cell(row, idx).number_format = pct_fmt
                        ratio_sum += value
                        has_ratio = True
                if has_ratio:
                    ws.cell(row, comp_col, ratio_sum)
                    ws.cell(row, comp_col).number_format = pct_fmt

        for row in block_rows:
            for c in range(1, max_col + 1):
                cell = ws.cell(row, c)
                if c in {2, 4}:
                    cell.alignment = LEFT
                else:
                    cell.alignment = CENTER
                cell.border = GRID_BORDER
        data_row += 7

    widths = {
        1: 8,
        2: 14,
        3: 14,
        4: 28,
        5: 14,
        6: 12,
        comp_col: 12,
        qty_col: 12,
        bom_col: 10,
        bom_ratio_col: 10,
    }
    for idx in range(raw_start, comp_col):
        widths[idx] = 12
    for idx in range(cost_start, cost_start + len(cost_headers)):
        widths[idx] = 14
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 18
    return ws.max_row


def move_sheet_after(wb, sheet_name: str, after_name: str) -> None:
    if sheet_name not in wb.sheetnames or after_name not in wb.sheetnames:
        return
    sheet = wb[sheet_name]
    sheets = wb._sheets
    sheets.remove(sheet)
    after_idx = wb.sheetnames.index(after_name)
    sheets.insert(after_idx + 1, sheet)


def process_workbook(path: Path, dry_run: bool = False) -> dict:
    wb = load_workbook(path)
    wb_values = load_workbook(path, data_only=True)

    leg_sheet = detect_detail_sheet(wb.sheetnames, "腿肉")
    chest_sheet = detect_detail_sheet(wb.sheetnames, "胸肉")
    prefix = leg_sheet[:-2]
    other_sheet_name = f"{prefix}其他"
    tsc_template_name = "胸肉TSC" if wb["胸肉TSC"].max_row > 1 else "腿肉TSC"

    half_mats = collect_half_materials(wb_values["半成品"])
    leg_mats = collect_detail_materials(wb_values[leg_sheet])
    chest_mats = collect_detail_materials(wb_values[chest_sheet])
    other_mats = half_mats - leg_mats - chest_mats

    last_other_row = create_other_detail_sheet(wb, wb_values, other_sheet_name, other_mats, chest_sheet)
    series_map = build_series_map(wb_values["成品-Q2-4"]) if "成品-Q2-4" in wb_values.sheetnames else {}
    labor_unit_map = build_labor_unit_map(wb_values["人工"])
    metrics, raw_desc_map = build_other_metrics(wb_values["半成品"], other_mats, series_map, labor_unit_map)
    tsc_rows = create_other_tsc_sheet(wb, "其他TSC", metrics, raw_desc_map, tsc_template_name)

    anchor_sheet = "胸肉TSC" if "胸肉TSC" in wb.sheetnames else chest_sheet
    move_sheet_after(wb, other_sheet_name, anchor_sheet)
    move_sheet_after(wb, "其他TSC", other_sheet_name)

    if hasattr(wb, "calculation"):
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
        except Exception:
            pass

    if not dry_run:
        wb.save(path)

    wb.close()
    wb_values.close()
    return {
        "file": path.name,
        "other_materials": len(other_mats),
        "detail_rows": max(last_other_row - 2, 0),
        "tsc_rows": tsc_rows,
    }


def inspect_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    leg_sheet = detect_detail_sheet(wb.sheetnames, "腿肉")
    chest_sheet = detect_detail_sheet(wb.sheetnames, "胸肉")
    prefix = leg_sheet[:-2]
    other_sheet_name = f"{prefix}其他"
    info = {
        "file": path.name,
        "has_other": other_sheet_name in wb.sheetnames,
        "has_other_tsc": "其他TSC" in wb.sheetnames,
        "sheetnames": wb.sheetnames,
    }
    if other_sheet_name in wb.sheetnames:
        info["other_rows"] = wb[other_sheet_name].max_row
    if "其他TSC" in wb.sheetnames:
        info["other_tsc_rows"] = wb["其他TSC"].max_row
    wb.close()
    return info


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inspect", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if args.inspect:
        for path in TARGET_FILES:
            print(inspect_workbook(path))
        return

    for path in TARGET_FILES:
        result = process_workbook(path, dry_run=args.dry_run)
        print(result)


if __name__ == "__main__":
    main()

